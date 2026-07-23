import os
import io
import json
import time
import math
import secrets
import threading
import logging
import requests
from datetime import datetime, timezone, timedelta
from flask import Blueprint, request, redirect, session, jsonify, url_for, Response
from pymongo import MongoClient
import redis as sync_redis
from jinja2.sandbox import SandboxedEnvironment

from data_harvester import get_harvester, RIVER_RACE_TOTAL_PERIODS, RIVER_RACE_MAX_FAME

# Idea #29/#40: shareable player-card PNG + profile QR code. Both are optional at
# import time — if Pillow/qrcode aren't installed yet (see requirements.txt),
# the rest of the site still works; only these two routes 404 with a clear error.
try:
    from PIL import Image, ImageDraw, ImageFont
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False
try:
    import qrcode
    _QRCODE_AVAILABLE = True
except ImportError:
    _QRCODE_AVAILABLE = False

# Idea #26: card-category map for the "mastery by category" progress bars.
# Clash Royale's own API doesn't expose troop/spell/building/champion category on
# a player's card list, so this is a hand-maintained lookup covering the common
# card pool as of mid-2025 — anything not in this dict falls back to "Other"
# rather than crashing, but it will need occasional updates as new cards release.
CARD_CATEGORIES = {
    "Knight": "Troop", "Archers": "Troop", "Goblins": "Troop", "Giant": "Troop",
    "P.E.K.K.A": "Troop", "Minions": "Troop", "Balloon": "Troop", "Witch": "Troop",
    "Skeleton Army": "Troop", "Bomber": "Troop", "Musketeer": "Troop", "Baby Dragon": "Troop",
    "Prince": "Troop", "Wizard": "Troop", "Mini P.E.K.K.A": "Troop", "Spear Goblins": "Troop",
    "Giant Skeleton": "Troop", "Hog Rider": "Troop", "Minion Horde": "Troop", "Ice Wizard": "Troop",
    "Royal Giant": "Troop", "Guards": "Troop", "Princess": "Troop", "Dark Prince": "Troop",
    "Three Musketeers": "Troop", "Lava Hound": "Troop", "Ice Spirit": "Troop", "Fire Spirit": "Troop",
    "Miner": "Troop", "Sparky": "Troop", "Bowler": "Troop", "Lumberjack": "Troop",
    "Battle Ram": "Troop", "Inferno Dragon": "Troop", "Ice Golem": "Troop", "Mega Minion": "Troop",
    "Dart Goblin": "Troop", "Goblin Gang": "Troop", "Electro Wizard": "Troop", "Elite Barbarians": "Troop",
    "Hunter": "Troop", "Executioner": "Troop", "Bandit": "Troop", "Royal Recruits": "Troop",
    "Night Witch": "Troop", "Bats": "Troop", "Royal Ghost": "Troop", "Ram Rider": "Troop",
    "Zappies": "Troop", "Rascals": "Troop", "Cannon Cart": "Troop", "Mega Knight": "Troop",
    "Skeleton Barrel": "Troop", "Flying Machine": "Troop", "Wall Breakers": "Troop", "Royal Hogs": "Troop",
    "Goblin Giant": "Troop", "Fisherman": "Troop", "Magic Archer": "Troop", "Electro Dragon": "Troop",
    "Firecracker": "Troop", "Elixir Golem": "Troop", "Battle Healer": "Troop", "Skeleton Dragons": "Troop",
    "Archer Queen": "Troop", "Golden Knight": "Troop", "Skeleton King": "Troop", "Mighty Miner": "Troop",
    "Monk": "Troop", "Little Prince": "Troop", "Phoenix": "Troop", "Goblin Machine": "Troop",
    "Suspicious Bush": "Troop", "Goblinstein": "Troop", "Boss Bandit": "Troop", "Berserker": "Troop",
    "Void": "Spell", "Fireball": "Spell", "Arrows": "Spell", "Zap": "Spell", "Rocket": "Spell",
    "Goblin Barrel": "Spell", "Freeze": "Spell", "Poison": "Spell", "Lightning": "Spell",
    "Rage": "Spell", "Tornado": "Spell", "Clone": "Spell", "Mirror": "Spell", "Earthquake": "Spell",
    "Barbarian Barrel": "Spell", "Heal Spirit": "Spell", "Giant Snowball": "Spell", "Royal Delivery": "Spell",
    "The Log": "Spell",
    "Cannon": "Building", "Tesla": "Building", "Inferno Tower": "Building", "Bomb Tower": "Building",
    "Mortar": "Building", "X-Bow": "Building", "Elixir Collector": "Building", "Furnace": "Building",
    "Goblin Cage": "Building", "Goblin Hut": "Building", "Barbarian Hut": "Building", "Tombstone": "Building",
    "Goblin Drill": "Building",
    "Golem": "Troop", "Barbarians": "Troop", "Valkyrie": "Troop", "Skeletons": "Troop",
}


def categorize_card(name: str) -> str:
    return CARD_CATEGORIES.get(name, "Other")


# ---------------------------------------------------------------------------
# 250-ideas implementation pass — Section 14: Card meta & deck intelligence
# (216-230). CARD_ELIXIR_COST, WIN_CONDITION_CARDS, and CARD_COUNTERS are all
# hand-maintained lookups, same convention/caveat as CARD_CATEGORIES above:
# non-exhaustive, current as of mid-2025's card pool, and any card missing
# from these dicts degrades gracefully (elixir cost falls back to a neutral
# average, archetype naming falls back to "Mixed Deck", counters return an
# empty list) rather than crashing.
# ---------------------------------------------------------------------------
CARD_ELIXIR_COST = {
    "Skeletons": 1, "Ice Spirit": 1, "Heal Spirit": 1, "Fire Spirit": 1,
    "Goblins": 2, "Spear Goblins": 2, "Bats": 2, "Zap": 2, "The Log": 2,
    "Ice Golem": 2, "Cannon": 3, "Tesla": 4, "Archers": 3, "Knight": 3,
    "Musketeer": 4, "Mini P.E.K.K.A": 4, "Valkyrie": 4, "Hog Rider": 4,
    "Miner": 3, "Goblin Barrel": 3, "Princess": 3, "Ice Wizard": 3,
    "Dart Goblin": 3, "Electro Wizard": 4, "Battle Ram": 4, "Mega Minion": 3,
    "Inferno Dragon": 4, "Bandit": 3, "Royal Ghost": 3, "Fisherman": 3,
    "Magic Archer": 4, "Firecracker": 3, "Wizard": 5, "Witch": 5,
    "Baby Dragon": 4, "Giant": 5, "Royal Giant": 6, "Balloon": 5,
    "Prince": 5, "Dark Prince": 4, "Three Musketeers": 9, "Lava Hound": 7,
    "Sparky": 6, "Bowler": 5, "Lumberjack": 4, "Executioner": 5,
    "Mega Knight": 7, "Golem": 8, "P.E.K.K.A": 7, "Elite Barbarians": 6,
    "Barbarians": 5, "Minion Horde": 5, "Minions": 3, "Giant Skeleton": 6,
    "Skeleton Army": 3, "Guards": 3, "Royal Recruits": 7, "Night Witch": 4,
    "Ram Rider": 5, "Zappies": 4, "Rascals": 5, "Cannon Cart": 5,
    "Skeleton Barrel": 3, "Flying Machine": 4, "Wall Breakers": 2, "Royal Hogs": 5,
    "Goblin Giant": 6, "Electro Dragon": 5, "Elixir Golem": 3, "Battle Healer": 4,
    "Archer Queen": 5, "Golden Knight": 4, "Skeleton King": 4, "Mighty Miner": 4,
    "Monk": 5, "Little Prince": 3, "Phoenix": 4, "Fireball": 4, "Arrows": 3,
    "Rocket": 6, "Freeze": 4, "Poison": 4, "Lightning": 6, "Rage": 2,
    "Tornado": 3, "Clone": 3, "Mirror": 1, "Earthquake": 3, "Barbarian Barrel": 2,
    "Giant Snowball": 2, "Royal Delivery": 3, "Void": 3, "Inferno Tower": 5,
    "Bomb Tower": 4, "Mortar": 4, "X-Bow": 6, "Elixir Collector": 6,
    "Furnace": 4, "Goblin Cage": 4, "Goblin Hut": 5, "Barbarian Hut": 6,
    "Tombstone": 3, "Goblin Drill": 4, "Goblin Gang": 3, "Hunter": 4,
    "Bomber": 3, "Suspicious Bush": 3, "Goblinstein": 5, "Boss Bandit": 5,
    "Berserker": 3, "Goblin Machine": 5, "Skeleton Dragons": 4,
}
DEFAULT_ELIXIR_COST = 4  # neutral fallback for any card missing above

# Idea #222 (lightweight version — not full ML-style deck clustering, which
# 250_IDEAS.md itself flags as needing a much bigger card-role tagging
# system): identify a deck's likely "win condition" card and name the
# archetype after it, the same way players colloquially name decks.
WIN_CONDITION_CARDS = {
    "Hog Rider": "Hog Cycle", "Golem": "Golem Beatdown", "Giant": "Giant Beatdown",
    "Royal Giant": "Royal Giant Beatdown", "Lava Hound": "LavaLoon",
    "Balloon": "Balloon Beatdown", "X-Bow": "X-Bow Siege", "Mortar": "Mortar Siege",
    "Miner": "Miner Control", "Goblin Barrel": "Barrel Cycle", "Graveyard": "Graveyard Control",
    "Three Musketeers": "3M Bridge Spam", "Elite Barbarians": "Elite Barbarians Bridge Spam",
    "Royal Hogs": "Royal Hogs Bridge Spam", "Ram Rider": "Ram Rider Bridge Spam",
    "P.E.K.K.A": "PEKKA Beatdown", "Mega Knight": "Mega Knight Beatdown",
    "Electro Giant": "Electro Giant Beatdown", "Goblin Giant": "Goblin Giant Beatdown",
    "Wall Breakers": "Wall Breakers Cycle", "Skeleton Barrel": "Skeleton Barrel Cycle",
    "Battle Ram": "Battle Ram Bridge Spam",
}

# Idea #216/#219: a small hand-picked counters map (not exhaustive — the CR
# meta is far too large to encode fully by hand, so this covers the most
# commonly-cited counters for well-known cards; anything missing just returns
# an empty counters list rather than a wrong guess).
CARD_COUNTERS = {
    "Hog Rider": ["Cannon", "Tombstone", "Skeletons", "Ice Golem"],
    "Golem": ["Inferno Tower", "Inferno Dragon", "P.E.K.K.A", "Mini P.E.K.K.A"],
    "Balloon": ["Musketeer", "Mega Minion", "Inferno Tower", "Tesla"],
    "Graveyard": ["Skeleton Army", "Guards", "Barbarians"],
    "Sparky": ["Zap", "Rocket", "Skeleton Army", "Barbarians"],
    "Mega Knight": ["Skeleton Army", "Barbarians", "Zap"],
    "Lava Hound": ["Musketeer", "Mega Minion", "Minions", "Baby Dragon"],
    "Royal Giant": ["Inferno Tower", "P.E.K.K.A", "Mini P.E.K.K.A"],
    "Miner": ["Ice Golem", "Skeletons", "Tombstone"],
    "X-Bow": ["Knight", "Archers", "Rocket"],
    "Three Musketeers": ["Fireball", "Poison", "Lightning"],
    "Goblin Barrel": ["Skeleton Army", "Guards", "The Log"],
    "P.E.K.K.A": ["Skeleton Army", "Minion Horde", "Goblin Gang"],
    "Electro Giant": ["Inferno Tower", "Inferno Dragon", "P.E.K.K.A"],
}


def _archetype_name(cards: list) -> str:
    """Idea #222: name a deck signature after its win condition, if we
    recognize one. Falls back to a generic label — see WIN_CONDITION_CARDS'
    docstring note above for why this isn't full clustering."""
    for card in cards:
        if card in WIN_CONDITION_CARDS:
            return WIN_CONDITION_CARDS[card]
    return "Mixed Deck"


def _elixir_curve(cards: list) -> dict:
    """Idea #221: average elixir cost + a simple cost-bucket histogram for a
    deck signature, using CARD_ELIXIR_COST (falls back to a neutral 4 for any
    unrecognized card, same convention as categorize_card)."""
    costs = [CARD_ELIXIR_COST.get(c, DEFAULT_ELIXIR_COST) for c in cards]
    if not costs:
        return {"average": 0, "histogram": {}}
    histogram = {}
    for c in costs:
        histogram[c] = histogram.get(c, 0) + 1
    return {"average": round(sum(costs) / len(costs), 2), "histogram": histogram}

# ---------------------------------------------------------------------------
# 1. WEB SETUP & DATABASE SYNC CONNECTIONS
# ---------------------------------------------------------------------------
web_bp = Blueprint("web", __name__)
log = logging.getLogger("web_routes")

# Idea #188: custom 404/500 error pages matching the site's dark visual
# identity instead of Flask's default plain-text error output. Kept as a
# small inline template (rather than a Mongo-backed get_template() entry)
# since it has to work even when Mongo itself is the thing that's down.
#
# Moved here from mainbot.py (which originally registered these via
# @app.errorhandler on its own Flask app object) so they live on the
# blueprint itself via @web_bp.app_errorhandler instead. Flask applies a
# blueprint's app_errorhandler to whatever app the blueprint gets
# registered on — mainbot.py's real app still gets them exactly as before,
# but so does any other app that registers web_bp, including the sandboxed
# test harnesses used throughout this project's development (which build a
# bare `Flask(__name__)` + `register_blueprint(web_bp)` and nothing else,
# since discord.py isn't installed there). Previously those harnesses could
# only verify the handler by re-executing an extracted copy of the mainbot.py
# code against a throwaway app; the "custom 404 page renders" check in
# run_harness_s11.py has been failing for exactly that reason — it hits a
# real 404 against the harness's own bare app, which never had mainbot.py's
# handler registered on it. Registering here fixes that for real instead of
# further special-casing the test.
_ERROR_PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} | Graveyard Squad</title>
<link rel="icon" href="/static/favicon.svg" type="image/svg+xml">
<link rel="stylesheet" href="/static/theme.css">
<style>
  body {{ background: var(--gy-bg, #0b0c10); color: var(--gy-text, #c5c6c7); font-family: 'Segoe UI', system-ui, sans-serif;
          min-height: 100vh; display: flex; align-items: center; justify-content: center; margin: 0; text-align: center; padding: 24px; }}
  .err-code {{ font-size: 72px; font-weight: 700; color: var(--gy-accent, #00e5ff); text-shadow: 0 0 24px rgba(0,229,255,0.35); margin-bottom: 8px; }}
  .err-skull {{ font-size: 40px; margin-bottom: 12px; }}
  .err-msg {{ font-size: 16px; color: var(--gy-dim, #888); margin-bottom: 24px; }}
  .err-link {{ color: var(--gy-accent, #00e5ff); text-decoration: none; font-weight: 700; }}
  .err-link:hover {{ opacity: 0.8; }}
</style>
</head>
<body>
  <div>
    <div class="err-skull">☠</div>
    <div class="err-code">{code}</div>
    <div class="err-msg">{message}</div>
    <a class="err-link" href="/">← Back to the Roster</a>
  </div>
</body>
</html>"""

@web_bp.app_errorhandler(404)
def _handle_404(e):
    return _ERROR_PAGE.format(title="Not Found", code="404", message="That page doesn't exist — maybe it wandered into the graveyard."), 404

@web_bp.app_errorhandler(500)
def _handle_500(e):
    log.error(f"Unhandled 500 error: {e}")
    return _ERROR_PAGE.format(title="Server Error", code="500", message="Something broke on our end. It's been logged — try again shortly."), 500

CLAN_TAG = os.getenv("CLAN_TAG", "9LVY89UP").strip().upper().replace("#", "")
MAX_CARD_LEVEL = int(os.getenv("MAX_CARD_LEVEL", 15))

# Discord OAuth2 (Authorization Code flow)
DISCORD_CLIENT_ID = os.getenv("DISCORD_CLIENT_ID", "")
DISCORD_CLIENT_SECRET = os.getenv("DISCORD_CLIENT_SECRET", "")
DISCORD_REDIRECT_URI = os.getenv("DISCORD_REDIRECT_URI", "")
DISCORD_API = "https://discord.com/api"
DISCORD_OAUTH_SCOPES = "identify guilds.members.read"
GUILD_ID = os.getenv("GUILD_ID", "")
# Same bot token mainbot.py uses to actually log the bot in (DISCORD_TOKEN,
# checked as a required env var at mainbot.py startup) -- reused here for a
# plain bot-authenticated REST call (GET /guilds/{id}/roles), not a second
# Discord connection. Distinct from DISCORD_CLIENT_ID/SECRET above, which are
# the OAuth *application* credentials used for user login, not the bot itself.
DISCORD_BOT_TOKEN = os.getenv("DISCORD_TOKEN", "")
GUILD_ROLE_NAMES_CACHE_TTL_SECONDS = 3600  # role names/colors change rarely

# Database connections
# Idea #238: connection pooling review. pymongo defaults to maxPoolSize=100 PER
# MongoClient instance, and this codebase constructs FOUR separate MongoClient
# instances across web_routes.py, data_harvester.py, admin.py, and
# harvest_data.py (each with its own independent connection pool rather than
# sharing one client) — worst case that's up to 400 concurrent connections to
# Mongo just from pooling defaults, before counting anything else talking to
# the same DB. Making the pool size explicit and env-tunable here (rather than
# silently relying on the pymongo default) means it can be turned down on a
# small deployment without editing code. admin.py/harvest_data.py look like
# older/legacy entry points reusing the same pattern; if they're confirmed
# dead code they should be removed rather than tuned.
MONGO_MAX_POOL_SIZE = int(os.getenv("MONGO_MAX_POOL_SIZE", "50"))
mongo_client_sync = MongoClient(os.getenv("MONGO_URL", "mongodb://localhost:27017"), maxPoolSize=MONGO_MAX_POOL_SIZE)
db_sync = mongo_client_sync["graveyardbot"]
redis_sync_client = sync_redis.from_url(os.getenv("REDIS_URL", "redis://localhost:6379"))

# API session for efficiency
cr_api_session = requests.Session()
cr_api_session.headers.update({
    "Authorization": f"Bearer {os.getenv('CR_TOKEN', '').strip()}",
    "User-Agent": "GraveyardBot/1.2 (Web Blueprint)",
    "Cache-Control": "max-age=300"
})

sandbox_env = SandboxedEnvironment(autoescape=True)
_cache_lock = threading.Lock()
_HTML_CACHE = {}

# ---------------------------------------------------------------------------
# 2. LEGACY HELPERS & DATA BRIDGES
# ---------------------------------------------------------------------------
def clean_tag(tag: str) -> str:
    return tag.strip().upper().replace("#", "")

def discord_avatar_url(discord_id: str, avatar_hash: str | None) -> str:
    """Builds a real, loadable Discord CDN avatar URL from the two fields
    Discord's own OAuth /users/@me response gives us (idea: "connect the
    discord id with information like their name or picture"). If the user
    never set a custom avatar, `avatar_hash` is None/absent -- Discord's own
    fallback in that case is one of 6 flat-color default avatars, chosen by
    `(discord_id >> 22) % 6` for the modern (post-username-migration)
    numeric-ID scheme. Falls back to index 0 if discord_id isn't a clean
    integer (shouldn't happen for a real Discord snowflake, but this must
    never raise just to render an avatar)."""
    if avatar_hash:
        return f"https://cdn.discordapp.com/avatars/{discord_id}/{avatar_hash}.png?size=64"
    try:
        default_index = (int(discord_id) >> 22) % 6
    except (TypeError, ValueError):
        default_index = 0
    return f"https://cdn.discordapp.com/embed/avatars/{default_index}.png"

def _normalize_card_names(cards) -> list:
    """Bugfix: some older battle_history records store team_cards/
    opponent_cards entries as raw CR API card objects ({"name": ...,
    "level": ...}) instead of plain name strings. harvest_battles() has
    always extracted just the name going forward (see the equivalent
    "Ensure card lists are name strings" normalization already applied in the
    paginated /admin/api/battles and /api/player/<tag>/battles routes), but
    records saved before that convention was established can still have
    dict-shaped entries. Every card-stats function in this file assumes a
    flat list of name strings — using a dict as a dict key raises
    `TypeError: unhashable type: 'dict'`, and sorting a mix of dicts and
    strings raises `TypeError: '<' not supported between instances of 'dict'
    and 'dict'` — which is exactly the live crash this fixes (a specific
    player's deck-breakdown computation on /player/<tag> hit an old
    dict-shaped record and 500'd). Call this right after reading
    team_cards/opponent_cards from Mongo, before doing anything else with it.
    """
    out = []
    for c in (cards or []):
        if not c:
            continue
        if isinstance(c, str):
            out.append(c)
        elif isinstance(c, dict):
            name = c.get("name")
            if name:
                out.append(name)
        else:
            out.append(str(c))
    return out

SENSITIVE_ACTION_MAX_SESSION_AGE = timedelta(hours=4)

def require_recent_login() -> bool:
    """Idea #146: gate the most sensitive admin actions (promote/demote admin,
    flush cache) behind a recently-refreshed Discord session, since this
    project has no separate MFA provider to hook into. Returns False (caller
    should 401) if the session is missing a login timestamp or it's stale —
    stale sessions predate this feature and should just be treated as needing
    a fresh login rather than silently trusted."""
    login_at_raw = session.get("login_at")
    if not login_at_raw:
        return False
    try:
        login_at = datetime.fromisoformat(login_at_raw)
    except (TypeError, ValueError):
        return False
    return (datetime.now(timezone.utc) - login_at) <= SENSITIVE_ACTION_MAX_SESSION_AGE

def _client_ip() -> str:
    return request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()

def rate_limited(bucket: str, max_attempts: int, window_seconds: int) -> bool:
    """Idea #147: lightweight rate limiting for /login and /link POST, since
    neither route had any before. Implemented against Mongo (not an in-memory
    dict) so it holds up correctly even if this app is ever run with multiple
    worker processes. Returns True if the caller is OVER the limit (i.e.
    should be blocked) — check-and-record in one call to keep call sites simple."""
    key = f"ratelimit_{bucket}_{_client_ip()}"
    now = datetime.now(timezone.utc)
    window_start = now - timedelta(seconds=window_seconds)
    doc = db_sync["config"].find_one({"_id": key}) or {}
    attempts = [_as_aware_utc(t) for t in doc.get("attempts", [])]
    attempts = [t for t in attempts if t and t >= window_start]
    if len(attempts) >= max_attempts:
        return True
    attempts.append(now)
    db_sync["config"].update_one({"_id": key}, {"$set": {"attempts": attempts}}, upsert=True)
    return False

def _as_aware_utc(dt):
    """PyMongo returns naive datetimes by default (tz_aware isn't set on this
    project's MongoClient), but every value we write is UTC — so any datetime
    read back from Mongo needs tzinfo re-attached before it can be subtracted
    from datetime.now(timezone.utc) without raising. Bare passthrough for
    already-aware values (e.g. ones we just constructed in this same request)."""
    if dt is None:
        return None
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)

def _record_cr_api_event(status_code: int | None, retried: bool):
    """Idea #156/#157: track CR API health (rate-limit hits, retries) in Mongo
    so the Diagnostics tab and a clanwide degradation banner can both read the
    same picture, regardless of whether the call came from this web process or
    the harvester's separate fetch_api(). Never raises."""
    try:
        update = {"$inc": {"total_calls": 1}, "$set": {"last_call_at": datetime.now(timezone.utc)}}
        if status_code == 429:
            update["$inc"]["rate_limited_count"] = 1
            update["$set"]["last_rate_limited_at"] = datetime.now(timezone.utc)
        if retried:
            update["$inc"]["retry_count"] = 1
        db_sync["config"].update_one({"_id": "cr_api_health"}, update, upsert=True)
    except Exception:
        pass

# Idea #231: cache CR API clan/player responses in Redis with a short TTL —
# the roster page, admin analytics tab, and several player-profile cards can
# each trigger 2-3 fetch_cr_api calls to the SAME endpoint within one page
# render (and across concurrent visitors hitting the same page), so a short
# cache meaningfully cuts redundant live calls without staling data
# noticeably (60s default; CR data itself doesn't change faster than that in
# practice). Same cache-key format is reused by data_harvester.py's
# pre-warming step (idea #232) — see cr_api_cache_key() there.
CR_API_CACHE_TTL_SECONDS = int(os.getenv("CR_API_CACHE_TTL_SECONDS", "60"))

def _cr_api_cache_key(endpoint: str) -> str:
    return f"crapi_cache:{endpoint}"

def fetch_cr_api(endpoint: str, retries: int = 3, use_cache: bool = True) -> dict | None:
    cache_key = _cr_api_cache_key(endpoint)
    if use_cache:
        try:
            cached = redis_sync_client.get(cache_key)
            if cached:
                return json.loads(cached)
        except Exception as e:
            # Redis being unavailable should never take the whole site down —
            # just fall through to a live fetch as if caching were off.
            log.warning(f"CR API cache read failed (non-fatal, fetching live): {e}")

    url = f"https://proxy.royaleapi.dev/v1/{endpoint}"
    for attempt in range(retries):
        try:
            response = cr_api_session.get(url, timeout=10)
            _record_cr_api_event(response.status_code, retried=(attempt > 0))
            if response.status_code == 200:
                data = response.json()
                if use_cache:
                    try:
                        redis_sync_client.setex(cache_key, CR_API_CACHE_TTL_SECONDS, json.dumps(data))
                    except Exception as e:
                        log.warning(f"CR API cache write failed (non-fatal): {e}")
                return data
        except Exception as e:
            log.error(f"Flask API exception: {e}")
            _record_cr_api_event(None, retried=(attempt > 0))
            time.sleep(2 ** attempt)
    return None

def fetch_cr_api_with_fallback(endpoint: str, retries: int = 3) -> tuple[dict | None, bool]:
    """Like fetch_cr_api, but for endpoints that don't already have a dedicated
    Mongo collection to fall back to (the base clan fetch, mainly -- player
    fetches fall back to player_profiles directly, and war/river-race fetches
    fall back to war_tracking/war_history directly, both at their own call
    sites, since those collections already store a fuller equivalent).

    Every successful live call is cached verbatim in config under
    last_known_api::<endpoint>. On a failed live call, that cache is returned
    instead so a page shows last-known real data instead of going blank --
    matching the "last known" pattern already used by
    data_harvester.py's _merge_leaderboard_entry() and the war-fame fallback
    in api_public_leaderboards().

    Returns (data, is_stale). is_stale is True only when the live call failed
    and a cached fallback was used -- callers can surface that to the page
    ("as of <date>") the same way stale Hall of Fame entries do.
    """
    data = fetch_cr_api(endpoint, retries=retries)
    cache_id = f"last_known_api::{endpoint}"
    if data:
        db_sync["config"].update_one(
            {"_id": cache_id},
            {"$set": {"data": data, "cached_at": datetime.now(timezone.utc)}},
            upsert=True,
        )
        return data, False
    cached = db_sync["config"].find_one({"_id": cache_id})
    if cached and cached.get("data"):
        return cached["data"], True
    return None, False

# Idea #236: several Analytics-tab routes (archetypes, tier-list, underused-gems,
# hard-counters, matchup-breakdown, leaderboards) each scan/aggregate 2000-3000
# battle_history docs in Python. Opening the Analytics tab once fires ALL of them
# essentially at the same time (see loadAnalytics() in admin.html), so a single
# tab click was re-running ~7 full scans of the same underlying data. These don't
# need to be second-fresh — the harvester only pulls new battles every
# HARVEST_INTERVAL_MINUTES anyway — so we cache each computed response in Redis
# for a few minutes and serve the cached copy on repeat loads/repeat admins,
# instead of recomputing from scratch every time. Same non-fatal-on-Redis-down
# pattern as the CR API cache above; a `?refresh=1` query param busts the cache
# for anyone who wants a guaranteed-fresh read.
ANALYTICS_CACHE_TTL_SECONDS = int(os.getenv("ANALYTICS_CACHE_TTL_SECONDS", "300"))

def _analytics_cache_key(name: str, *parts) -> str:
    suffix = ":".join(str(p) for p in parts)
    return f"analytics_cache:{name}" + (f":{suffix}" if suffix else "")

def _analytics_cache_get(key: str):
    try:
        cached = redis_sync_client.get(key)
        if cached:
            return json.loads(cached)
    except Exception as e:
        log.warning(f"Analytics cache read failed (non-fatal, computing live): {e}")
    return None

def _analytics_cache_set(key: str, data) -> None:
    try:
        redis_sync_client.setex(key, ANALYTICS_CACHE_TTL_SECONDS, json.dumps(data, default=str))
    except Exception as e:
        log.warning(f"Analytics cache write failed (non-fatal): {e}")

# Idea #239: a "slow query" flag in Diagnostics. Rather than bolting on a real
# APM/profiler, this reuses the same lightweight accumulate-in-a-config-doc
# pattern already established by _record_cr_api_event()'s cr_api_health doc —
# per-route call_count/total_ms/max_ms counters plus a rolling log (last 20) of
# individual calls that crossed the threshold, so admin_diagnostics() can
# surface "these routes are slow" without a dependency on external tooling.
SLOW_QUERY_THRESHOLD_MS = int(os.getenv("SLOW_QUERY_THRESHOLD_MS", "500"))

def _record_query_timing(name: str, duration_ms: float) -> None:
    try:
        update = {
            "$max": {f"routes.{name}.max_ms": round(duration_ms, 1)},
            "$inc": {f"routes.{name}.call_count": 1, f"routes.{name}.total_ms": duration_ms},
        }
        if duration_ms >= SLOW_QUERY_THRESHOLD_MS:
            update["$push"] = {"recent_slow": {"$each": [{
                "route": name,
                "duration_ms": round(duration_ms, 1),
                "at": datetime.now(timezone.utc),
            }], "$slice": -20}}
        db_sync["config"].update_one({"_id": "slow_query_log"}, update, upsert=True)
    except Exception as e:
        log.warning(f"Slow-query logging failed (non-fatal): {e}")

def get_user_guild_roles(access_token: str) -> list:
    """Fetch the caller's role IDs in GUILD_ID using their OAuth access token.
    Ported from admin.py — needed so admin_role_ids-based access (role-granted,
    not just explicit user ID) works the same way the old standalone app did.
    """
    if not GUILD_ID:
        return []
    try:
        r = requests.get(
            f"{DISCORD_API}/users/@me/guilds/{GUILD_ID}/member",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=5,
        )
        if r.status_code == 200:
            return r.json().get("roles", [])
    except Exception as e:
        log.error(f"Failed to fetch guild roles: {e}")
    return []

def fetch_guild_role_names() -> dict:
    """Resolves the server's role IDs to their display names, for showing
    something readable ("War General") instead of a raw snowflake in the
    admin User Access table. Every place that stores a role ID (session
    user_roles, config.system_config.admin_role_ids) has only ever kept IDs
    -- nothing in this codebase previously called Discord's role-list
    endpoint at all. Uses the bot's own token (Bot-authenticated, not a
    per-user OAuth token) since that's a server-wide read, not something
    scoped to whoever's currently logged in. Cached in Redis since role
    names/colors change rarely and this could otherwise fire once per admin
    page load."""
    if not GUILD_ID or not DISCORD_BOT_TOKEN:
        return {}
    cache_key = f"discord_role_names:{GUILD_ID}"
    try:
        cached = redis_sync_client.get(cache_key)
        if cached:
            return json.loads(cached)
    except Exception as e:
        log.warning(f"Role-name cache read failed (non-fatal): {e}")
    try:
        r = requests.get(
            f"{DISCORD_API}/guilds/{GUILD_ID}/roles",
            headers={"Authorization": f"Bot {DISCORD_BOT_TOKEN}"},
            timeout=8,
        )
        if r.status_code != 200:
            return {}
        role_map = {str(role["id"]): role["name"] for role in r.json()}
        redis_sync_client.setex(cache_key, GUILD_ROLE_NAMES_CACHE_TTL_SECONDS, json.dumps(role_map))
        return role_map
    except Exception as e:
        log.warning(f"Failed to fetch guild role names (non-fatal, falling back to raw IDs): {e}")
        return {}

def is_admin() -> bool:
    if "discord_id" not in session: return False
    discord_id = str(session.get("discord_id"))
    master_admin = os.getenv("MASTER_ADMIN_ID", "")
    if master_admin and discord_id == master_admin: return True
    # Check Mongo for authorized admins
    config = db_sync["config"].find_one({"_id": "system_config"}) or {}
    if discord_id in config.get("admin_user_ids", []):
        # Idea #149: session rotation on privilege change. admin_toggle_privilege
        # bumps users.session_version whenever this specific grant/revoke path is
        # used; a session minted before that bump is treated as stale and must
        # re-login, closing the window where a just-revoked admin's existing
        # session cookie would otherwise keep working. Scoped to only this grant
        # path (not master-admin or role-based access) to avoid touching
        # mechanisms this feature isn't about.
        user_doc = db_sync["users"].find_one({"discord_id": discord_id}, {"session_version": 1})
        current_version = (user_doc or {}).get("session_version", 0)
        if session.get("session_version", 0) != current_version:
            return False
        return True
    # Role-based access — ported from admin.py so members granted admin via a
    # Discord role (not just an explicit user ID) aren't locked out.
    allowed_roles = set(str(r) for r in config.get("admin_role_ids", []))
    if allowed_roles:
        user_roles = session.get("user_roles", [])
        if any(str(r) in allowed_roles for r in user_roles):
            return True
    return False

def has_full_admin() -> bool:
    """Idea #65: a lightweight permission tier on top of the existing flat
    is_admin() boolean. Master admin and anyone without an explicit tier
    recorded default to "full" (so this is additive, not a breaking change for
    admins granted before this feature existed) — only discord_ids explicitly
    downgraded to "analytics_only" via /admin/api/users/tier lose access to
    settings/template/user-management routes.
    """
    if not is_admin():
        return False
    discord_id = str(session.get("discord_id"))
    master_admin = os.getenv("MASTER_ADMIN_ID", "")
    if master_admin and discord_id == master_admin:
        return True
    config = db_sync["config"].find_one({"_id": "system_config"}) or {}
    tiers = config.get("admin_tiers", {})
    return tiers.get(discord_id, "full") == "full"


def get_template(template_name: str) -> str:
    with _cache_lock:
        if template_name in _HTML_CACHE: return _HTML_CACHE[template_name]
        doc = db_sync["config"].find_one({"_id": "html_templates"})
        content = doc.get(template_name) if doc else None
        if not content:
            # Nothing deployed to Mongo yet for this template — fall back to the
            # canonical .html file on disk so pages never show "Template Missing"
            # just because nobody has hit Deploy in the UI editor.
            import pathlib
            disk_path = pathlib.Path(__file__).parent / "templates" / f"{template_name}.html"
            if disk_path.exists():
                content = disk_path.read_text(encoding="utf-8")
                log.warning(f"'{template_name}' not found in Mongo html_templates — served from disk fallback.")
            else:
                content = "<h2>Template Missing</h2>"
                log.error(f"'{template_name}' missing from both Mongo and disk ({disk_path}).")
        _HTML_CACHE[template_name] = content
        return content

def get_csrf_token() -> str:
    """One token per browser session, generated lazily and reused for every
    admin page render so the frontend can echo it back on state-changing POSTs."""
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token

@web_bp.before_request
def _csrf_protect():
    # Idea #159 (least-privilege CSRF review): this used to only cover POST to
    # /admin/* — a prior audit explicitly flagged /api/lfg and /link's POST as
    # relying on session-auth alone. Since then, this project has grown many
    # more session-authenticated /api/* POST/DELETE routes (onboarding, shop
    # redeem, notifications, mentor pairs, flair, leaderboard opt-out...), so
    # the gate now covers every state-changing method against /admin/, /api/,
    # and /link — every POST route in this file lives under one of those three
    # prefixes (confirmed by grepping every `methods=[...POST/DELETE...]` route).
    if request.method in ("POST", "DELETE", "PUT") and (
        request.path.startswith("/admin/") or request.path.startswith("/api/") or request.path == "/link"
    ):
        sent = request.headers.get("X-CSRF-Token") or request.form.get("csrf_token", "")
        expected = session.get("csrf_token", "")
        if not expected or not secrets.compare_digest(str(sent), str(expected)):
            return jsonify({"error": "CSRF token missing or invalid. Refresh the page and try again."}), 403

@web_bp.after_request
def _security_headers(response):
    """Idea #153: CSP on the admin panel to reduce XSS blast radius — worth
    calling out that this is defense-in-depth, not a hard guarantee: the UI
    Editor already lets a full admin inject arbitrary HTML/JS into templates
    by design (that's the feature), and 'unsafe-inline' is required below
    because this app's pages use inline <script>/<style> blocks throughout
    rather than external files with nonces. This mainly blocks a DIFFERENT
    class of attack — injected content from an untrusted source other than a
    trusted admin's own deploy (e.g. a compromised CDN script), not a
    malicious admin. Applied to /admin/* specifically per the idea's scope.
    """
    if request.path.startswith("/admin"):
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; "
            # admin.html's <head> links a Google Fonts stylesheet
            # (fonts.googleapis.com) which itself serves @font-face rules
            # pointing at fonts.gstatic.com for the actual woff2 files — both
            # need to be allowlisted or the browser silently blocks the whole
            # font (falls back to the OS default font, plus a noisy CSP
            # console warning on every admin page load).
            "style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com; "
            "img-src 'self' data: https:; "
            "connect-src 'self'; "
            "frame-ancestors 'none'"
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
    return response

def render_sandboxed(template_str: str, **context) -> str:
    template = sandbox_env.from_string(template_str)
    if "session" not in context: context["session"] = session
    return template.render(**context)

# Analytical Bridges (Syncing logic with ClashCog)
def get_player_analytical_data(tag):
    clean = clean_tag(tag)
    player = fetch_cr_api(f"players/%23{clean}")
    is_last_known_data = False
    if not player:
        # player_profiles already stores a near-complete copy of this exact
        # CR API response (the harvester writes the raw {**profile_data, ...}
        # spread on every cycle -- see data_harvester.py) -- fall back to it
        # instead of returning None and 404ing/blanking the whole profile
        # page over a transient live-API failure.
        stored = db_sync["player_profiles"].find_one({"tag": f"#{clean}"})
        if not stored:
            return None
        player = {k: v for k, v in stored.items() if k != "_id"}
        is_last_known_data = True
    # Defensive: exclude _id (ObjectId isn't JSON-serializable) even though
    # this currently only reaches Jinja templates, not jsonify() directly —
    # same bug class as the config_backups crash above, cheap to prevent here
    # before some future response dict starts including recent_battles as-is.
    history = list(db_sync["battle_history"].find({"player_tag": clean}, {"_id": 0}).sort("battle_time", -1).limit(10))
    player["recent_battles"] = history

    # Collection completion — how much of their full card collection is maxed.
    cards = player.get("cards") or []
    player["collection_total_count"] = len(cards)
    player["collection_maxed_count"] = sum(
        1 for c in cards if c.get("level", 0) >= c.get("maxLevel", 999)
    )

    now = datetime.now(timezone.utc)

    # 7-day trophy trend — same baseline logic as the clan-wide "climbers" leaderboard,
    # just scoped to one player so their own page can show a personal "most improved" stat.
    week_ago = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    old_snap_7d = db_sync["player_snapshots"].find_one(
        {"tag": f"#{clean}", "date": {"$gte": week_ago}},
        sort=[("date", 1)],
    )
    player["trophy_trend_7d"] = (player.get("trophies", 0) - old_snap_7d.get("trophies", 0)) if old_snap_7d else None

    # Idea #28 (revised from "on this day a year ago" to a short-term callback):
    # trophies ~1 week ago and ~1 month ago, so the callback is actually useful
    # against a clan that's only been harvesting data for weeks/months, not years.
    month_ago = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    old_snap_30d = db_sync["player_snapshots"].find_one(
        {"tag": f"#{clean}", "date": {"$gte": month_ago}},
        sort=[("date", 1)],
    )
    player["trophies_1_week_ago"] = old_snap_7d.get("trophies") if old_snap_7d else None
    player["trophies_1_month_ago"] = old_snap_30d.get("trophies") if old_snap_30d else None
    player["trophy_trend_30d"] = (player.get("trophies", 0) - old_snap_30d.get("trophies", 0)) if old_snap_30d else None

    # Idea #35: tenure — "Player Since" this clan started tracking them. Uses the
    # earliest player_snapshots row as a proxy for join date (the harvester only
    # starts snapshotting once a player is in the roster), falling back to the
    # `users.linked_at` date if they've linked a Discord account but somehow have
    # no snapshot yet.
    earliest_snap = db_sync["player_snapshots"].find_one({"tag": f"#{clean}"}, sort=[("date", 1)])
    if earliest_snap:
        player["player_since"] = earliest_snap.get("date")
    else:
        user_doc = db_sync["users"].find_one({"cr_tag": f"#{clean}"})
        player["player_since"] = user_doc["linked_at"].strftime("%Y-%m-%d") if user_doc and user_doc.get("linked_at") else None

    # Idea #23: personal bests — highest trophies ever seen in our own snapshot
    # history (separate from CR's own `bestTrophies`, which is season-scoped),
    # longest win streak, and biggest single-battle crown count, all from battle_history.
    all_snaps = list(db_sync["player_snapshots"].find({"tag": f"#{clean}"}, {"trophies": 1}))
    player["personal_best_trophies_tracked"] = max((s.get("trophies", 0) for s in all_snaps), default=player.get("trophies", 0))

    all_battles_chrono = list(
        db_sync["battle_history"].find({"player_tag": clean}, {"result": 1, "team_crowns": 1}).sort("battle_time", 1)
    )
    longest_streak = running = 0
    biggest_crowns = 0
    for b in all_battles_chrono:
        if b.get("result") == "win":
            running += 1
            longest_streak = max(longest_streak, running)
        else:
            running = 0
        biggest_crowns = max(biggest_crowns, b.get("team_crowns", 0) or 0)
    player["personal_best_win_streak"] = longest_streak
    player["personal_best_crowns"] = biggest_crowns

    # Idea #24: deck usage breakdown — which of THEIR OWN decks (8-card signature)
    # they actually play most, mirroring the clan-wide archetype grouping already
    # done in admin_analytics_archetypes, just scoped to one player.
    deck_stats = {}
    for b in db_sync["battle_history"].find({"player_tag": clean}, {"team_cards": 1, "result": 1}):
        cards_played = _normalize_card_names(b.get("team_cards"))
        if len(cards_played) < 8:
            continue
        sig = tuple(sorted(cards_played[:8]))
        entry = deck_stats.setdefault(sig, {"games": 0, "wins": 0})
        entry["games"] += 1
        if b.get("result") == "win":
            entry["wins"] += 1
    player["own_deck_breakdown"] = sorted(
        [
            {"cards": list(sig), "games": v["games"], "win_rate": round(v["wins"] / v["games"] * 100, 1)}
            for sig, v in deck_stats.items()
        ],
        key=lambda x: -x["games"],
    )[:5]

    # Idea #36: deck grade (S/A/B/C) — the player's most-played deck's win rate,
    # graded against the clan-wide archetype win-rate distribution so "62%" means
    # something relative to this clan's own meta rather than an arbitrary cutoff.
    if player["own_deck_breakdown"]:
        top_deck_wr = player["own_deck_breakdown"][0]["win_rate"]
        clan_archetype_wrs = []
        for b in db_sync["battle_history"].find({}, {"team_cards": 1, "result": 1}).limit(3000):
            if b.get("result") not in ("win", "loss"):
                continue
        # Reuse the clan-wide archetype aggregation the admin analytics endpoint
        # already computes, rather than duplicating the full grouping logic here.
        archetypes_resp = _compute_archetypes(min_games=3)
        clan_archetype_wrs = [a["win_rate"] for a in archetypes_resp.get("top_by_usage", [])]
        if clan_archetype_wrs:
            better_than = sum(1 for wr in clan_archetype_wrs if top_deck_wr >= wr)
            percentile = better_than / len(clan_archetype_wrs)
            grade = "S" if percentile >= 0.9 else "A" if percentile >= 0.7 else "B" if percentile >= 0.4 else "C"
        else:
            grade = "—"
        player["deck_grade"] = grade
    else:
        player["deck_grade"] = "—"

    # Idea #38: player-vs-clan-average comparison. Clan average trophies comes
    # from already-harvested player_profiles (no extra live CR API call); clan
    # average win rate reuses the same battle_history aggregation the admin
    # Analytics tab's overview card already does.
    active_profiles = list(db_sync["player_profiles"].find({"left_clan_at": {"$exists": False}}, {"tag": 1, "trophies": 1}))
    if active_profiles:
        player["clan_avg_trophies"] = round(sum(p.get("trophies", 0) for p in active_profiles) / len(active_profiles))
    else:
        player["clan_avg_trophies"] = None

    # Idea #30 (clan-local half only — global rank is a separate admin-gated
    # live-API route below, per the project's new-API-call gating rule): this
    # player's rank among currently-active clan members by trophies, computed
    # entirely from already-harvested Mongo data.
    ranked = sorted(active_profiles, key=lambda p: p.get("trophies", 0), reverse=True)
    player["clan_rank_local"] = next(
        (i + 1 for i, p in enumerate(ranked) if p.get("tag") == f"#{clean}"), None
    )
    clan_wins = db_sync["battle_history"].count_documents({"result": "win"})
    clan_losses = db_sync["battle_history"].count_documents({"result": "loss"})
    player["clan_avg_win_rate"] = round(clan_wins / (clan_wins + clan_losses) * 100, 1) if (clan_wins + clan_losses) else None

    # Idea #37: achievements sorted "closest to completion" first, so a player
    # sees what to chase next instead of scanning a flat unsorted list. Fully
    # completed achievements (or ones with no target) sink to the bottom.
    achievements = player.get("achievements") or []
    def _completion_key(a):
        target = a.get("target")
        value = a.get("value", 0)
        if not target:
            return (2, 0)  # no target — always last
        frac = min(value / target, 1.0)
        if frac >= 1.0:
            return (1, -frac)  # completed — after in-progress, but still grouped together
        return (0, -frac)  # in-progress, closest-first
    player["achievements"] = sorted(achievements, key=_completion_key)

    player["is_last_known_data"] = is_last_known_data
    return player


def _compute_archetypes(min_games: int = 3) -> dict:
    """Shared archetype-grouping logic used by both the admin analytics endpoint
    and the per-player deck-grade calculation (idea #36), so the two never drift
    out of sync by duplicating the grouping rules independently."""
    recent = list(
        db_sync["battle_history"]
        .find({}, {"_id": 0, "team_cards": 1, "result": 1, "player_tag": 1, "battle_time": 1})
        .sort("battle_time", -1)
        .limit(3000)
    )
    archetypes = {}
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        cards = _normalize_card_names(b.get("team_cards"))
        if len(cards) < 8:
            continue
        signature = tuple(sorted(cards[:8]))
        entry = archetypes.setdefault(signature, {"wins": 0, "games": 0})
        entry["games"] += 1
        if result == "win":
            entry["wins"] += 1
    scored = [
        {"cards": list(sig), "games": v["games"], "win_rate": round((v["wins"] / v["games"]) * 100, 1)}
        for sig, v in archetypes.items() if v["games"] >= min_games
    ]
    return {
        "top_by_win_rate": sorted(scored, key=lambda x: (-x["win_rate"], -x["games"]))[:15],
        "top_by_usage": sorted(scored, key=lambda x: -x["games"])[:15],
    }

def get_clan_war_summary():
    latest_war = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)])
    if not latest_war: return {"status": "No War Data"}
    participants = latest_war.get("participants", [])

    # Idea #9: members leadership has already excused (vacation, known absence)
    # or benched (idea #16) shouldn't keep tripping the slacker banner/nudges.
    excused_tags = {
        p.get("tag") for p in db_sync["player_profiles"].find(
            {"war_excused": True}, {"tag": 1}
        )
    }
    now = datetime.now(timezone.utc)
    benched_tags = {
        p.get("tag") for p in db_sync["player_profiles"].find(
            {"benched_until": {"$gte": now}}, {"tag": 1}
        )
    }
    exempt_tags = excused_tags | benched_tags

    slackers = [
        m for m in participants
        if m.get("decksUsed", 0) == 0 and m.get("tag") not in exempt_tags
    ]
    return {"participants_count": len(participants), "slackers": slackers, "exempt_count": len(exempt_tags)}


def compute_projected_finish(war_data: dict) -> dict | None:
    """Idea #4: projected finish, using the requested formula — current fame
    divided by time elapsed so far gives a fame-per-period rate, extrapolated
    across the full RIVER_RACE_TOTAL_PERIODS-day race, expressed as a fraction
    of the RIVER_RACE_MAX_FAME (10,000) a clan can earn. periodIndex is 0-based,
    so we treat "time elapsed" as periodIndex + 1 completed periods.
    """
    if not war_data:
        return None
    clan = war_data.get("clan") or {}
    fame_so_far = clan.get("fame", 0)
    period_index = war_data.get("periodIndex")
    if period_index is None:
        return None
    time_elapsed = period_index + 1  # periods, not hours — a coarse but API-honest unit
    if time_elapsed <= 0 or fame_so_far <= 0:
        return {"projected_fame": fame_so_far, "projected_pct": 0.0, "periods_elapsed": time_elapsed}
    rate_per_period = fame_so_far / time_elapsed
    projected_fame = round(rate_per_period * RIVER_RACE_TOTAL_PERIODS)
    projected_fame = min(projected_fame, RIVER_RACE_MAX_FAME)
    return {
        "projected_fame": projected_fame,
        "projected_pct": round(projected_fame / RIVER_RACE_MAX_FAME * 100, 1),
        "periods_elapsed": time_elapsed,
        "periods_total": RIVER_RACE_TOTAL_PERIODS,
    }

# ---------------------------------------------------------------------------
# 3. DISCORD OAUTH (login / callback / logout)
# ---------------------------------------------------------------------------
@web_bp.route("/login")
def login():
    # Idea #147: rate limit — neither this nor /link POST had any before.
    if rate_limited("login", max_attempts=15, window_seconds=300):
        return "Too many login attempts from this address. Please wait a few minutes and try again.", 429
    state = secrets.token_urlsafe(16)
    session["oauth_state"] = state
    session["post_login_redirect"] = request.args.get("next", "/")
    params = (
        f"client_id={DISCORD_CLIENT_ID}&redirect_uri={requests.utils.quote(DISCORD_REDIRECT_URI, safe='')}"
        f"&response_type=code&scope={DISCORD_OAUTH_SCOPES}&state={state}"
    )
    return redirect(f"{DISCORD_API}/oauth2/authorize?{params}")

@web_bp.route("/callback")
def oauth_callback():
    error = request.args.get("error")
    if error:
        return f"Discord login was cancelled or failed: {error}", 400

    code = request.args.get("code")
    state = request.args.get("state")
    if not code or not state or state != session.get("oauth_state"):
        return "Invalid or expired login attempt. Please try again.", 400

    token_resp = requests.post(
        f"{DISCORD_API}/oauth2/token",
        data={
            "client_id": DISCORD_CLIENT_ID,
            "client_secret": DISCORD_CLIENT_SECRET,
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": DISCORD_REDIRECT_URI,
        },
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=10,
    )
    if token_resp.status_code != 200:
        log.error(f"Discord token exchange failed: {token_resp.text}")
        return "Login failed while exchanging the authorization code.", 400
    access_token = token_resp.json().get("access_token")

    user_resp = requests.get(
        f"{DISCORD_API}/users/@me",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=10,
    )
    if user_resp.status_code != 200:
        return "Login failed while fetching your Discord profile.", 400
    user_data = user_resp.json()

    session["discord_id"] = user_data["id"]
    session["discord_name"] = f"{user_data['username']}"
    session["discord_avatar"] = user_data.get("avatar")
    session["user_roles"] = get_user_guild_roles(access_token)
    session.pop("oauth_state", None)
    # Idea #146: a fresh-Discord-login timestamp, used as a lightweight re-auth
    # prompt gate for the most sensitive admin actions (promote/demote admin,
    # flush cache) — full MFA isn't feasible on top of Discord OAuth alone, but
    # requiring a *recent* login is the same spirit the idea calls for.
    session["login_at"] = datetime.now(timezone.utc).isoformat()
    existing_user = db_sync["users"].find_one({"discord_id": session["discord_id"]}, {"session_version": 1})
    session["session_version"] = (existing_user or {}).get("session_version", 0)

    # Persist name + avatar every login, not just at /link time. Previously
    # session["discord_avatar"] was captured from Discord's own API response
    # (right above) and then discarded the moment the session expired -- it
    # was never written to Mongo, so nothing (not the roster, not the admin
    # User Access table, not the player page) could ever show a Discord
    # picture even though Discord handed it to us for free on every login.
    # Upserting here (rather than only in link_account()) also means this
    # stays fresh even for someone who's logged in but hasn't linked a CR tag
    # yet, and picks up avatar changes on their next login.
    db_sync["users"].update_one(
        {"discord_id": session["discord_id"]},
        {"$set": {
            "discord_name": session["discord_name"],
            "discord_avatar": session["discord_avatar"],
            # Idea (admin panel Discord identity view): this was already being
            # fetched every login for the is_admin() role-based-access check
            # and held only in the session, same story as the avatar above --
            # persisting it means the admin panel can show "what server roles
            # does this person actually have" without a live Discord call.
            "discord_roles": session["user_roles"],
        }},
        upsert=True,
    )

    dest = session.pop("post_login_redirect", "/")
    return redirect(dest)

@web_bp.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ---------------------------------------------------------------------------
# 4. ACCOUNT LINKING (Clash Royale tag <-> Discord ID)
# ---------------------------------------------------------------------------
@web_bp.route("/link", methods=["GET", "POST"])
def link_account():
    if "discord_id" not in session:
        return redirect(url_for("web.login", next="/link"))

    # Idea #130: a clear "already linked" vs "not linked yet" state — this bot
    # has no manual-approval step (linking takes effect immediately after tag
    # verification), so the honest UI is a binary state, not a fake "pending" one.
    existing_link = db_sync["users"].find_one({"discord_id": session["discord_id"]}, {"cr_tag": 1})
    # Idea #200: recruitment video/trailer embed, admin-configured.
    recruit_video_url = (db_sync["config"].find_one({"_id": "bot_settings"}) or {}).get("recruit_video_url", "")
    link_ctx = {
        "name": session.get("discord_name", "Warrior"),
        "already_linked": bool(existing_link and existing_link.get("cr_tag")),
        "linked_tag": (existing_link or {}).get("cr_tag", "").replace("#", ""),
        "csrf_token": get_csrf_token(),
        "recruit_video_url": recruit_video_url,
    }

    if request.method == "GET":
        return render_sandboxed(get_template("link"), **link_ctx)

    # Idea #147: rate limit repeated link-verification attempts (each one is a
    # live CR API call, so this also protects against accidentally hammering
    # the CR API proxy, not just abuse).
    if rate_limited("link_post", max_attempts=10, window_seconds=300):
        return render_sandboxed(get_template("link"), error="Too many attempts — please wait a few minutes and try again.", **link_ctx)

    tag = clean_tag(request.form.get("tag", ""))
    if not tag:
        return render_sandboxed(get_template("link"), error="Please enter a player tag.", **link_ctx)

    player = fetch_cr_api(f"players/%23{tag}")
    if not player:
        return render_sandboxed(get_template("link"), error="Couldn't find that player tag. Double-check it and try again.", **link_ctx)

    db_sync["users"].update_one(
        {"discord_id": session["discord_id"]},
        {"$set": {
            "discord_id": session["discord_id"],
            "discord_name": session.get("discord_name"),
            "cr_tag": f"#{tag}",
            "linked_at": datetime.now(timezone.utc),
        }},
        upsert=True,
    )
    return redirect(f"/player/{tag}")

# ---------------------------------------------------------------------------
# 5. PUBLIC FRONTEND ROUTES
# ---------------------------------------------------------------------------
@web_bp.route("/healthz")
def healthz():
    """Lightweight keep-alive endpoint for uptime pings / Render Cron Jobs.
    Deliberately does NOT call fetch_cr_api() or render any template — a
    keep-alive check has no business making a live CR API call or paying the
    cost of a full Jinja render every ping interval. This exists specifically
    because a cron job was pinging '/' directly instead: that returns the
    entire roster page (which has grown substantially across this project's
    many feature additions), and its response size eventually exceeded
    Render's cron-job output-capture limit, showing as "Failed (output too
    large)" even though the app itself wasn't erroring. Point any uptime/
    keep-alive cron job at this route instead of the homepage.
    """
    return jsonify({"status": "ok"}), 200

@web_bp.route("/")
def index():
    # Falls back to the last successful live fetch (cached in config) instead
    # of rendering an empty roster whenever the CR API is unreachable -- see
    # fetch_cr_api_with_fallback(). clan_data_is_stale drives the "last known
    # as of ..." note in roster.html rather than silently showing 0 members.
    clan_data, clan_data_is_stale = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    if not clan_data:
        log.error("Live clan fetch failed on index() and no cached fallback exists yet — check CR_TOKEN / IP whitelist on this host.")
        clan_data = {"memberList": [], "memberCount": 0}

    # Idea #126: a "what changed since you were last here" banner for members
    # returning after time away. Tracked per-Discord-account via
    # users.last_dashboard_visit — read the OLD value before overwriting it.
    whats_changed = None
    show_onboarding = False
    if "discord_id" in session:
        user_doc = db_sync["users"].find_one({"discord_id": session["discord_id"]})
        last_visit = _as_aware_utc((user_doc or {}).get("last_dashboard_visit"))
        if last_visit and (datetime.now(timezone.utc) - last_visit) > timedelta(days=14):
            recent = list(db_sync["pending_actions"].find(
                {"kind": {"$in": ["changelog_post", "war_summary_post", "milestone_post", "role_change_post"]},
                 "created_at": {"$gte": last_visit}},
                {"_id": 0, "message": 1},
            ).sort("created_at", -1).limit(5))
            if recent:
                whats_changed = [r["message"].split("\n")[0] for r in recent]
        db_sync["users"].update_one(
            {"discord_id": session["discord_id"]},
            {"$set": {"last_dashboard_visit": datetime.now(timezone.utc)}},
            upsert=True,
        )
        # Idea #116: only nudge into the onboarding flow once per account.
        show_onboarding = not (user_doc or {}).get("onboarding_completed", False)

    # "Member Since" column: joined_clan_at is stamped once, the first time our
    # harvester ever sees a tag in the clan (see harvest_clan_and_profiles's
    # $setOnInsert in data_harvester.py) — the closest proxy available, since
    # neither the clan endpoint nor the player endpoint exposes a true
    # clan-join date. For anyone who was already in the clan before this bot
    # started tracking, this shows when tracking began for them, not their
    # real join date — same caveat as "clan tracked since" below.
    join_dates = {
        p["tag"]: p.get("joined_clan_at")
        for p in db_sync["player_profiles"].find({}, {"tag": 1, "joined_clan_at": 1})
    }
    for m in clan_data.get("memberList", []):
        joined_at = _as_aware_utc(join_dates.get(m.get("tag")))
        m["joined_clan_at"] = joined_at.strftime("%Y-%m-%d") if joined_at else None

    # "Clan tracked since": the CR API doesn't expose a true clan-creation
    # date anywhere (the /clans endpoint has no such field), so the best
    # available answer is "the earliest point our own harvester has a record
    # of this clan" — the first clan_snapshots document, which gets one new
    # row every harvest cycle and is never pruned. Explicitly not the clan's
    # real founding date; just how far back this bot's own history goes.
    earliest_snapshot = db_sync["clan_snapshots"].find_one({}, sort=[("timestamp", 1)])
    clan_tracked_since = None
    if earliest_snapshot and earliest_snapshot.get("timestamp"):
        clan_tracked_since = _as_aware_utc(earliest_snapshot["timestamp"]).strftime("%Y-%m-%d")

    # is_admin passed so roster.html can show an Admin Panel link (previously
    # there was no way to reach /admin from the public pages at all).
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    return render_sandboxed(
        get_template("roster"), clan_data=clan_data, is_admin=is_admin(),
        whats_changed=whats_changed, show_onboarding=show_onboarding,
        is_logged_in=("discord_id" in session), csrf_token=get_csrf_token(),
        family_page_enabled=bot_settings.get("family_page_enabled", False),
        beta_features_enabled=bot_settings.get("beta_features_enabled", False),
        recruiting_banner_enabled=bot_settings.get("recruiting_banner_enabled", False),
        member_count=clan_data.get("memberCount", 0),
        clan_data_is_stale=clan_data_is_stale,
        clan_tracked_since=clan_tracked_since,
        # "Our Discord" card: Discord's own official embeddable widget iframe
        # (discord.com/widget?id=...), which the server owner already has —
        # it only renders anything if "Server Widget" is enabled in Discord's
        # own Server Settings, same feature the old /api/discord/widget
        # attempt depended on. GUILD_ID is the same server-wide constant used
        # for role-based admin access elsewhere, kept as the single source of
        # truth rather than hardcoding the guild ID into the template.
        discord_guild_id=GUILD_ID,
        discord_invite_url=bot_settings.get("discord_invite_url", ""),
        discord_widget_style=bot_settings.get("discord_widget_style", "official"),
    )

@web_bp.route("/player/<tag>")
def player_profile(tag):
    player_data = get_player_analytical_data(tag)
    db_player = db_sync["player_profiles"].find_one({"tag": f"#{clean_tag(tag)}"})
    # Idea #118: progressive disclosure — a brand-new member (joined in the last
    # 3 days, per the "player_since" tracking already computed above) gets the
    # deeper analytics cards collapsed by default rather than a wall of stats on
    # day one. Still fully accessible via a click — never actually hidden.
    is_new_member = False
    if player_data and player_data.get("player_since"):
        try:
            since_dt = datetime.strptime(str(player_data["player_since"])[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
            is_new_member = (datetime.now(timezone.utc) - since_dt).days < 3
        except ValueError:
            pass
    # Idea #125: show an assigned mentor pairing on a new member's profile, if one exists.
    mentor_pair = db_sync["mentor_pairs"].find_one({"mentee_tag": f"#{clean_tag(tag)}", "active": True}, {"_id": 0})
    # Idea #132/#138: the notification-preferences card only makes sense on your
    # OWN profile — viewing a teammate's page shouldn't edit your own prefs.
    own_profile = False
    if "discord_id" in session:
        viewer = db_sync["users"].find_one({"discord_id": session["discord_id"]}, {"cr_tag": 1})
        own_profile = bool(viewer and clean_tag(viewer.get("cr_tag", "")) == clean_tag(tag))

    # "Connect the Discord ID with information like their name or picture" —
    # if this player has linked their Discord account (see link_account()),
    # show that identity publicly on their profile the same way "Member
    # Since" already is: name + avatar, both sourced straight from Discord's
    # own OAuth response (captured in oauth_callback(), never surfaced until
    # now). None if this member never linked an account.
    linked_user = db_sync["users"].find_one(
        {"cr_tag": f"#{clean_tag(tag)}"}, {"discord_id": 1, "discord_name": 1, "discord_avatar": 1}
    )
    linked_discord = None
    if linked_user and linked_user.get("discord_id"):
        linked_discord = {
            "name": linked_user.get("discord_name") or "Unknown",
            "avatar_url": discord_avatar_url(linked_user["discord_id"], linked_user.get("discord_avatar")),
        }

    return render_sandboxed(
        get_template("player"), player=player_data, db_player=db_player, is_admin=is_admin(),
        csrf_token=get_csrf_token(), is_new_member=is_new_member, mentor_pair=mentor_pair,
        own_profile=own_profile, linked_discord=linked_discord,
    )

# ---------------------------------------------------------------------------
# 4. ADMIN & MANAGEMENT ROUTES
# ---------------------------------------------------------------------------
@web_bp.route("/admin")
def admin_panel():
    if not is_admin(): return "Unauthorized", 403
    clan_data, clan_data_is_stale = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    db_players = {
        p["tag"].replace("#", ""): p
        for p in db_sync["player_profiles"].find({}, {"tag": 1, "admin_notes": 1, "strikes": 1})
    }
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    system_config = db_sync["config"].find_one({"_id": "system_config"}) or {}
    return render_sandboxed(
        get_template("admin"),
        clan_data=clan_data,
        clan_data_is_stale=clan_data_is_stale,
        db_players=db_players,
        bot_settings=bot_settings,
        system_config=system_config,
        clan_tag=CLAN_TAG,
        csrf_token=get_csrf_token(),
    )


# "Spreadsheet lover" export: human-readable column labels shared across the
# JSON preview table, the CSV, and the XLSX -- previously the raw camelCase/
# snake_case Mongo field names (decksUsedToday, current_streak) went straight
# into every export as-is, which is fine for code but not what someone opening
# this in Excel/Sheets wants to see as a column header.
EXPORT_FIELD_LABELS = {
    "name": "Name", "tag": "Tag", "role": "Role", "trophies": "Trophies",
    "donations": "Donations", "fame": "War Fame",
    "decksUsedToday": "Decks Used Today", "decksRemaining": "Decks Remaining",
    "warDayWins": "War Day Wins", "totalWins": "Total Wins", "totalLosses": "Total Losses",
    "current_streak": "Win Streak",
    "win_rate_pct": "Win Rate %", "war_participation_pct": "War Participation %",
}
# Percentage columns get a real Excel percent number format (see the xlsx
# branch below) rather than being written as a "87.5%" text string, which
# Excel/Sheets can't sort, sum, or average as a number.
EXPORT_PERCENT_LABELS = {"Win Rate %", "War Participation %"}
# Numeric columns worth a bottom TOTALS/AVERAGE row in the xlsx export --
# something a spreadsheet-native leadership review would otherwise have to
# add a formula for themselves every single export.
EXPORT_SUM_LABELS = {"Donations", "War Fame", "Total Wins", "Total Losses", "War Day Wins", "Decks Used Today"}

@web_bp.route("/admin/export/custom", methods=["POST"])
def admin_export_csv():
    """Returns JSON when export_format=json (for the JS preview table), a real
    spreadsheet-native CSV for export_format=csv, or a styled .xlsx for
    export_format=xlsx. All three share the exact same row-building and
    field-labeling logic below, so the preview, the CSV, and the Excel file
    can never silently drift apart from each other the way the old
    client-side-rebuilt CSV could."""
    if not is_admin(): return "Unauthorized", 403
    from flask import Response
    export_format = request.form.get("export_format", "csv")
    requested_fields = request.form.getlist("fields") or ["name", "tag", "role", "trophies", "donations"]
    requested_formulas = set(request.form.getlist("formulas"))

    # Pull latest war participant data to enrich the export
    latest_war = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)]) or {}
    war_participants = {
        p.get("tag", "").replace("#", "").upper(): p
        for p in latest_war.get("clan", {}).get("participants", [])
    }

    clan_data, _ = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    # Default sort: highest trophies first -- a plain "whatever order the CR
    # API happened to return" isn't a useful default for something meant to
    # be opened and read in a spreadsheet.
    members = sorted(clan_data.get("memberList", []), key=lambda m: m.get("trophies", 0), reverse=True)

    selected_keys = list(requested_fields)
    if "win_rate" in requested_formulas and "win_rate_pct" not in selected_keys:
        selected_keys.append("win_rate_pct")
    if "war_participation" in requested_formulas and "war_participation_pct" not in selected_keys:
        selected_keys.append("war_participation_pct")

    records = []
    for m in members:
        tag = m.get("tag", "").replace("#", "").upper()
        wp = war_participants.get(tag, {})
        db_profile = db_sync["player_profiles"].find_one({"tag": f"#{tag}"}) or {}
        total_wins = db_profile.get("wins", 0)
        total_losses = db_profile.get("losses", 0)
        decks_used = wp.get("decksUsedToday", 0)
        decks_remaining = 4 - decks_used
        row = {
            "name": m.get("name", ""),
            "tag": m.get("tag", ""),
            "role": m.get("role", ""),
            "trophies": m.get("trophies", 0),
            "donations": m.get("donations", 0),
            "fame": wp.get("fame", 0),
            "decksUsedToday": decks_used,
            "decksRemaining": decks_remaining,
            "warDayWins": db_profile.get("warDayWins", 0),
            "totalWins": total_wins,
            "totalLosses": total_losses,
            "current_streak": db_profile.get("current_streak", 0),
            # Computed server-side now (previously duplicated in admin.html's
            # JS, and only applied to the client-rebuilt CSV -- never to the
            # JSON preview or the Excel export, so those two never actually
            # showed the formula columns despite the checkboxes implying they
            # would). Stored as a plain number (e.g. 87.5, not "87.5%") so
            # CSV/JSON keep it usable for real spreadsheet math; the xlsx
            # branch below applies a genuine percent number format on top.
            "win_rate_pct": round((total_wins / (total_wins + total_losses)) * 100, 1) if (total_wins + total_losses) else 0.0,
            "war_participation_pct": round((decks_used / (decks_used + decks_remaining)) * 100, 1) if (decks_used + decks_remaining) else 0.0,
        }
        # Only keep requested fields/formulas, relabeled to human-readable headers.
        records.append({EXPORT_FIELD_LABELS.get(k, k): row[k] for k in selected_keys if k in row})

    export_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if export_format == "json":
        return jsonify(records)

    if export_format == "xlsx":
        # Idea #245: a formatted/styled export beyond plain CSV for leadership
        # that lives in spreadsheets — bold header row, frozen header, and
        # already-colored cells (red for members with decks remaining unused
        # this war day / a strike-worthy 0-donation row, green for top
        # donators) instead of leadership manually conditional-formatting a
        # CSV import every time. Extended with an autofilter, real percent
        # number formats, and a bottom totals/average row.
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Clan Roster"
        header_fill = PatternFill(start_color="1E2530", end_color="1E2530", fill_type="solid")
        header_font = Font(color="00E5CC", bold=True)
        slacker_fill = PatternFill(start_color="4A1F1F", end_color="4A1F1F", fill_type="solid")
        top_fill = PatternFill(start_color="1F4A2A", end_color="1F4A2A", fill_type="solid")
        totals_font = Font(bold=True)
        totals_fill = PatternFill(start_color="16191F", end_color="16191F", fill_type="solid")

        fields = list(records[0].keys()) if records else [EXPORT_FIELD_LABELS.get(k, k) for k in selected_keys]
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"

        donation_values = sorted((r.get("Donations", 0) for r in records), reverse=True)
        top_donation_cutoff = donation_values[max(0, min(2, len(donation_values) - 1))] if donation_values else 0

        last_row = 1
        for row_idx, record in enumerate(records, start=2):
            last_row = row_idx
            is_slacker = record.get("Decks Remaining", 0) and record.get("Decks Remaining", 0) > 0
            is_top_donator = "Donations" in record and donation_values and record.get("Donations", 0) >= top_donation_cutoff and top_donation_cutoff > 0
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record.get(field, ""))
                if field in EXPORT_PERCENT_LABELS:
                    # Value is already a plain number like 87.5 -- this format
                    # just appends the "%" glyph for display without Excel's
                    # usual x100 rescale (which is only correct for fractions
                    # like 0.875), so the number keeps sorting/summing sanely.
                    cell.number_format = '0.0"%"'
                if is_slacker:
                    cell.fill = slacker_fill
                elif is_top_donator:
                    cell.fill = top_fill
            ws.row_dimensions[row_idx].height = 16

        # Bottom totals/average row -- sums for count-style columns, an
        # average for percent columns, blank for identity columns (name/tag/
        # role) rather than a meaningless sum of trophies or streaks.
        if records:
            totals_row = last_row + 1
            for col_idx, field in enumerate(fields, start=1):
                if col_idx == 1:
                    cell = ws.cell(row=totals_row, column=col_idx, value="TOTAL / AVG")
                elif field in EXPORT_SUM_LABELS:
                    values = [r.get(field, 0) for r in records if isinstance(r.get(field), (int, float))]
                    cell = ws.cell(row=totals_row, column=col_idx, value=sum(values) if values else "")
                elif field in EXPORT_PERCENT_LABELS:
                    values = [r.get(field, 0) for r in records if isinstance(r.get(field), (int, float))]
                    cell = ws.cell(row=totals_row, column=col_idx, value=round(sum(values) / len(values), 1) if values else "")
                    cell.number_format = '0.0"%"'
                else:
                    cell = ws.cell(row=totals_row, column=col_idx, value="")
                cell.font = totals_font
                cell.fill = totals_fill

        # Autofilter over the header + all data rows (not the totals row --
        # filtering that away with the rest of the data would hide it) so
        # sorting/filtering by column works the moment the file is opened,
        # same as any hand-built spreadsheet.
        if records:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{last_row}"

        for col_idx, field in enumerate(fields, start=1):
            width = max(12, min(30, len(field) + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=graveyard_roster_{export_date}.xlsx"},
        )

    import csv, io
    buf = io.StringIO()
    if records:
        writer = csv.DictWriter(buf, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)
    return Response(
        # A leading UTF-8 BOM ("﻿") so Excel (which otherwise guesses the
        # system codepage instead of UTF-8) renders accented/non-ASCII player
        # names correctly instead of showing mojibake -- Google Sheets/Numbers
        # handle plain UTF-8 fine either way, so this is purely for Excel's
        # benefit and is invisible everywhere else. No stray title line before
        # the header row anymore either (the old client-side CSV builder
        # prepended one) -- row 1 is always the real header, so Excel/Sheets'
        # own "format as table"/autofilter detection works immediately.
        "﻿" + buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=graveyard_roster_{export_date}.csv"},
    )

@web_bp.route("/admin/api/player/update", methods=["POST"])
def admin_player_update():
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Expected a JSON body."}), 400
    db_sync["player_profiles"].update_one({"tag": f"#{clean_tag(data.get('tag'))}"}, {"$set": {"admin_notes": data.get("notes")}}, upsert=True)
    log_admin_activity("Saved admin notes", target=clean_tag(data.get("tag")))
    return jsonify({"success": True})

@web_bp.route("/admin/api/player/strike", methods=["POST"])
def admin_add_strike():
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Expected a JSON body."}), 400
    db_sync["player_profiles"].update_one({"tag": f"#{clean_tag(data.get('tag'))}"}, {"$inc": {"strikes": 1}}, upsert=True)
    log_admin_activity("Issued strike", target=clean_tag(data.get("tag")))
    notify_user_by_tag(data.get("tag"), "⚠️ A strike was added to your record.", kind="strike")
    return jsonify({"success": True})

@web_bp.route("/admin/api/player/admin_toggle", methods=["POST"])
def admin_toggle_privilege():
    # Idea #146: this is one of the two most sensitive admin actions (the other
    # being cache flush) — require a recently-refreshed Discord login rather
    # than trusting an arbitrarily old session cookie.
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    if not require_recent_login():
        return jsonify({"error": "reauth_required", "message": "Please log in again to confirm this sensitive action."}), 401
    data = request.json
    tag = clean_tag(data.get("tag"))
    # is_admin() checks discord_id against admin_user_ids, so we need this player's
    # linked Discord ID, not their in-game tag, for the grant to actually take effect.
    user = db_sync["users"].find_one({"cr_tag": f"#{tag}"})
    if not user or not user.get("discord_id"):
        return jsonify({"error": "This player hasn't linked a Discord account yet — link them first."}), 400
    if data.get("is_admin"):
        db_sync["config"].update_one({"_id": "system_config"}, {"$addToSet": {"admin_user_ids": user["discord_id"]}}, upsert=True)
    else:
        db_sync["config"].update_one({"_id": "system_config"}, {"$pull": {"admin_user_ids": user["discord_id"]}})
    # Idea #149: invalidate any existing session this user has open, forcing a
    # fresh login before the (revoked or newly-granted) privilege takes effect.
    db_sync["users"].update_one({"discord_id": user["discord_id"]}, {"$inc": {"session_version": 1}})
    log_admin_activity("Toggled admin privilege", target=tag, details=f"is_admin={data.get('is_admin')}")
    notify_user_by_tag(
        tag, "🎉 You were granted admin access." if data.get("is_admin") else "Your admin access was removed.",
        kind="promotion",
    )
    return jsonify({"success": True})

@web_bp.route("/admin/flush-cache", methods=["POST"])
def admin_flush_cache():
    if not has_full_admin(): return jsonify({"error": "unauthorized — flushing cache requires full-admin tier"}), 403
    if not require_recent_login():
        return jsonify({"error": "reauth_required", "message": "Please log in again to confirm this sensitive action."}), 401
    redis_sync_client.flushall()
    _HTML_CACHE.clear()
    log_admin_activity("Flushed cache")
    return jsonify({"message": "Cache flushed."})

@web_bp.route("/admin/api/cleanup/battle-history-nulls", methods=["POST"])
def admin_cleanup_battle_history_nulls():
    """Fixes the recurring boot-time error: 'Failed to create index unique_battle_id
    ... E11000 duplicate key ... unique_battle_id: null'. That index can't build
    while multiple battle_history docs have a null/missing unique_battle_id (a
    unique index rejects duplicate nulls same as any other duplicate value).
    Deletes those specific bad records, then immediately re-runs the harvester's
    _ensure_indexes() so the index gets built right away rather than needing a
    full redeploy/restart to notice it succeeded. Same sensitive-action gate as
    /admin/flush-cache above since this deletes data (nothing that isn't already
    corrupt/unusable — a null unique_battle_id can never be queried by battle_time
    for that record's own player anyway — but still a delete, so full-admin +
    fresh-login only)."""
    if not has_full_admin(): return jsonify({"error": "unauthorized — this requires full-admin tier"}), 403
    if not require_recent_login():
        return jsonify({"error": "reauth_required", "message": "Please log in again to confirm this sensitive action."}), 401

    result = db_sync["battle_history"].delete_many({"unique_battle_id": None})

    get_harvester()._ensure_indexes()
    index_now_exists = "unique_battle_id_1" in db_sync["battle_history"].index_information()

    log_admin_activity("Cleaned up battle_history null unique_battle_id records",
                        details=f"deleted={result.deleted_count}, index_rebuilt={index_now_exists}")
    return jsonify({
        "success": True,
        "deleted_count": result.deleted_count,
        "index_rebuilt": index_now_exists,
        "message": (
            f"Deleted {result.deleted_count} bad record(s). "
            + ("The unique_battle_id index is now active." if index_now_exists
               else "Index still didn't build — there may be non-null duplicates too; check the server log for details.")
        ),
    })

@web_bp.route("/admin/api/player/link", methods=["POST"])
def admin_manual_link():
    """Manually associate a Discord ID with a player tag (admin.html's manualLinkDiscord)."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.json or {}
    tag = clean_tag(data.get("tag", ""))
    discord_id = str(data.get("discord_id", "")).strip()
    # Idea #154: real Discord snowflake IDs are 17-20 digits — bounding the
    # length (beyond the pre-existing bare isdigit() check) rejects obviously
    # malformed input (e.g. a pasted phone number or a 40-digit string) before
    # it ever reaches Mongo, and also bounds the tag length defensively.
    if not tag or len(tag) > 20 or not discord_id.isdigit() or not (17 <= len(discord_id) <= 20):
        return jsonify({"error": "A valid tag and a 17-20 digit numeric Discord ID are required."}), 400
    db_sync["users"].update_one(
        {"discord_id": discord_id},
        {"$set": {"discord_id": discord_id, "cr_tag": f"#{tag}", "linked_at": datetime.now(timezone.utc), "linked_by": "admin"}},
        upsert=True,
    )
    log_admin_activity("Manually linked Discord account", target=tag, details=f"discord_id={discord_id}")
    return jsonify({"success": True})

@web_bp.route("/admin/api/player/dm-warning", methods=["POST"])
def admin_dm_warning():
    """Queue a single-player DM warning; clash_cog.py's pending-actions loop delivers it."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.json or {}
    tag = clean_tag(data.get("tag", ""))
    user = db_sync["users"].find_one({"cr_tag": f"#{tag}"})
    if not user:
        return jsonify({"error": "That player hasn't linked a Discord account yet."}), 400
    db_sync["pending_actions"].insert_one({
        "kind": "dm_warning",
        "discord_id": user["discord_id"],
        "message": data.get("message", "You're missing war participation — please use your decks!"),
        "created_at": datetime.now(timezone.utc),
        "processed": False,
    })
    log_admin_activity("Sent DM warning", target=tag)
    return jsonify({"success": True, "message": "DM queued for delivery."})

@web_bp.route("/admin/war/nudges", methods=["POST"])
def admin_war_nudges():
    """Queue DM reminders to every linked member who still has war decks left today."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    summary = get_clan_war_summary()
    slacker_tags = {s.get("tag") for s in summary.get("slackers", []) if s.get("tag")}
    if not slacker_tags:
        return jsonify({"message": "No slackers found — nothing to send."})

    linked_users = list(db_sync["users"].find({"cr_tag": {"$in": list(slacker_tags)}}))
    discord_ids = [u["discord_id"] for u in linked_users if u.get("discord_id")]
    if not discord_ids:
        return jsonify({"message": "Slackers found, but none have linked a Discord account yet."})

    db_sync["pending_actions"].insert_one({
        "kind": "war_nudge",
        "discord_ids": discord_ids,
        "created_at": datetime.now(timezone.utc),
        "processed": False,
    })
    return jsonify({"success": True, "message": f"Nudges queued for {len(discord_ids)} member(s)."})

@web_bp.route("/admin/api/settings/save", methods=["POST"])
def admin_save_settings():
    """Persist bot settings (admin.html's Settings tab); the bot reloads these every 30s.
    Covers everything admin.py used to save across its two forms (General Bot
    Customizations + Live System File Configurations), consolidated into one endpoint
    and one Settings tab instead of two disconnected config docs.
    """
    if not has_full_admin(): return jsonify({"error": "unauthorized — settings require full-admin tier"}), 403
    data = request.json or {}

    # -- bot_settings doc: runtime bot behavior --------------------------------
    new_maintenance_mode = bool(data.get("maintenance_mode", False))
    update = {
        "maintenance_mode": new_maintenance_mode,
        "feature_auto_pings": bool(data.get("feature_auto_pings", False)),
        "war_channel_id": data.get("war_channel_id", 0),
    }
    # Idea #70: track when maintenance mode was turned ON so a stale-warning
    # banner can flag it if it's been left on too long.
    prev_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    if new_maintenance_mode and not prev_settings.get("maintenance_mode"):
        update["maintenance_enabled_at"] = datetime.now(timezone.utc)
    elif not new_maintenance_mode:
        update["maintenance_enabled_at"] = None
    if "command_prefix" in data:
        prefix = str(data.get("command_prefix") or "!").strip()[:3]
        update["command_prefix"] = prefix or "!"
    if "ignored_channels" in data:
        raw = data.get("ignored_channels", [])
        if isinstance(raw, str):
            raw = [c.strip() for c in raw.split(",") if c.strip()]
        update["ignored_channels"] = [str(c).strip() for c in raw if str(c).strip()]
    # Ported from admin.py's old /admin/save-config ("General Bot Customizations") —
    # these three lived in a separate, disconnected doc before; folded in here so
    # there's one Settings tab and one save button instead of two.
    if "min_trophies" in data:
        try:
            update["min_trophies"] = int(data.get("min_trophies") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "min_trophies must be a number"}), 400
    if "war_reminders" in data:
        update["war_reminders"] = bool(data.get("war_reminders", False))
    if "welcome_msg" in data:
        update["welcome_msg"] = str(data.get("welcome_msg") or "Welcome to the Squad!").strip()[:500]
    # Idea #5: configurable under-strength threshold.
    if "min_clan_size" in data:
        try:
            update["min_clan_size"] = max(0, int(data.get("min_clan_size") or 40))
        except (TypeError, ValueError):
            return jsonify({"error": "min_clan_size must be a number"}), 400
    # Idea #13: the tiered war-reminder day-progress approximation needs to know
    # roughly when the war day resets, since the CR API doesn't expose it directly.
    if "war_reset_hour_utc" in data:
        try:
            update["war_reset_hour_utc"] = max(0, min(23, int(data.get("war_reset_hour_utc") or 10)))
        except (TypeError, ValueError):
            return jsonify({"error": "war_reset_hour_utc must be 0-23"}), 400
    # Idea #82/#83: the beta nickname/role sync is manual-only (/syncbeta slash
    # command) — this just tells that command which role counts as "beta test group".
    if "beta_sync_role_id" in data:
        update["beta_sync_role_id"] = str(data.get("beta_sync_role_id") or "").strip() or None
    # Idea #86: off (0) by default — auto-strike only kicks in once an admin
    # explicitly sets a consecutive-missed-war threshold.
    if "auto_strike_missed_war_threshold" in data:
        try:
            update["auto_strike_missed_war_threshold"] = max(0, int(data.get("auto_strike_missed_war_threshold") or 0))
        except (TypeError, ValueError):
            return jsonify({"error": "auto_strike_missed_war_threshold must be a number"}), 400
    # Idea #98: which channel gets the auto-posted changelog line when settings/templates change.
    if "changelog_channel_id" in data:
        update["changelog_channel_id"] = data.get("changelog_channel_id") or None
    # Idea #134: dedicated announcements channel convention for all automated posts.
    if "announcements_channel_id" in data:
        update["announcements_channel_id"] = data.get("announcements_channel_id") or None
    # Idea #131/#139: leadership escalation channel.
    if "leadership_channel_id" in data:
        update["leadership_channel_id"] = data.get("leadership_channel_id") or None
    # Idea #99: quiet hours (UTC) during which automated DMs are held, not dropped.
    if "quiet_hours_start" in data or "quiet_hours_end" in data:
        try:
            start = data.get("quiet_hours_start")
            end = data.get("quiet_hours_end")
            update["quiet_hours_start"] = int(start) if start not in (None, "") else None
            update["quiet_hours_end"] = int(end) if end not in (None, "") else None
        except (TypeError, ValueError):
            return jsonify({"error": "quiet_hours_start/end must be numbers 0-23"}), 400

    # ---- Section 12/13/16 feature-flag toggles (per your review notes, all
    # default OFF/hidden until an admin explicitly enables them here) --------
    # Idea #198: sub-clan/family page — dark until the clan actually grows.
    if "family_page_enabled" in data:
        update["family_page_enabled"] = bool(data.get("family_page_enabled"))
    if "family_clan_tags" in data:
        raw = data.get("family_clan_tags", [])
        if isinstance(raw, str):
            raw = [c.strip() for c in raw.split(",") if c.strip()]
        update["family_clan_tags"] = [clean_tag(c) for c in raw if clean_tag(c)]
    # Idea #205: testimonial carousel — dark until enabled.
    if "testimonials_enabled" in data:
        update["testimonials_enabled"] = bool(data.get("testimonials_enabled"))
    # Idea #200: recruitment video/trailer embed.
    if "recruit_video_url" in data:
        update["recruit_video_url"] = str(data.get("recruit_video_url") or "").strip()[:500]
    # "A little applet for the roster page showing our discord" — shown as an
    # explicit "Join our Discord" button under the embedded server widget
    # iframe (see index() / roster.html), since nothing in this codebase
    # stored an invite URL anywhere before now.
    if "discord_invite_url" in data:
        update["discord_invite_url"] = str(data.get("discord_invite_url") or "").strip()[:300]
    # Lets an admin pick between Discord's own official embeddable widget
    # (Discord's UI/chrome, zero maintenance on our end) and the custom-built
    # version that matches this site's own dark theme exactly (see
    # /api/discord/widget). Whitelisted to the two known values rather than
    # storing whatever string was posted, since roster.html branches on this
    # exact value.
    if "discord_widget_style" in data:
        style = str(data.get("discord_widget_style") or "official").strip().lower()
        update["discord_widget_style"] = style if style in ("official", "custom") else "official"
    # Idea #203: clan founding date, used for the auto-generated anniversary page/post.
    if "clan_founding_date" in data:
        update["clan_founding_date"] = str(data.get("clan_founding_date") or "").strip()[:10]
    # Idea #204: clan culture doc, shown on /clan-history.
    if "culture_page_content" in data:
        update["culture_page_content"] = str(data.get("culture_page_content") or "").strip()[:4000]
    # Idea #215: public "we're recruiting" banner toggle.
    if "recruiting_banner_enabled" in data:
        update["recruiting_banner_enabled"] = bool(data.get("recruiting_banner_enabled"))
    # Idea #214: auto-decline stale applications — off by default, admin turns on.
    if "auto_decline_stale_applications_enabled" in data:
        update["auto_decline_stale_applications_enabled"] = bool(data.get("auto_decline_stale_applications_enabled"))
    if "auto_decline_days" in data:
        try:
            update["auto_decline_days"] = max(1, int(data.get("auto_decline_days") or 14))
        except (TypeError, ValueError):
            return jsonify({"error": "auto_decline_days must be a number"}), 400
    # Idea #243: Twitch/YouTube go-live integration — dark until enabled (also
    # requires real Twitch/YouTube API credentials to actually poll, which
    # this project doesn't have configured yet; see clash_cog.py note).
    if "streaming_integration_enabled" in data:
        update["streaming_integration_enabled"] = bool(data.get("streaming_integration_enabled"))
    if "streaming_twitch_channel" in data:
        update["streaming_twitch_channel"] = (data.get("streaming_twitch_channel") or "").strip()
    if "streaming_youtube_channel" in data:
        update["streaming_youtube_channel"] = (data.get("streaming_youtube_channel") or "").strip()
    # Idea #247: Reddit recruitment auto-poster — same reasoning as streaming
    # above (needs Reddit API credentials this project doesn't have configured).
    if "reddit_autopost_enabled" in data:
        update["reddit_autopost_enabled"] = bool(data.get("reddit_autopost_enabled"))
    if "reddit_subreddit" in data:
        update["reddit_subreddit"] = (data.get("reddit_subreddit") or "").strip()
    # Idea #250: opt-in beta-features toggle, so leadership can trial
    # upcoming features (this pass gates the poll/gallery board behind it as
    # the concrete example) before rolling them out clan-wide.
    if "beta_features_enabled" in data:
        update["beta_features_enabled"] = bool(data.get("beta_features_enabled"))

    db_sync["config"].update_one(
        {"_id": "bot_settings"},
        {"$set": update},
        upsert=True,
    )
    log_admin_activity("Saved bot settings", details=", ".join(f"{k}={v}" for k, v in update.items()))

    # -- system_config doc: access control -------------------------------------
    # admin_role_ids lives in the same doc is_admin() already reads for admin_user_ids,
    # so role-based access changes take effect immediately without a separate lookup.
    if "admin_role_ids" in data:
        raw = data.get("admin_role_ids", [])
        if isinstance(raw, str):
            raw = [r.strip() for r in raw.split(",") if r.strip()]
        role_ids = [str(r).strip() for r in raw if str(r).strip()]
        db_sync["config"].update_one(
            {"_id": "system_config"},
            {"$set": {"admin_role_ids": role_ids}},
            upsert=True,
        )

    return jsonify({"success": True})

@web_bp.route("/admin/api/harvest/trigger", methods=["POST"])
def admin_trigger_harvest():
    """Manually kick a full harvest cycle (admin.html's triggerHarvester) without blocking the request."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    harvester = get_harvester()
    threading.Thread(target=harvester.run_full_cycle, daemon=True).start()
    return jsonify({"success": True, "message": "Harvest started in the background."})

# ---------------------------------------------------------------------------
# 5. NEW ADMIN API ROUTES (required by the new admin dashboard)
# ---------------------------------------------------------------------------

@web_bp.route("/admin/diagnostics")
def admin_diagnostics():
    """Full health-check JSON — powers the Diagnostics tab and raw JSON viewer."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    import socket, time as _time

    result = {
        "version": "1.3",
        "environment": os.getenv("FLASK_ENV", "production"),
        "hostname": socket.gethostname(),
    }

    # Redis
    t0 = _time.monotonic()
    try:
        info = redis_sync_client.info("memory")
        ping_ms = round((_time.monotonic() - t0) * 1000, 1)
        result["redis"] = {
            "status": "ok",
            "ping_ms": ping_ms,
            "used_memory": info.get("used_memory_human", "N/A"),
            "total_keys": redis_sync_client.dbsize(),
        }
    except Exception as e:
        result["redis"] = {"status": "error", "error": str(e)}

    # MongoDB
    t0 = _time.monotonic()
    try:
        mongo_client_sync.admin.command("ping")
        ping_ms = round((_time.monotonic() - t0) * 1000, 1)
        result["mongo"] = {
            "status": "ok",
            "ping_ms": ping_ms,
            "snapshot_count": db_sync["player_snapshots"].estimated_document_count(),
            "battle_count": db_sync["battle_history"].estimated_document_count(),
        }
    except Exception as e:
        result["mongo"] = {"status": "error", "error": str(e)}

    # CR API
    t0 = _time.monotonic()
    try:
        test_url = f"https://proxy.royaleapi.dev/v1/clans/%23{CLAN_TAG}"
        resp = cr_api_session.get(test_url, timeout=8)
        latency_ms = round((_time.monotonic() - t0) * 1000, 1)
        result["cr_api"] = {
            "status": "ok" if resp.status_code == 200 else ("rate_limited" if resp.status_code == 429 else "error"),
            "status_code": resp.status_code,
            "latency_ms": latency_ms,
            "endpoint_tested": f"/clans/%23{CLAN_TAG}",
        }
    except Exception as e:
        result["cr_api"] = {"status": "error", "error": str(e)}

    # Idea #157: cumulative retry/rate-limit visibility, not just this single
    # test call above — tracked across every real fetch_cr_api() call by
    # _record_cr_api_event().
    cr_health = db_sync["config"].find_one({"_id": "cr_api_health"}, {"_id": 0}) or {}
    result["cr_api"]["cumulative"] = cr_health

    # Bot (written by the bot process into Mongo every 30s via reload_config_loop)
    bot_heartbeat = db_sync["config"].find_one({"_id": "bot_heartbeat"}) or {}
    result["bot"] = {
        "connected": bot_heartbeat.get("connected", False),
        "latency_ms": bot_heartbeat.get("latency_ms"),
        "uptime": bot_heartbeat.get("uptime"),
    }

    # Cache
    result["cache"] = {
        "backend": "redis" if result.get("redis", {}).get("status") == "ok" else "mongo",
        "total_keys": result.get("redis", {}).get("total_keys", 0),
        "html_cache_entries": len(_HTML_CACHE),
    }

    # Idea #239: slow-query visibility — per-route call_count/total_ms/max_ms
    # counters plus a rolling log of individual calls that crossed
    # SLOW_QUERY_THRESHOLD_MS, accumulated by _record_query_timing() on the
    # heavy Analytics-tab routes (archetypes, tier-list, underused-gems,
    # hard-counters, matchup-breakdown, battles, leaderboards). Surfaced here
    # rather than only discovered via a slow-page complaint.
    sq_doc = db_sync["config"].find_one({"_id": "slow_query_log"}, {"_id": 0}) or {}
    sq_routes = sq_doc.get("routes", {})
    for _name, _stats in sq_routes.items():
        calls = _stats.get("call_count", 0)
        _stats["avg_ms"] = round(_stats.get("total_ms", 0) / calls, 1) if calls else 0
    result["slow_queries"] = {
        "threshold_ms": SLOW_QUERY_THRESHOLD_MS,
        "routes": sq_routes,
        "recent_slow": sq_doc.get("recent_slow", []),
    }

    # Harvest metadata
    harv = get_harvester()._harvest_meta.copy()
    # Snapshot history: distinct dates from player_snapshots
    dates = sorted(db_sync["player_snapshots"].distinct("date"), reverse=True)[:14]
    harv["history_dates"] = dates
    harv["member_count"] = db_sync["player_profiles"].estimated_document_count()
    harv["war_participants_found"] = db_sync["war_tracking"].find_one(
        {}, sort=[("harvest_time", -1)]
    ) or {}
    harv["war_participants_found"] = len(
        (harv["war_participants_found"].get("clan") or {}).get("participants", [])
    )
    result["harvest"] = harv

    result["tasks"] = {
        "snapshot_loop": f"every {os.getenv('HARVEST_INTERVAL_MINUTES', 30)} min",
        "next_snapshot": "scheduled",
    }

    return jsonify(result)


@web_bp.route("/admin/api/roster")
def admin_api_roster():
    """Enriched member list for the admin Roster tab.
    Merges: live clan API → player_profiles (strikes, notes, war stats)
            → war_tracking (decks used today, fame)
            → users (discord link, discord_name)
            → config (who is a site admin)
    """
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403

    # Falls back to the last successful clan fetch instead of rendering an
    # empty table when the CR API is briefly unreachable.
    clan_data, _ = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    members = clan_data.get("memberList", [])

    # Current war participant data keyed by clean tag
    latest_war = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)]) or {}
    war_participants = {
        p.get("tag", "").replace("#", "").upper(): p
        for p in (latest_war.get("clan") or {}).get("participants", [])
    }

    # DB profiles keyed by clean tag
    profiles = {
        p.get("tag", "").replace("#", "").upper(): p
        for p in db_sync["player_profiles"].find(
            {}, {"tag": 1, "strikes": 1, "admin_notes": 1, "warDayWins": 1,
                 "bestTrophies": 1, "wins": 1, "losses": 1,
                 "trial_member": 1, "trial_started_at": 1}  # idea #213
        )
    }

    # Users (Discord links) keyed by clean cr_tag
    users = {
        u.get("cr_tag", "").replace("#", "").upper(): u
        for u in db_sync["users"].find({}, {"cr_tag": 1, "discord_id": 1, "discord_name": 1})
    }

    # Site admin discord IDs
    config = db_sync["config"].find_one({"_id": "system_config"}) or {}
    admin_ids = set(config.get("admin_user_ids", []))
    master_admin = os.getenv("MASTER_ADMIN_ID", "")

    result = []
    for m in members:
        tag = m.get("tag", "").replace("#", "").upper()
        wp  = war_participants.get(tag, {})
        db  = profiles.get(tag, {})
        usr = users.get(tag, {})
        discord_id = usr.get("discord_id", "")
        result.append({
            "tag":           m.get("tag", ""),
            "name":          m.get("name", ""),
            "role":          m.get("role", "member"),
            "trophies":      m.get("trophies", 0),
            "bestTrophies":  db.get("bestTrophies") or m.get("bestTrophies", 0),
            "donations":     m.get("donations", 0),
            "fame":          wp.get("fame", 0),
            "decksUsedToday": wp.get("decksUsedToday"),   # None = not in war
            "warDayWins":    wp.get("warDayWins") or db.get("warDayWins", 0),
            "in_war":        tag in war_participants,
            "strikes":       db.get("strikes", 0),
            "admin_notes":   db.get("admin_notes", ""),
            "discord_id":    discord_id,
            "discord_name":  usr.get("discord_name", ""),
            "is_site_admin": bool(discord_id and (discord_id in admin_ids or discord_id == master_admin)),
            "trial_member":  bool(db.get("trial_member")),  # idea #213
        })

    return jsonify(result)


def _current_riverrace_with_fallback():
    """currentriverrace already gets stored verbatim into war_tracking every
    harvest cycle (see data_harvester.py) -- fall back to the latest stored
    doc instead of erroring when the live call fails. Returns (data, is_stale)."""
    data = fetch_cr_api(f"clans/%23{CLAN_TAG}/currentriverrace")
    if data:
        return data, False
    stored = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)])
    if stored:
        return {k: v for k, v in stored.items() if k not in ("_id", "harvest_time")}, True
    return None, False


@web_bp.route("/admin/api/war")
def admin_api_war():
    """Current River Race data."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data, data_is_stale = _current_riverrace_with_fallback()
    if not data:
        return jsonify({"error": "Could not fetch current war data from CR API, and no cached fallback exists yet."})

    # Idea #1: is today a War Day or a Training Day, plus decks remaining per
    # member — periodType comes straight from the CR API, decks_remaining is
    # just 4 minus what's already used.
    data["period_type_label"] = "War Day" if data.get("periodType") == "warDay" else "Training Day"
    for p in (data.get("clan") or {}).get("participants", []):
        p["decksRemaining"] = max(0, 4 - p.get("decksUsedToday", 0))

    # Idea #4: projected finish (see compute_projected_finish for the formula).
    data["projected_finish"] = compute_projected_finish(data)
    data["is_stale"] = data_is_stale

    return jsonify(data)


@web_bp.route("/admin/api/war/day-breakdown")
def admin_api_war_day_breakdown():
    """Idea #2: per-war-day fame breakdown (day 1-4) instead of only the
    aggregate race total, using the periodLogs the CR API already returns on
    the current-race endpoint (one entry per completed period this race).
    """
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data, data_is_stale = _current_riverrace_with_fallback()
    if not data:
        return jsonify({"error": "Could not fetch current war data from CR API, and no cached fallback exists yet."})
    own_tag = f"#{CLAN_TAG}"
    logs = data.get("periodLogs", [])
    breakdown = []
    for i, log_entry in enumerate(logs):
        items = log_entry.get("items", [])
        own = next((it for it in items if it.get("clan", {}).get("tag") == own_tag), None)
        breakdown.append({
            "day": i + 1,
            "fame": (own or {}).get("clan", {}).get("fame", 0) if own else 0,
        })
    return jsonify({"days": breakdown, "current_period_index": data.get("periodIndex"), "is_stale": data_is_stale})


@web_bp.route("/admin/api/war/scouting")
def admin_api_war_scouting():
    """Idea #3: opponent clan scouting — pulls each rival clan's member list and
    average trophies once a race is active. This IS a set of new live CR API
    calls (one per rival clan), so admin-gated per the project's existing rule.
    """
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    war_data, war_data_is_stale = _current_riverrace_with_fallback()
    if not war_data:
        return jsonify({"error": "No active race to scout, and no cached fallback exists yet."})
    own_tag = f"#{CLAN_TAG}"
    rival_summaries = []
    for c in war_data.get("clans", []):
        tag = c.get("tag", "")
        if tag == own_tag or not tag:
            continue
        # Rival clans are external, unaffiliated clans with no Mongo snapshot
        # of their own here -- appropriately live-only, unlike our own clan's data.
        detail = fetch_cr_api(f"clans/{tag.replace('#', '%23')}")
        members = (detail or {}).get("memberList", [])
        avg_trophies = round(sum(m.get("trophies", 0) for m in members) / len(members)) if members else None
        rival_summaries.append({
            "tag": tag, "name": c.get("name", detail.get("name") if detail else "Unknown"),
            "member_count": len(members), "avg_trophies": avg_trophies,
            "current_fame": c.get("fame", 0),
        })
    return jsonify({"rivals": rival_summaries, "is_stale": war_data_is_stale})


@web_bp.route("/admin/api/war/calendar")
def admin_api_war_calendar():
    """Idea #6: a month-grid, GitHub-contributions-style calendar of war outcomes
    (win/loss) built from the already-harvested war_history collection — no new
    CR API calls, pure Mongo read.
    """
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    own_tag = f"#{CLAN_TAG}"
    races = list(db_sync["war_history"].find({}, {"_id": 0}))
    days = []
    for race in races:
        data = race.get("data", {})
        clan_list = sorted(data.get("standings", []) or [], key=lambda s: s.get("rank", 99))
        own_rank = next((s.get("rank") for s in clan_list if (s.get("clan") or {}).get("tag") == own_tag), None)
        outcome = "win" if own_rank == 1 else ("loss" if own_rank else "unknown")
        days.append({
            "unique_war_id": race.get("unique_war_id"),
            "season_id": data.get("seasonId"),
            "section_index": data.get("sectionIndex"),
            "outcome": outcome,
            "rank": own_rank,
        })
    return jsonify({"races": days})


@web_bp.route("/admin/api/war/missed-streaks")
def admin_api_war_missed_streaks():
    """Idea #8: consecutive recent war weeks with 0 decks used, per member —
    catches a slow decline before it becomes a formal strike. Reads the last 10
    completed races from war_history (no new CR API calls)."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    races = list(db_sync["war_history"].find({}, {"_id": 0}).sort("data.seasonId", -1).limit(10))
    per_member = {}
    for race in races:
        participants = (race.get("data", {}).get("clan") or {}).get("participants", [])
        for p in participants:
            tag = p.get("tag")
            if not tag:
                continue
            per_member.setdefault(tag, {"name": p.get("name"), "streak": 0, "broke": False})
            entry = per_member[tag]
            if entry["broke"]:
                continue
            if p.get("decksUsed", p.get("decksUsedToday", 0)) == 0:
                entry["streak"] += 1
            else:
                entry["broke"] = True
    streaks = [
        {"tag": t, "name": v["name"], "missed_streak": v["streak"]}
        for t, v in per_member.items() if v["streak"] >= 2
    ]
    streaks.sort(key=lambda x: -x["missed_streak"])
    return jsonify({"streaks": streaks})


@web_bp.route("/admin/api/player/excuse", methods=["POST"])
def admin_player_excuse():
    """Idea #9: one-click excuse from the current war so the slacker banner and
    tiered nudges (idea #13) stop nagging about a member leadership already
    knows is away. Toggle off manually once they're back."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    tag = clean_tag(data.get("tag", ""))
    excused = bool(data.get("excused", True))
    db_sync["player_profiles"].update_one(
        {"tag": f"#{tag}"}, {"$set": {"war_excused": excused}}, upsert=True
    )
    log_admin_activity("Excused from war" if excused else "Un-excused from war", target=tag)
    return jsonify({"success": True, "tag": tag, "excused": excused})


@web_bp.route("/admin/api/player/bench", methods=["POST"])
def admin_player_bench():
    """Idea #16: vacation-mode / bench designation, specified in WEEKS (not
    days, per review feedback) — excludes the member from slacker/strike logic
    until the bench period ends. POST {"tag": ..., "weeks": 2} to bench,
    {"tag": ..., "weeks": 0} to clear early."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    tag = clean_tag(data.get("tag", ""))
    try:
        weeks = float(data.get("weeks", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "weeks must be a number"}), 400
    if weeks <= 0:
        db_sync["player_profiles"].update_one({"tag": f"#{tag}"}, {"$unset": {"benched_until": ""}}, upsert=True)
        return jsonify({"success": True, "tag": tag, "benched_until": None})
    until = datetime.now(timezone.utc) + timedelta(weeks=weeks)
    db_sync["player_profiles"].update_one({"tag": f"#{tag}"}, {"$set": {"benched_until": until}}, upsert=True)
    return jsonify({"success": True, "tag": tag, "benched_until": until.isoformat()})


@web_bp.route("/admin/api/war/retro/<unique_war_id>", methods=["GET", "POST"])
def admin_war_retro(unique_war_id):
    """Idea #20: a free-text retro-notes box per completed race, PAIRED with an
    auto-generated data-driven recap (top fame, participation %, slacker list)
    computed from the same war_history doc — not just a blank text box."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    race = db_sync["war_history"].find_one({"unique_war_id": unique_war_id})
    if not race:
        return jsonify({"error": "Unknown war id"}), 404

    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        db_sync["war_history"].update_one(
            {"unique_war_id": unique_war_id},
            {"$set": {"retro_notes": str(data.get("notes", ""))[:5000],
                      "retro_notes_updated_at": datetime.now(timezone.utc)}},
        )
        return jsonify({"success": True})

    clan = race.get("data", {}).get("clan", {})
    participants = sorted(clan.get("participants", []), key=lambda p: p.get("fame", 0), reverse=True)
    total_possible_decks = len(participants) * 4
    decks_used = sum(p.get("decksUsed", p.get("decksUsedToday", 0)) for p in participants)
    recap = {
        "clan_name": clan.get("name"),
        "total_fame": clan.get("fame", 0),
        "top_fame": participants[:5],
        "slackers": [p for p in participants if p.get("decksUsed", p.get("decksUsedToday", 0)) == 0],
        "participation_pct": round(decks_used / total_possible_decks * 100, 1) if total_possible_decks else 0,
    }
    return jsonify({
        "recap": recap,
        "retro_notes": race.get("retro_notes", ""),
        "retro_notes_updated_at": race.get("retro_notes_updated_at"),
    })


def _recruit_quality_score(player: dict) -> int:
    """Idea #211: a single at-a-glance 0-100 recruit-quality score, combining
    trophies (the strongest available signal), an estimate of war
    participation (the CR API doesn't expose a player's per-clan war history
    outside their current clan, so this falls back to their current clan's
    warDayWins if present), and donation habits. Weighted 50/30/20 — trophies
    dominate since that's the only signal that's always present and reliable;
    the other two are soft signals that only apply when visible."""
    trophies = player.get("trophies", 0) or 0
    trophy_score = min(50, (trophies / 8000) * 50)
    war_wins = player.get("warDayWins", 0) or 0
    war_score = min(30, (war_wins / 50) * 30)
    donations = player.get("donations", 0) or 0
    donation_score = min(20, (donations / 500) * 20)
    return round(trophy_score + war_score + donation_score)


@web_bp.route("/admin/api/scout/<tag>")
def admin_scout_player(tag):
    """Idea #206: global player search by tag — a real live lookup for
    "check a recruit before inviting", not just a client-side filter over the
    current roster (which is all the sidebar's global search box did before).
    This IS a new live CR API call, so it's admin-gated per this project's
    existing rule for any new per-view external call. Idea #211's
    recruit-quality score is computed on the same response."""
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    clean = clean_tag(tag)
    player = fetch_cr_api(f"players/%23{clean}")
    if not player:
        return jsonify({"error": "Player not found."}), 404
    current_clan = (player.get("clan") or {})
    return jsonify({
        "tag": player.get("tag"), "name": player.get("name"),
        "trophies": player.get("trophies"), "bestTrophies": player.get("bestTrophies"),
        "expLevel": player.get("expLevel"), "donations": player.get("donations"),
        "warDayWins": player.get("warDayWins"),
        "wins": player.get("wins"), "losses": player.get("losses"),
        "current_clan_name": current_clan.get("name"), "current_clan_tag": current_clan.get("tag"),
        "current_clan_role": player.get("role"),
        "quality_score": _recruit_quality_score(player),
    })


@web_bp.route("/admin/api/rival-clan/<tag>")
def admin_rival_clan_snapshot(tag):
    """Idea #208: a rival-clan comparison snapshot, fetched once per scouting
    session (client caches the response rather than this app polling a rival
    clan continuously) rather than a live per-view call."""
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    clean = clean_tag(tag)
    rival = fetch_cr_api(f"clans/%23{clean}")
    if not rival:
        return jsonify({"error": "Clan not found."}), 404
    members = rival.get("memberList", []) or []
    avg_trophies = round(sum(m.get("trophies", 0) for m in members) / len(members)) if members else 0
    # Our own clan's side has a stored fallback (unlike the rival above, an
    # external clan with nothing to fall back to) -- use it so a transient CR
    # API hiccup doesn't make OUR side of the comparison look worse than theirs.
    our_clan, _ = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    our_clan = our_clan or {}
    our_members = our_clan.get("memberList", []) or []
    our_avg_trophies = round(sum(m.get("trophies", 0) for m in our_members) / len(our_members)) if our_members else 0
    return jsonify({
        "rival": {
            "name": rival.get("name"), "tag": rival.get("tag"),
            "memberCount": rival.get("memberCount"), "avg_trophies": avg_trophies,
            "clanWarTrophies": rival.get("clanWarTrophies"),
        },
        "us": {
            "name": our_clan.get("name"), "tag": our_clan.get("tag"),
            "memberCount": our_clan.get("memberCount"), "avg_trophies": our_avg_trophies,
            "clanWarTrophies": our_clan.get("clanWarTrophies"),
        },
    })


@web_bp.route("/admin/api/recruit-templates", methods=["GET", "POST"])
def admin_recruit_templates():
    """Idea #209: a small library of canned invite/recruiting messages admins
    can quickly copy, instead of retyping the same pitch every time."""
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    if request.method == "GET":
        doc = db_sync["config"].find_one({"_id": "recruit_templates"}, {"_id": 0}) or {}
        return jsonify({"templates": doc.get("templates", [])})
    data = request.get_json(silent=True) or {}
    name = str(data.get("name", "")).strip()[:60]
    message = str(data.get("message", "")).strip()[:500]
    if not name or not message:
        return jsonify({"error": "name and message are required"}), 400
    db_sync["config"].update_one(
        {"_id": "recruit_templates"},
        {"$push": {"templates": {"name": name, "message": message}}},
        upsert=True,
    )
    return jsonify({"success": True})


@web_bp.route("/apply", methods=["GET", "POST"])
def application_form():
    """Idea #212: an application form for prospective members instead of raw
    Discord DMs, so there's a durable, reviewable record rather than a DM that
    gets lost in someone's inbox."""
    if request.method == "GET":
        bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
        return render_sandboxed(
            get_template("apply"),
            min_trophies=bot_settings.get("min_trophies", 6500),
            csrf_token=get_csrf_token(), submitted=False,
        )
    data = request.form if request.form else (request.get_json(silent=True) or {})
    tag = clean_tag(data.get("tag", ""))
    name = str(data.get("name", "")).strip()[:60]
    message = str(data.get("message", "")).strip()[:500]
    if not tag or not name:
        return render_sandboxed(get_template("apply"), csrf_token=get_csrf_token(), submitted=False, error="Tag and name are required."), 400
    db_sync["applications"].insert_one({
        "tag": tag, "name": name, "message": message,
        "status": "pending",  # pending | invited | declined
        "created_at": datetime.now(timezone.utc),
    })
    return render_sandboxed(get_template("apply"), csrf_token=get_csrf_token(), submitted=True)


@web_bp.route("/admin/api/applications")
def admin_list_applications():
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    items = list(db_sync["applications"].find({}, {"_id": 0}).sort("created_at", -1).limit(100))
    return jsonify({"applications": items})


@web_bp.route("/admin/api/applications/status", methods=["POST"])
def admin_set_application_status():
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    status = str(data.get("status", "")).strip()
    if status not in ("pending", "invited", "declined"):
        return jsonify({"error": "invalid status"}), 400
    db_sync["applications"].update_one(
        {"tag": clean_tag(data.get("tag", "")), "created_at": data.get("created_at")},
        {"$set": {"status": status}},
    )
    return jsonify({"success": True})


@web_bp.route("/admin/api/player/<tag>/trial-status", methods=["POST"])
def admin_set_trial_status(tag):
    """Idea #213: a temporary trial-member flag for new recruits under a
    probation period before full member status — purely informational (shown
    as a badge in the roster drawer); doesn't touch Discord roles."""
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    is_trial = bool(data.get("trial_member", True))
    update = {"trial_member": is_trial}
    if is_trial:
        update["trial_started_at"] = datetime.now(timezone.utc)
    db_sync["player_profiles"].update_one({"tag": f"#{clean_tag(tag)}"}, {"$set": update}, upsert=True)
    return jsonify({"success": True})


@web_bp.route("/admin/api/recruits", methods=["GET", "POST"])
def admin_recruits():
    """Idea #14: a saved recruiting shortlist (scouted tags + notes + status) so
    scouting work isn't lost between sessions."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        tag = clean_tag(data.get("tag", ""))
        if not tag:
            return jsonify({"error": "tag is required"}), 400
        db_sync["recruit_shortlist"].update_one(
            {"tag": tag},
            {"$set": {
                "tag": tag,
                "name": data.get("name", ""),
                "notes": data.get("notes", ""),
                "status": data.get("status", "scouted"),  # scouted | invited | joined | declined
                "updated_at": datetime.now(timezone.utc),
            }, "$setOnInsert": {"added_at": datetime.now(timezone.utc)}},
            upsert=True,
        )
        return jsonify({"success": True})
    recruits = list(db_sync["recruit_shortlist"].find({}, {"_id": 0}).sort("updated_at", -1))
    return jsonify({"recruits": recruits})


@web_bp.route("/admin/api/recruits/<tag>", methods=["DELETE"])
def admin_recruits_delete(tag):
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    db_sync["recruit_shortlist"].delete_one({"tag": clean_tag(tag)})
    return jsonify({"success": True})


@web_bp.route("/admin/api/clan/departed")
def admin_departed_members():
    """Idea #17 (revised): visibility into who's in the ~23-week retention
    window before their historical data is purged by the harvester."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    from data_harvester import DEPARTED_MEMBER_RETENTION_WEEKS
    departed = list(db_sync["player_profiles"].find(
        {"left_clan_at": {"$exists": True}}, {"tag": 1, "name": 1, "left_clan_at": 1}
    ))
    now = datetime.now(timezone.utc)
    out = []
    for d in departed:
        # Bugfix: pymongo returns naive datetimes (this project's MongoClient
        # never sets tz_aware=True), so left_clan_at came back naive while
        # `now` above is timezone-aware — subtracting them crashed with
        # "can't subtract offset-naive and offset-aware datetimes" the moment
        # a real departed member existed. Same pattern already fixed
        # elsewhere in this file via _as_aware_utc(); this line just predated
        # that fix and wasn't caught by the earlier grep sweep.
        left_at = _as_aware_utc(d.get("left_clan_at"))
        purge_at = left_at + timedelta(weeks=DEPARTED_MEMBER_RETENTION_WEEKS) if left_at else None
        out.append({
            "tag": d.get("tag"), "name": d.get("name"),
            "left_clan_at": left_at, "purge_at": purge_at,
            "days_until_purge": (purge_at - now).days if purge_at else None,
        })
    return jsonify({"departed": out, "retention_weeks": DEPARTED_MEMBER_RETENTION_WEEKS})


@web_bp.route("/admin/api/war/previous")
def admin_api_war_previous():
    """Most recent completed River Race from the log."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = fetch_cr_api(f"clans/%23{CLAN_TAG}/riverracelog?limit=1")
    items = data.get("items", []) if data else []
    if items:
        return jsonify(items[0])
    # Live riverracelog call failed (or came back empty) -- war_history already
    # stores exactly this kind of completed-race data every time backfill_missed_wars
    # runs, so fall back to the most recent race there instead of an empty result.
    last_race = db_sync["war_history"].find_one(
        {}, sort=[("data.seasonId", -1), ("data.sectionIndex", -1)]
    )
    if last_race and last_race.get("data"):
        result = dict(last_race["data"])
        result["is_stale"] = True
        return jsonify(result)
    return jsonify({"standings": []})


@web_bp.route("/admin/api/war/aggregate")
def admin_api_war_aggregate():
    """Aggregate war stats from war_history collection."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    pipeline = [
        {"$unwind": "$data.clan.participants"},
        {"$group": {"_id": None, "total_fame": {"$sum": "$data.clan.fame"}}},
    ]
    rows = list(db_sync["war_history"].aggregate(pipeline))
    total_fame = rows[0]["total_fame"] if rows else 0

    # Also pull from live war_tracking collection as a fallback
    if total_fame == 0:
        for doc in db_sync["war_tracking"].find({}, {"_id": 0, "clan.fame": 1}):
            total_fame += (doc.get("clan") or {}).get("fame", 0)

    return jsonify({"total_fame": total_fame})


@web_bp.route("/admin/api/battles")
def admin_api_battles():
    """Battle records across all clan members, most recent first.

    Idea #234: paginated instead of always loading the last 100 in one
    response — `page` (0-indexed) and `page_size` (default 25, capped at 100
    to keep a single response bounded) control the window; `total_count` in
    the response lets the frontend render Prev/Next controls. Old callers
    that don't pass either param still get a working response (page 0,
    default page_size), so this isn't a breaking change."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    try:
        page = max(int(request.args.get("page", 0)), 0)
        page_size = min(max(int(request.args.get("page_size", 25)), 1), 100)
    except (TypeError, ValueError):
        page, page_size = 0, 25
    total_count = db_sync["battle_history"].count_documents({})
    battles = list(
        db_sync["battle_history"]
        .find({}, {"_id": 0})
        .sort("battle_time", -1)
        .skip(page * page_size)
        .limit(page_size)
    )
    # Ensure card lists are name strings for the JS renderer
    for b in battles:
        b["team_cards"] = [
            c if isinstance(c, str) else c.get("name", str(c))
            for c in (b.get("team_cards") or [])
        ]
        b["opponent_cards"] = [
            c if isinstance(c, str) else c.get("name", str(c))
            for c in (b.get("opponent_cards") or [])
        ]
    return jsonify({
        "battles": battles, "page": page, "page_size": page_size,
        "total_count": total_count,
        "has_more": (page + 1) * page_size < total_count,
    })


@web_bp.route("/api/player/<tag>/battles")
def api_player_battles(tag):
    """Per-player recent battles — used by the player profile page's async loader."""
    battles = list(
        db_sync["battle_history"]
        .find({"player_tag": clean_tag(tag)}, {"_id": 0})
        .sort("battle_time", -1)
        .limit(20)
    )
    for b in battles:
        b["team_cards"] = [
            c if isinstance(c, str) else c.get("name", str(c))
            for c in (b.get("team_cards") or [])
        ]
        b["opponent_cards"] = [
            c if isinstance(c, str) else c.get("name", str(c))
            for c in (b.get("opponent_cards") or [])
        ]
    return jsonify(battles)


@web_bp.route("/api/cards/icons")
def api_cards_icons():
    """Name -> icon URL map for rendering real card art instead of the 🃏
    placeholder on the player page's battle log.

    Root cause this exists to fix: harvest_battles() in data_harvester.py has
    always flattened team_cards/opponent_cards down to plain name strings
    before storage (needed to avoid the dict-sorting/dict-key crashes fixed
    elsewhere in this file via _normalize_card_names() — storing rich card
    objects there reintroduces that exact bug class). That means no icon URL
    ever reached the frontend, and every card chip on every player's battle
    log always fell back to a placeholder, on every deploy, forever.

    Rather than resurrecting rich card objects in battle_history, the
    harvester now separately upserts name->icon-URL pairs (real data seen
    from the CR API, not hand-typed/guessed) into a small `card_icons`
    collection as battles are harvested. This route just serves that map.
    Public/no-auth: it's static card-art metadata, nothing player-specific.
    """
    icons = {
        doc["_id"]: doc["icon_url"]
        for doc in db_sync["card_icons"].find({}, {"icon_url": 1})
        if doc.get("icon_url")
    }
    return jsonify(icons)


@web_bp.route("/api/player/<tag>/insights")
def api_player_insights(tag):
    """Per-player battle intelligence — most-used cards, toughest/most-defeated
    matchups, win streaks, rival opponent, busiest battle day. Computed entirely
    from this player's own logged battle_history — no new Clash Royale API calls,
    same pattern as /api/player/<tag>/battles above.
    """
    clean = clean_tag(tag)
    battles = list(
        db_sync["battle_history"]
        .find(
            {"player_tag": clean},
            {"_id": 0, "team_cards": 1, "opponent_cards": 1, "result": 1,
             "battle_time": 1, "opponent_name": 1, "opponent_tag": 1},
        )
        .sort("battle_time", 1)  # chronological — needed for streak math
    )
    if not battles:
        return jsonify({"total_battles": 0})

    def card_name(c):
        return c if isinstance(c, str) else (c or {}).get("name")

    card_usage = {}       # my card -> {games, wins}
    opp_card_faced = {}   # opponent card -> {encounters, wins, losses}
    rival_record = {}     # (opponent name, tag) -> {w, l, d}
    day_counts = {}
    running_streak = 0
    longest_streak = 0

    for b in battles:
        result = b.get("result")

        for raw in (b.get("team_cards") or []):
            c = card_name(raw)
            if not c: continue
            e = card_usage.setdefault(c, {"games": 0, "wins": 0})
            e["games"] += 1
            if result == "win": e["wins"] += 1

        for raw in (b.get("opponent_cards") or []):
            c = card_name(raw)
            if not c: continue
            e = opp_card_faced.setdefault(c, {"encounters": 0, "wins": 0, "losses": 0})
            e["encounters"] += 1
            if result == "win": e["wins"] += 1
            elif result == "loss": e["losses"] += 1

        opp_key = (b.get("opponent_name") or "Unknown", b.get("opponent_tag") or "")
        rec = rival_record.setdefault(opp_key, {"w": 0, "l": 0, "d": 0})
        if result == "win": rec["w"] += 1
        elif result == "loss": rec["l"] += 1
        else: rec["d"] += 1

        # CR API battle_time looks like "20240312T183045.000Z" — first 8 chars = YYYYMMDD
        day = (b.get("battle_time") or "")[:8]
        if len(day) == 8:
            day_counts[day] = day_counts.get(day, 0) + 1

        if result == "win":
            running_streak += 1
            longest_streak = max(longest_streak, running_streak)
        else:
            running_streak = 0

    most_used_cards = sorted(
        [
            {"card": c, "games": v["games"], "win_rate": round(v["wins"] / v["games"] * 100, 1)}
            for c, v in card_usage.items()
        ],
        key=lambda x: -x["games"],
    )[:5]

    MIN_ENCOUNTERS = 3
    qualifying = [
        {
            "card": c, "encounters": v["encounters"], "wins": v["wins"], "losses": v["losses"],
            "win_rate": round(v["wins"] / v["encounters"] * 100, 1),
        }
        for c, v in opp_card_faced.items() if v["encounters"] >= MIN_ENCOUNTERS
    ]
    most_defeated = sorted(qualifying, key=lambda x: (-x["wins"], -x["win_rate"]))[:1]
    toughest_matchup = sorted(qualifying, key=lambda x: (x["win_rate"], -x["encounters"]))[:1]

    rival = None
    if rival_record:
        (r_name, r_tag), rec = max(rival_record.items(), key=lambda kv: kv[1]["w"] + kv[1]["l"] + kv[1]["d"])
        rival = {"name": r_name, "tag": r_tag, "wins": rec["w"], "losses": rec["l"], "draws": rec["d"]}

    busiest_day = None
    if day_counts:
        d, count = max(day_counts.items(), key=lambda kv: kv[1])
        busiest_day = {"date": f"{d[0:4]}-{d[4:6]}-{d[6:8]}", "battles": count}

    return jsonify({
        "total_battles": len(battles),
        "most_used_cards": most_used_cards,
        "most_defeated_card": most_defeated[0] if most_defeated else None,
        "toughest_matchup": toughest_matchup[0] if toughest_matchup else None,
        "current_streak": running_streak,
        "longest_win_streak": longest_streak,
        "rival": rival,
        "busiest_day": busiest_day,
    })


@web_bp.route("/admin/api/template/<name>")
def admin_get_template(name):
    """Serve a template to the UI editor — ?source=current pulls live DB, ?source=default pulls the DEFAULT_ constant."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    allowed = {"roster", "player", "admin", "link"}
    if name not in allowed:
        return jsonify({"error": f"Unknown template '{name}'"}), 400

    source = request.args.get("source", "current")
    if source == "current":
        doc = db_sync["config"].find_one({"_id": "html_templates"}) or {}
        html = doc.get(name, "")
        if not html:
            # Fall back to default if nothing is deployed yet
            source = "default"

    if source == "default":
        # Read the canonical file from disk, inside the templates/ subfolder.
        import pathlib
        candidates = [
            pathlib.Path(__file__).parent / "templates" / f"{name}.html",
        ]
        html = ""
        for path in candidates:
            if path.exists():
                html = path.read_text(encoding="utf-8")
                break
        if not html:
            return jsonify({"error": f"Default file for '{name}' not found on disk."}), 404

    return jsonify({"html": html, "template": name, "source": source})


@web_bp.route("/admin/update-html", methods=["POST"])
def admin_update_html():
    """Deploy an HTML template from the UI editor into the DB."""
    if not has_full_admin(): return jsonify({"error": "unauthorized — deploying requires full-admin tier"}), 403
    template_name = request.form.get("template_name", "").strip()
    html_content = request.form.get("html_content", "").strip()
    allowed = {"roster", "player", "admin", "link"}
    if template_name not in allowed:
        return jsonify({"error": f"Unknown template '{template_name}'"}), 400
    if not html_content:
        return jsonify({"error": "html_content is empty"}), 400

    db_sync["config"].update_one(
        {"_id": "html_templates"},
        {"$set": {template_name: html_content}},
        upsert=True,
    )
    # Invalidate the in-process HTML cache so the next page load picks up the change
    with _cache_lock:
        _HTML_CACHE.pop(template_name, None)

    # Idea #67: keep the last 5 deployed versions per template so a bad deploy
    # can be rolled back without digging through git or Mongo manually.
    history_doc = db_sync["config"].find_one({"_id": "html_template_history"}) or {}
    versions = history_doc.get(template_name, [])
    versions.insert(0, {"html": html_content, "deployed_at": datetime.now(timezone.utc).isoformat()})
    versions = versions[:5]
    db_sync["config"].update_one(
        {"_id": "html_template_history"}, {"$set": {template_name: versions}}, upsert=True
    )

    log_admin_activity("Deployed template", target=template_name, details=f"{len(html_content)} chars")
    return jsonify({"success": True, "template": template_name})


@web_bp.route("/admin/preview", methods=["POST"])
def admin_preview():
    """Render a template with dummy context so the editor's preview button works."""
    if not is_admin(): return "Unauthorized", 403
    template_name = request.form.get("template_name", "")
    html_content = request.form.get("html_content", "")
    if not html_content:
        return "Empty template", 400

    # Provide enough dummy context for the template to render without crashing
    dummy_clan = fetch_cr_api(f"clans/%23{CLAN_TAG}") or {"memberList": [], "memberCount": 0}
    dummy_player = {"name": "Preview Player", "tag": "#PREVIEW", "trophies": 9999,
                    "bestTrophies": 9999, "wins": 100, "losses": 50, "threeCrownWins": 30,
                    "battleCount": 150, "donations": 200, "donationsReceived": 180,
                    "warDayWins": 12, "expLevel": 14, "role": "member",
                    "arena": {"name": "Ultimate Champion"}, "currentDeck": [],
                    "current_streak": 3, "currentFavouriteCard": {"name": "Goblin Barrel"}}
    try:
        rendered = render_sandboxed(
            html_content,
            clan_data=dummy_clan,
            player=dummy_player,
            data=dummy_player,
            db_player={},
            is_admin=True,
            clan_tag=CLAN_TAG,
            max_lvl=MAX_CARD_LEVEL,
        )
        return rendered
    except Exception as e:
        return f"<pre>Template render error:\n{e}</pre>", 400


@web_bp.route("/admin/harvest/manual", methods=["POST"])
def admin_harvest_manual():
    """Alias for /admin/api/harvest/trigger — the new admin JS calls this path."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    harvester = get_harvester()
    threading.Thread(target=harvester.run_full_cycle, daemon=True).start()
    return jsonify({"success": True, "message": "Harvest started in the background."})


@web_bp.route("/admin/api/snapshot/<date>")
def admin_api_snapshot(date):
    """Return all player snapshots for a given date (YYYY-MM-DD) as JSON."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    docs = list(db_sync["player_snapshots"].find({"date": date}, {"_id": 0}))
    return jsonify(docs)


@web_bp.route("/admin/api/users/tier", methods=["POST"])
def admin_set_user_tier():
    """Idea #65: assign a permission tier — "full" (everything) or
    "analytics_only" (view Roster/Analytics/Diagnostics/Recruiting, but not
    Settings/UI Editor/Cache Flush/User Access changes)."""
    if not has_full_admin(): return jsonify({"error": "unauthorized — only full admins can change permission tiers"}), 403
    data = request.get_json(silent=True) or {}
    discord_id = str(data.get("discord_id", "")).strip()
    tier = data.get("tier", "full")
    if tier not in ("full", "analytics_only"):
        return jsonify({"error": "tier must be 'full' or 'analytics_only'"}), 400
    if not discord_id.isdigit():
        return jsonify({"error": "a valid numeric discord_id is required"}), 400
    db_sync["config"].update_one(
        {"_id": "system_config"}, {"$set": {f"admin_tiers.{discord_id}": tier}}, upsert=True
    )
    log_admin_activity("Changed admin tier", target=discord_id, details=tier)
    return jsonify({"success": True})


@web_bp.route("/admin/api/users")
def admin_api_users():
    """User table for the User Access tab."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403

    config = db_sync["config"].find_one({"_id": "system_config"}) or {}
    admin_ids = set(config.get("admin_user_ids", []))
    master_admin = os.getenv("MASTER_ADMIN_ID", "")

    # Build a tag -> profile map for clan rank
    profiles = {
        p.get("tag", ""): p
        for p in db_sync["player_profiles"].find({}, {"tag": 1, "role": 1, "name": 1})
    }

    # "Lay this info out in the admin panel" — resolves each stored server
    # role ID to its display name via fetch_guild_role_names() (one cached
    # Discord call for the whole table, not one per user).
    role_names = fetch_guild_role_names()

    users = list(db_sync["users"].find({}, {"_id": 0}))
    result = []
    for u in users:
        cr_tag = u.get("cr_tag", "")
        profile = profiles.get(cr_tag, {})
        discord_id = u.get("discord_id", "")
        role_ids = u.get("discord_roles") or []
        result.append({
            "discord_id": discord_id,
            "name": profile.get("name") or u.get("discord_name", "Unknown"),
            # The raw Discord username, distinct from the above -- "name" prefers
            # the in-clan CR player name when linked, which can differ from (or be
            # totally unrelated to) what someone actually goes by on Discord.
            "discord_username": u.get("discord_name") or "Unknown",
            "cr_tag": cr_tag.replace("#", "") if cr_tag else "",
            "is_linked": bool(cr_tag),
            "rank": profile.get("role", "—"),
            "status": "Admin" if (discord_id in admin_ids or discord_id == master_admin) else "Member",
            # idea: "connect the discord id with information like their name
            # or picture" — avatar hash was already captured off Discord's own
            # OAuth response (see oauth_callback) but never surfaced anywhere.
            "avatar_url": discord_avatar_url(discord_id, u.get("discord_avatar")) if discord_id else None,
            # Falls back to the raw ID string if fetch_guild_role_names()
            # couldn't resolve it (no bot token configured, Discord API hiccup,
            # or a role that's since been deleted server-side) -- always shows
            # *something* rather than silently dropping a role the user has.
            "roles": [role_names.get(rid, rid) for rid in role_ids],
        })

    return jsonify(result)


@web_bp.route("/admin/users/update", methods=["POST"])
def admin_users_update():
    """Form submit from the User Access tab — promote or demote by Discord ID."""
    if not is_admin(): return "Unauthorized", 403
    discord_id = request.form.get("discord_id", "").strip()
    status = request.form.get("status", "member")
    if not discord_id.isdigit():
        return redirect("/admin")
    if status == "admin":
        db_sync["config"].update_one(
            {"_id": "system_config"},
            {"$addToSet": {"admin_user_ids": discord_id}},
            upsert=True,
        )
    else:
        db_sync["config"].update_one(
            {"_id": "system_config"},
            {"$pull": {"admin_user_ids": discord_id}},
        )
    return redirect("/admin")


# ---------------------------------------------------------------------------
# 5b. ANALYTICS (data-driven member/clan dashboard)
# ---------------------------------------------------------------------------
@web_bp.route("/admin/api/analytics/overview")
def admin_analytics_overview():
    """Clan-wide headline numbers for the Analytics tab's stat row."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403

    clan_data, clan_data_is_stale = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    members = clan_data.get("memberList", [])
    member_count = len(members)

    total_trophies = sum(m.get("trophies", 0) for m in members)
    total_donations = sum(m.get("donations", 0) for m in members)
    inactive = sum(1 for m in members if m.get("donations", 0) == 0)

    latest_war = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)]) or {}
    participants = (latest_war.get("clan") or {}).get("participants", [])
    decks_used = sum(p.get("decksUsedToday", 0) for p in participants)
    max_decks = len(participants) * 4
    war_participation_pct = round((decks_used / max_decks) * 100, 1) if max_decks else 0

    battle_count = db_sync["battle_history"].estimated_document_count()

    total_w = db_sync["battle_history"].count_documents({"result": "win"})
    total_l = db_sync["battle_history"].count_documents({"result": "loss"})
    overall_win_rate = round((total_w / (total_w + total_l)) * 100, 1) if (total_w + total_l) else 0

    # Idea #5: flag the clan as under-strength ahead of matchmaking if member
    # count drops below a configurable Settings threshold (default 40).
    system_config = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    min_clan_size = int(system_config.get("min_clan_size", 40))
    under_strength = member_count < min_clan_size

    # Idea #7: boat battle tracking as its own stat, not just a CSV export field.
    # "boatBattle" is the CR API's own `type` value for these, already flattened
    # into battle_type by the harvester.
    boat_total = db_sync["battle_history"].count_documents({"battle_type": "boatBattle"})
    boat_wins = db_sync["battle_history"].count_documents({"battle_type": "boatBattle", "result": "win"})
    boat_win_rate = round(boat_wins / boat_total * 100, 1) if boat_total else 0

    return jsonify({
        "member_count": member_count,
        "avg_trophies": round(total_trophies / member_count) if member_count else 0,
        "total_donations_live": total_donations,
        "inactive_members": inactive,
        "war_participation_pct": war_participation_pct,
        "battles_logged": battle_count,
        "overall_win_rate": overall_win_rate,
        "under_strength": under_strength,
        "min_clan_size": min_clan_size,
        "boat_battles_logged": boat_total,
        "boat_battle_win_rate": boat_win_rate,
        "is_stale": clan_data_is_stale,
    })


@web_bp.route("/admin/api/analytics/leaderboards")
def admin_analytics_leaderboards():
    """Top-N leaderboards: donators, trophy climbers (7d delta), win rate, war fame, streaks."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403

    # Idea #236: this runs a $group aggregate over all of battle_history plus a
    # full war_history scan every time it's called — cache the response briefly
    # so opening the Analytics tab repeatedly doesn't redo that work each time.
    cache_key = _analytics_cache_key("leaderboards")
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()

    clan_data, _ = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    members = clan_data.get("memberList", [])

    top_donators = sorted(members, key=lambda m: m.get("donations", 0), reverse=True)[:10]
    top_donators = [{"name": m.get("name"), "tag": m.get("tag", "").replace("#", ""), "value": m.get("donations", 0)} for m in top_donators]

    # Trophy climbers: compare current trophies against the earliest snapshot in the last 7 days
    import datetime as _dt
    week_ago = (_dt.datetime.now(timezone.utc) - _dt.timedelta(days=7)).strftime("%Y-%m-%d")
    # Ascending sort, but keep only the FIRST (oldest) snapshot seen per tag — the
    # baseline to climb from. A plain dict comprehension here would keep the last
    # value instead, comparing current trophies against yesterday's snapshot rather
    # than a week ago.
    old_snaps = {}
    for s in db_sync["player_snapshots"].find({"date": {"$gte": week_ago}}).sort("date", 1):
        old_snaps.setdefault(s["tag"], s.get("trophies", 0))
    climbers = []
    for m in members:
        tag = m.get("tag", "")
        old = old_snaps.get(tag)
        if old is not None:
            climbers.append({"name": m.get("name"), "tag": tag.replace("#", ""), "value": m.get("trophies", 0) - old})
    climbers.sort(key=lambda x: x["value"], reverse=True)
    top_climbers = climbers[:10]

    # Win rate from logged battles (min 10 battles to qualify, avoids noisy small samples)
    win_pipeline = [
        {"$group": {
            "_id": "$player_tag",
            "wins": {"$sum": {"$cond": [{"$eq": ["$result", "win"]}, 1, 0]}},
            "total": {"$sum": 1},
        }},
        {"$match": {"total": {"$gte": 10}}},
    ]
    win_rows = list(db_sync["battle_history"].aggregate(win_pipeline))
    name_by_tag = {m.get("tag", "").replace("#", ""): m.get("name") for m in members}
    win_rates = [
        {
            "name": name_by_tag.get(r["_id"], r["_id"]),
            "tag": r["_id"],
            "value": round((r["wins"] / r["total"]) * 100, 1),
            "battles": r["total"],
        }
        for r in win_rows
    ]
    win_rates.sort(key=lambda x: x["value"], reverse=True)
    top_win_rate = win_rates[:10]

    latest_war = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)]) or {}
    war_participants = (latest_war.get("clan") or {}).get("participants", [])
    top_fame = sorted(war_participants, key=lambda p: p.get("fame", 0), reverse=True)[:10]
    top_fame = [{"name": p.get("name"), "tag": p.get("tag", "").replace("#", ""), "value": p.get("fame", 0)} for p in top_fame]

    # Idea #10: fame-per-deck efficiency — separates high-effort-low-skill from
    # low-effort members better than raw fame alone (min 1 deck used to avoid
    # divide-by-zero and to keep 0-deck members out of a "efficiency" ranking).
    fame_efficiency = [
        {
            "name": p.get("name"), "tag": p.get("tag", "").replace("#", ""),
            "value": round(p.get("fame", 0) / p.get("decksUsedToday", 1), 1),
            "fame": p.get("fame", 0), "decks_used": p.get("decksUsedToday", 0),
        }
        for p in war_participants if p.get("decksUsedToday", 0) > 0
    ]
    fame_efficiency.sort(key=lambda x: -x["value"])
    top_fame_efficiency = fame_efficiency[:10]

    # Idea #18: personal-best fame marker — each member's highest-ever single-race
    # fame from war_history, so the leaderboard can show "beat your own record"
    # alongside beating everyone else.
    personal_best_fame = {}
    for race in db_sync["war_history"].find({}, {"data.clan.participants": 1}):
        for p in (race.get("data", {}).get("clan") or {}).get("participants", []):
            tag = p.get("tag")
            fame = p.get("fame", 0)
            if tag and fame > personal_best_fame.get(tag, 0):
                personal_best_fame[tag] = fame
    for row in top_fame:
        row["personal_best"] = personal_best_fame.get(f"#{row['tag']}", row["value"])
        row["is_new_personal_best"] = row["value"] >= row["personal_best"]

    result = {
        "top_donators": top_donators,
        "top_climbers": top_climbers,
        "top_win_rate": top_win_rate,
        "top_war_fame": top_fame,
        "top_fame_efficiency": top_fame_efficiency,
    }
    _record_query_timing("leaderboards", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/cw2-trend")
def admin_analytics_cw2_trend():
    """Idea #11: Clan Wars 2 weekly score chart with a moving average, mirroring
    RoyaleAPI's own CW2 visualization — built from war_history (no new CR API
    calls), one point per completed race section."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    races = list(db_sync["war_history"].find({}, {"_id": 0}).sort("data.seasonId", 1))
    points = []
    for race in races:
        data = race.get("data", {})
        clan = data.get("clan", {})
        points.append({
            "unique_war_id": race.get("unique_war_id"),
            "season_id": data.get("seasonId"),
            "section_index": data.get("sectionIndex"),
            "fame": clan.get("fame", 0),
        })
    # 3-point moving average, matching RoyaleAPI's "moving average line" pattern.
    window = 3
    for i, p in enumerate(points):
        window_vals = [points[j]["fame"] for j in range(max(0, i - window + 1), i + 1)]
        p["moving_avg"] = round(sum(window_vals) / len(window_vals), 1)
    return jsonify({"points": points})


@web_bp.route("/admin/api/analytics/clan-trend")
def admin_analytics_clan_trend():
    """Clan-wide trophy/member-count time series for the trend chart (from clan_snapshots)."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    days = min(int(request.args.get("days", 30)), 180)
    import datetime as _dt
    cutoff = _dt.datetime.now(timezone.utc) - _dt.timedelta(days=days)
    docs = list(
        db_sync["clan_snapshots"]
        .find({"timestamp": {"$gte": cutoff}}, {"_id": 0, "timestamp": 1, "clanScore": 1, "memberCount": 1})
        .sort("timestamp", 1)
    )
    for d in docs:
        d["timestamp"] = d["timestamp"].isoformat() if hasattr(d["timestamp"], "isoformat") else d["timestamp"]
    return jsonify(docs)


@web_bp.route("/admin/api/analytics/member/<tag>/trend")
def admin_analytics_member_trend(tag):
    """Per-member trophy/donation time series for the member drill-down chart."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    days = min(int(request.args.get("days", 30)), 180)
    import datetime as _dt
    cutoff_date = (_dt.datetime.now(timezone.utc) - _dt.timedelta(days=days)).strftime("%Y-%m-%d")
    docs = list(
        db_sync["player_snapshots"]
        .find({"tag": f"#{clean_tag(tag)}", "date": {"$gte": cutoff_date}}, {"_id": 0})
        .sort("date", 1)
    )
    return jsonify(docs)


@web_bp.route("/admin/api/analytics/battles")
def admin_analytics_battles():
    """Clan-wide deck/card win-rate analysis from logged battles."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403

    # Idea #236: scans 2000 battle_history docs every call — cache briefly.
    cache_key = _analytics_cache_key("battles")
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()

    recent = list(
        db_sync["battle_history"]
        .find({}, {"_id": 0, "team_cards": 1, "result": 1, "battle_type": 1})
        .sort("battle_time", -1)
        .limit(2000)
    )
    card_stats = {}
    type_counts = {}
    total_wins = total_games = 0
    for b in recent:
        result = b.get("result")
        if result in ("win", "loss"):
            total_games += 1
            if result == "win": total_wins += 1
        btype = b.get("battle_type", "unknown")
        type_counts[btype] = type_counts.get(btype, 0) + 1
        if result not in ("win", "loss"):
            continue
        for card in _normalize_card_names(b.get("team_cards")):
            entry = card_stats.setdefault(card, {"wins": 0, "games": 0})
            entry["games"] += 1
            if result == "win": entry["wins"] += 1

    card_leaderboard = [
        {"card": c, "games": v["games"], "win_rate": round((v["wins"] / v["games"]) * 100, 1)}
        for c, v in card_stats.items() if v["games"] >= 5
    ]
    card_leaderboard.sort(key=lambda x: x["win_rate"], reverse=True)

    result = {
        "overall_win_rate": round((total_wins / total_games) * 100, 1) if total_games else 0,
        "sample_size": total_games,
        "battle_type_breakdown": type_counts,
        "top_cards": card_leaderboard[:15],
        "worst_cards": card_leaderboard[-15:][::-1] if len(card_leaderboard) > 15 else [],
    }
    _record_query_timing("battles", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/archetypes")
def admin_analytics_archetypes():
    """Deck archetype win-rate analysis — the 'battle algorithm' view. Card-level
    win rate (see admin_analytics_battles above) answers 'is this card good';
    this answers 'which whole decks are actually winning', by grouping battles
    on the sorted 8-card signature actually played rather than individual cards.
    No new Clash Royale API calls — reads battle_history the harvester already
    collected, same as the card win-rate endpoint above.

    Extended for section 14 (ideas #221/#222/#224/#227/#230): each archetype
    now also carries a human-readable name, an elixir curve, a 7-day usage
    trend, and a ready-to-share deck-copy deep link; the response as a whole
    also carries a clan-wide deck-diversity score.
    """
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    min_games = max(int(request.args.get("min_games", 3)), 1)

    # Idea #236: scans 3000 battle_history docs every call (and is itself
    # re-invoked in-process by the meta-snapshot GET route below) — cache by
    # min_games since that param changes the qualifying-archetype cutoff.
    cache_key = _analytics_cache_key("archetypes", min_games)
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()

    recent = list(
        db_sync["battle_history"]
        .find({}, {"_id": 0, "team_cards": 1, "result": 1, "player_tag": 1, "battle_time": 1})
        .sort("battle_time", -1)
        .limit(3000)
    )

    now = datetime.now(timezone.utc)

    def _battle_dt(bt):
        if not bt:
            return None
        try:
            return datetime.strptime(str(bt)[:15], "%Y%m%dT%H%M%S").replace(tzinfo=timezone.utc)
        except ValueError:
            return None

    archetypes = {}
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        cards = _normalize_card_names(b.get("team_cards"))
        if len(cards) < 8:
            continue  # incomplete deck record — skip rather than mis-group
        signature = tuple(sorted(cards[:8]))
        entry = archetypes.setdefault(signature, {
            "wins": 0, "games": 0, "players": set(), "last_seen": None,
            "games_last_7d": 0, "games_prior_7d": 0,
        })
        entry["games"] += 1
        if result == "win":
            entry["wins"] += 1
        if b.get("player_tag"):
            entry["players"].add(b["player_tag"])
        bt = b.get("battle_time")
        if bt and (entry["last_seen"] is None or bt > entry["last_seen"]):
            entry["last_seen"] = bt
        # Idea #224: 7-day usage trend.
        parsed = _battle_dt(bt)
        if parsed:
            age_days = (now - parsed).days
            if age_days <= 7:
                entry["games_last_7d"] += 1
            elif age_days <= 14:
                entry["games_prior_7d"] += 1

    scored = []
    for sig, v in archetypes.items():
        if v["games"] < min_games:
            continue
        cards_list = list(sig)
        trend = "flat"
        if v["games_prior_7d"] == 0 and v["games_last_7d"] > 0:
            trend = "up"
        elif v["games_prior_7d"] > 0:
            delta = (v["games_last_7d"] - v["games_prior_7d"]) / v["games_prior_7d"]
            trend = "up" if delta >= 0.25 else "down" if delta <= -0.25 else "flat"
        scored.append({
            "cards": cards_list,
            "name": _archetype_name(cards_list),  # idea #222
            "games": v["games"],
            "wins": v["wins"],
            "win_rate": round((v["wins"] / v["games"]) * 100, 1),
            "unique_players": len(v["players"]),
            "last_seen": v["last_seen"],
            "trend": trend,  # idea #224
            "elixir_curve": _elixir_curve(cards_list),  # idea #221
            "copy_deck_url": "https://link.clashroyale.com/en/?clashroyale://copyDeck?deck="
                              + ";".join(cards_list),  # idea #227
        })

    # Idea #230: a clan-wide deck-diversity score — 1.0 means every logged
    # game used a different archetype (fully diverse), 0 means everyone plays
    # the exact same deck. Computed as 1 minus the usage share of the single
    # most-played qualifying archetype.
    diversity_score = 1.0
    if scored:
        total_games = sum(a["games"] for a in scored)
        top_share = max(a["games"] for a in scored) / total_games if total_games else 0
        diversity_score = round(1 - top_share, 2)

    result = {
        "sample_size": len(recent),
        "distinct_archetypes": len(archetypes),
        "qualifying_archetypes": len(scored),
        "top_by_win_rate": sorted(scored, key=lambda x: (-x["win_rate"], -x["games"]))[:15],
        "top_by_usage": sorted(scored, key=lambda x: -x["games"])[:15],
        "deck_diversity_score": diversity_score,
    }
    _record_query_timing("archetypes", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/counters/<card>")
def admin_card_counters(card):
    """Idea #216: card counter suggestions, clan-wide. See CARD_COUNTERS'
    docstring — a hand-picked, non-exhaustive list rather than a full
    matchup-solver."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    return jsonify({"card": card, "counters": CARD_COUNTERS.get(card, [])})


@web_bp.route("/admin/api/analytics/counter-pick-prep", methods=["POST"])
def admin_counter_pick_prep():
    """Idea #219: a scrim/tournament prep tool — given a list of cards the
    opponent is known to run, suggests counters (from CARD_COUNTERS) and also
    surfaces which of the clan's own logged archetypes have the best win rate
    against battles where the opponent used any of those cards (a loose
    historical approximation, not a guarantee — future opponents don't always
    play like past ones)."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    opponent_cards = [str(c).strip() for c in (data.get("cards") or []) if str(c).strip()]
    if not opponent_cards:
        return jsonify({"error": "cards list is required"}), 400

    suggested_counters = {}
    for card in opponent_cards:
        counters = CARD_COUNTERS.get(card, [])
        if counters:
            suggested_counters[card] = counters

    recent = list(
        db_sync["battle_history"]
        .find({"opponent_cards": {"$in": opponent_cards}}, {"_id": 0, "team_cards": 1, "result": 1})
        .limit(500)
    )
    archetype_vs = {}
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        cards = _normalize_card_names(b.get("team_cards"))
        if len(cards) < 8:
            continue
        sig = tuple(sorted(cards[:8]))
        entry = archetype_vs.setdefault(sig, {"wins": 0, "games": 0})
        entry["games"] += 1
        if result == "win":
            entry["wins"] += 1
    best_archetypes = sorted(
        [
            {"cards": list(sig), "name": _archetype_name(list(sig)), "games": v["games"],
             "win_rate": round((v["wins"] / v["games"]) * 100, 1)}
            for sig, v in archetype_vs.items() if v["games"] >= 2
        ],
        key=lambda x: (-x["win_rate"], -x["games"]),
    )[:5]

    return jsonify({
        "opponent_cards": opponent_cards,
        "suggested_counters": suggested_counters,
        "sample_size": len(recent),
        "best_performing_decks_against_similar_opponents": best_archetypes,
    })


@web_bp.route("/admin/api/analytics/best-pilots/<card>")
def admin_best_pilots(card):
    """Idea #220: which clan member has the best win rate playing a specific
    card, gamifying card mastery beyond the individual player insights page."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    min_games = max(int(request.args.get("min_games", 5)), 1)
    recent = list(
        db_sync["battle_history"]
        .find({"team_cards": card}, {"_id": 0, "player_tag": 1, "result": 1})
        .limit(3000)
    )
    stats = {}
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        tag = b.get("player_tag")
        if not tag:
            continue
        entry = stats.setdefault(tag, {"wins": 0, "games": 0})
        entry["games"] += 1
        if result == "win":
            entry["wins"] += 1
    names = {p.get("tag", "").replace("#", "").upper(): p.get("name", "")
             for p in db_sync["player_profiles"].find({}, {"tag": 1, "name": 1})}
    leaderboard = [
        {"tag": t, "name": names.get(t, t), "games": v["games"], "win_rate": round((v["wins"] / v["games"]) * 100, 1)}
        for t, v in stats.items() if v["games"] >= min_games
    ]
    leaderboard.sort(key=lambda x: (-x["win_rate"], -x["games"]))
    return jsonify({"card": card, "pilots": leaderboard[:10]})


@web_bp.route("/admin/api/analytics/hard-counters")
def admin_hard_counters_report():
    """Idea #223: aggregate which opponent cards show up most often in the
    clan's own LOSSES specifically (not just "popular cards overall"), so
    practice focus can target what's actually beating the clan."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    # Idea #236: cache — see note on admin_analytics_leaderboards above.
    cache_key = _analytics_cache_key("hard_counters")
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()
    recent = list(
        db_sync["battle_history"]
        .find({"result": "loss"}, {"_id": 0, "opponent_cards": 1})
        .sort("battle_time", -1)
        .limit(2000)
    )
    counts = {}
    for b in recent:
        for card in _normalize_card_names(b.get("opponent_cards")):
            counts[card] = counts.get(card, 0) + 1
    report = sorted(
        [{"card": c, "loss_appearances": n} for c, n in counts.items()],
        key=lambda x: -x["loss_appearances"],
    )[:15]
    result = {"sample_size": len(recent), "cards_clan_struggles_against": report}
    _record_query_timing("hard_counters", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/tier-list")
def admin_tier_list():
    """Idea #225: a lightweight in-house tier list generated from the clan's
    own battle data (not a global meta site) — since local win rates can
    differ meaningfully from the overall meta depending on this clan's habits
    and skill level. Reuses the same card-level win-rate computation as
    admin_analytics_battles, bucketed into S/A/B/C/D tiers by win rate."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    min_games = max(int(request.args.get("min_games", 5)), 1)
    # Idea #236: cache — see note on admin_analytics_leaderboards above.
    cache_key = _analytics_cache_key("tier_list", min_games)
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()
    recent = list(
        db_sync["battle_history"].find({}, {"_id": 0, "team_cards": 1, "result": 1}).limit(3000)
    )
    card_stats = {}
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        for card in _normalize_card_names(b.get("team_cards")):
            entry = card_stats.setdefault(card, {"wins": 0, "games": 0})
            entry["games"] += 1
            if result == "win":
                entry["wins"] += 1

    def _tier_for(win_rate):
        if win_rate >= 58: return "S"
        if win_rate >= 53: return "A"
        if win_rate >= 48: return "B"
        if win_rate >= 43: return "C"
        return "D"

    tiers = {"S": [], "A": [], "B": [], "C": [], "D": []}
    for card, v in card_stats.items():
        if v["games"] < min_games:
            continue
        win_rate = round((v["wins"] / v["games"]) * 100, 1)
        tiers[_tier_for(win_rate)].append({"card": card, "games": v["games"], "win_rate": win_rate})
    for tier in tiers.values():
        tier.sort(key=lambda x: -x["win_rate"])
    result = {"sample_size": len(recent), "tiers": tiers}
    _record_query_timing("tier_list", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/underused-gems")
def admin_underused_gems():
    """Idea #229: cards with a high win rate but low usage — worth spotlighting
    as "hidden gem" recommendations, using the same underlying data as the
    tier list but inverted to surface the overlooked-but-effective end."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    # Idea #236: cache — see note on admin_analytics_leaderboards above.
    cache_key = _analytics_cache_key("underused_gems")
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()
    recent = list(
        db_sync["battle_history"].find({}, {"_id": 0, "team_cards": 1, "result": 1}).limit(3000)
    )
    card_stats = {}
    total_games = 0
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        total_games += 1
        for card in _normalize_card_names(b.get("team_cards")):
            entry = card_stats.setdefault(card, {"wins": 0, "games": 0})
            entry["games"] += 1
            if result == "win":
                entry["wins"] += 1
    gems = []
    for card, v in card_stats.items():
        if v["games"] < 3:
            continue  # too few games to trust the win rate at all
        usage_pct = round((v["games"] / total_games) * 100, 1) if total_games else 0
        win_rate = round((v["wins"] / v["games"]) * 100, 1)
        if win_rate >= 55 and usage_pct <= 8:
            gems.append({"card": card, "win_rate": win_rate, "usage_pct": usage_pct, "games": v["games"]})
    gems.sort(key=lambda x: -x["win_rate"])
    result = {"underused_gems": gems[:10]}
    _record_query_timing("underused_gems", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/matchup-breakdown")
def admin_matchup_breakdown():
    """Idea #226: win rate split by battle_type (ladder/war/friendly) instead
    of one number that mixes all three — a deck that's great on ladder isn't
    necessarily great in war, and vice versa."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    # Idea #236: cache — see note on admin_analytics_leaderboards above.
    cache_key = _analytics_cache_key("matchup_breakdown")
    if request.args.get("refresh") != "1":
        cached = _analytics_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)
    _t0 = time.monotonic()
    recent = list(
        db_sync["battle_history"].find({}, {"_id": 0, "battle_type": 1, "result": 1}).limit(3000)
    )
    by_type = {}
    for b in recent:
        result = b.get("result")
        if result not in ("win", "loss"):
            continue
        btype = b.get("battle_type") or "unknown"
        entry = by_type.setdefault(btype, {"wins": 0, "games": 0})
        entry["games"] += 1
        if result == "win":
            entry["wins"] += 1
    breakdown = {
        t: {"games": v["games"], "win_rate": round((v["wins"] / v["games"]) * 100, 1)}
        for t, v in by_type.items()
    }
    result = {"sample_size": len(recent), "by_battle_type": breakdown}
    _record_query_timing("matchup_breakdown", (time.monotonic() - _t0) * 1000)
    _analytics_cache_set(cache_key, result)
    return jsonify(result)


@web_bp.route("/admin/api/analytics/meta-snapshot", methods=["GET", "POST"])
def admin_meta_snapshot():
    """Idea #228: a season-start meta snapshot for later before/after
    comparison. POST captures the current archetype table into
    config.meta_snapshots (admin-triggered — there's no clean way to detect
    "season start" automatically from the CR API); GET returns the most
    recent snapshot plus a diff against the current live archetype table."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    if request.method == "POST":
        current = admin_analytics_archetypes().get_json()
        snapshot = {
            "taken_at": datetime.now(timezone.utc),
            "top_by_usage": current.get("top_by_usage", []),
        }
        db_sync["config"].update_one({"_id": "meta_snapshots"}, {"$push": {"snapshots": {"$each": [snapshot], "$slice": -6}}}, upsert=True)
        return jsonify({"success": True})
    doc = db_sync["config"].find_one({"_id": "meta_snapshots"}, {"_id": 0}) or {}
    snapshots = doc.get("snapshots", [])
    if not snapshots:
        return jsonify({"snapshots": [], "diff": []})
    last = snapshots[-1]
    last_by_name = {tuple(a["cards"]): a for a in last.get("top_by_usage", [])}
    current = admin_analytics_archetypes().get_json()
    diff = []
    for a in current.get("top_by_usage", []):
        prev = last_by_name.get(tuple(a["cards"]))
        diff.append({
            "name": a["name"], "current_win_rate": a["win_rate"],
            "previous_win_rate": prev["win_rate"] if prev else None,
            "delta": round(a["win_rate"] - prev["win_rate"], 1) if prev else None,
        })
    return jsonify({"snapshots": snapshots, "diff": diff, "last_taken_at": last.get("taken_at")})


# ---------------------------------------------------------------------------
# 5d. ADMIN OPS TOOLING (250-ideas pass, block 4: items 61-80)
# ---------------------------------------------------------------------------

CHANGELOG_WORTHY_ACTIONS = {"Saved bot settings", "Deployed template", "Rolled back template"}


def log_admin_activity(action: str, target: str | None = None, details: str | None = None):
    """Idea #64/#79: a clan-wide admin activity feed (who did what, when),
    separate from the existing per-player `audit_log` field. Called from every
    state-changing admin route below. Never raises — a logging failure should
    never break the action it's describing."""
    try:
        db_sync["admin_activity_log"].insert_one({
            "admin_discord_id": session.get("discord_id"),
            "admin_name": session.get("discord_name", "Unknown"),
            "action": action,
            "target": target,
            "details": details,
            "timestamp": datetime.now(timezone.utc),
            # Idea #150/#151: IP + user-agent alongside the existing who/what/when,
            # so a hijacked session cookie is at least detectable after the fact
            # (a login from a new IP/UA doing admin actions is a visible anomaly
            # in this log, even without full device-pinning enforcement).
            "ip_address": request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip(),
            "user_agent": request.headers.get("User-Agent", "unknown")[:250],
        })
        # Idea #98: auto-post a changelog line for the handful of actions that
        # actually change what members experience (settings, template deploys).
        if action in CHANGELOG_WORTHY_ACTIONS:
            who = session.get("discord_name", "An admin")
            msg = f"{who} — {action}" + (f" ({target})" if target else "")
            db_sync["pending_actions"].insert_one({
                "kind": "changelog_post", "message": msg,
                "created_at": datetime.now(timezone.utc), "processed": False,
            })
    except Exception as e:
        log.error(f"log_admin_activity failed (non-fatal): {e}")


def notify_user_by_tag(tag: str, message: str, kind: str = "general"):
    """Idea #140: in-dashboard notification bell/inbox — records go here
    whenever something relevant to a specific logged-in member happens
    (strikes, promotions, DMs), independent of whatever Discord delivery
    channel (or lack thereof) is used. Never raises."""
    try:
        user = db_sync["users"].find_one({"cr_tag": f"#{clean_tag(tag)}"})
        if not user or not user.get("discord_id"):
            return
        db_sync["notifications"].insert_one({
            "discord_id": user["discord_id"], "message": message, "kind": kind,
            "created_at": datetime.now(timezone.utc), "read": False,
        })
    except Exception as e:
        log.error(f"notify_user_by_tag failed (non-fatal): {e}")


@web_bp.route("/admin/api/activity-log")
def admin_api_activity_log():
    """Idea #64: clan-wide admin activity feed. Idea #79: pass ?by=<discord_id>
    to scope it to one admin, for spotting accounts that aren't being used."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    query = {}
    by = request.args.get("by")
    if by:
        query["admin_discord_id"] = by
    entries = list(
        db_sync["admin_activity_log"].find(query, {"_id": 0}).sort("timestamp", -1).limit(200)
    )
    # Idea #79: per-admin usage counts, so inactive admin accounts are visible at a glance.
    by_admin = {}
    for e in db_sync["admin_activity_log"].find({}, {"admin_discord_id": 1, "admin_name": 1}):
        key = e.get("admin_discord_id") or "unknown"
        entry = by_admin.setdefault(key, {"name": e.get("admin_name", "Unknown"), "count": 0})
        entry["count"] += 1
    return jsonify({"entries": entries, "by_admin": by_admin})


@web_bp.route("/admin/api/player/unstrike", methods=["POST"])
def admin_player_unstrike():
    """Idea #63: a short undo window for the last strike issued — decrements
    instead of a full audit-trail rewrite, kept simple on purpose."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    tag = clean_tag(data.get("tag", ""))
    db_sync["player_profiles"].update_one(
        {"tag": f"#{tag}"}, {"$inc": {"strikes": -1}}, upsert=True
    )
    log_admin_activity("Undo strike", target=tag)
    return jsonify({"success": True})


@web_bp.route("/admin/api/template/versions/<name>")
def admin_template_versions(name):
    """Idea #67: last 5 deployed versions per template, with rollback — stored
    as a small history array alongside the live content, capped at 5 entries."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    doc = db_sync["config"].find_one({"_id": "html_template_history"}) or {}
    versions = doc.get(name, [])
    return jsonify({"versions": [{"deployed_at": v["deployed_at"], "chars": len(v["html"])} for v in versions]})


@web_bp.route("/admin/api/template/rollback", methods=["POST"])
def admin_template_rollback():
    """Idea #67: roll back to one of the last 5 deployed versions."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    name = data.get("template_name", "")
    index = data.get("index", 0)
    allowed = {"roster", "player", "admin", "link"}
    if name not in allowed:
        return jsonify({"error": f"Unknown template '{name}'"}), 400
    doc = db_sync["config"].find_one({"_id": "html_template_history"}) or {}
    versions = doc.get(name, [])
    try:
        target = versions[index]
    except IndexError:
        return jsonify({"error": "No such version"}), 404
    db_sync["config"].update_one({"_id": "html_templates"}, {"$set": {name: target["html"]}}, upsert=True)
    with _cache_lock:
        _HTML_CACHE.pop(name, None)
    log_admin_activity("Rolled back template", target=name, details=f"to version from {target['deployed_at']}")
    return jsonify({"success": True})


@web_bp.route("/admin/api/settings/schedule", methods=["POST"])
def admin_schedule_settings():
    """Idea #77: schedule a settings change for a future time instead of only
    immediate effect. A lightweight mechanism — the harvester's existing 30-min
    loop (or the settings-reload path) applies any due scheduled change; nothing
    new to run continuously, since the harvester already wakes up on a cadence."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    apply_at = data.get("apply_at")  # ISO datetime string, e.g. "2026-07-21T02:00:00Z"
    changes = data.get("changes", {})
    if not apply_at or not changes:
        return jsonify({"error": "apply_at and changes are required"}), 400
    db_sync["scheduled_settings"].insert_one({
        "apply_at": apply_at, "changes": changes,
        "created_by": session.get("discord_name", "admin"),
        "created_at": datetime.now(timezone.utc), "applied": False,
    })
    log_admin_activity("Scheduled a settings change", details=f"for {apply_at}: {changes}")
    return jsonify({"success": True})


@web_bp.route("/admin/api/settings/scheduled")
def admin_list_scheduled_settings():
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    pending = list(db_sync["scheduled_settings"].find({"applied": False}, {"_id": 0}))
    return jsonify({"scheduled": pending})


@web_bp.route("/admin/api/settings/test-welcome-dm", methods=["POST"])
def admin_test_welcome_dm():
    """Idea #69: "test this setting" — queue the CURRENT welcome message as a DM
    to the admin's own linked Discord account, using the same pending_actions
    delivery path every other DM already uses."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    if "discord_id" not in session:
        return jsonify({"error": "You need a Discord session to receive a test DM."}), 400
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    message = bot_settings.get("welcome_msg", "Welcome to the Squad!")
    db_sync["pending_actions"].insert_one({
        "kind": "dm_warning",  # reuses the existing generic DM delivery kind
        "discord_id": session["discord_id"],
        "message": f"[TEST] {message}",
        "created_at": datetime.now(timezone.utc),
        "processed": False,
    })
    log_admin_activity("Sent test welcome DM to self")
    return jsonify({"success": True, "message": "Test DM queued — check your Discord DMs shortly."})


@web_bp.route("/admin/api/roster/bulk-strike", methods=["POST"])
def admin_roster_bulk_strike():
    """Idea #62: bulk actions — issue a strike to every selected member at once."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    tags = [clean_tag(t) for t in data.get("tags", [])]
    if not tags:
        return jsonify({"error": "tags is required"}), 400
    for tag in tags:
        db_sync["player_profiles"].update_one({"tag": f"#{tag}"}, {"$inc": {"strikes": 1}}, upsert=True)
    log_admin_activity("Bulk strike", details=f"{len(tags)} member(s): {', '.join(tags)}")
    return jsonify({"success": True, "count": len(tags)})


@web_bp.route("/admin/api/roster/bulk-dm", methods=["POST"])
def admin_roster_bulk_dm():
    """Idea #62: bulk DM to every selected member."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    tags = [f"#{clean_tag(t)}" for t in data.get("tags", [])]
    message = data.get("message", "Please remember clan expectations — thanks!")
    if not tags:
        return jsonify({"error": "tags is required"}), 400
    users = list(db_sync["users"].find({"cr_tag": {"$in": tags}}))
    discord_ids = [u["discord_id"] for u in users if u.get("discord_id")]
    if discord_ids:
        db_sync["pending_actions"].insert_one({
            "kind": "war_nudge",  # reuses the existing multi-recipient DM delivery kind
            "discord_ids": discord_ids,
            "created_at": datetime.now(timezone.utc),
            "processed": False,
        })
    log_admin_activity("Bulk DM", details=f"{len(discord_ids)} recipient(s)")
    return jsonify({"success": True, "count": len(discord_ids)})


@web_bp.route("/admin/api/maintenance-status")
def admin_maintenance_status():
    """Idea #70: warn if maintenance mode has been left on for a long time —
    tracks when it was last turned ON via `maintenance_enabled_at`."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    enabled = bool(bot_settings.get("maintenance_mode", False))
    enabled_at = _as_aware_utc(bot_settings.get("maintenance_enabled_at"))
    hours_on = None
    if enabled and enabled_at:
        hours_on = round((datetime.now(timezone.utc) - enabled_at).total_seconds() / 3600, 1)
    return jsonify({"maintenance_mode": enabled, "hours_on": hours_on, "stale_warning": bool(hours_on and hours_on > 6)})


# ---------------------------------------------------------------------------
# 5c. PLAYER-PROFILE ENHANCEMENTS (250-ideas pass, block 2)
# ---------------------------------------------------------------------------

@web_bp.route("/api/player/<tag>/trophy-history")
def api_player_trophy_history(tag):
    """Idea #21: full trophy history series for a public line graph on the
    player page — the admin analytics member-trend chart already renders this
    exact series server-side; this just exposes it without requiring is_admin,
    since it's a player's own already-public trophy count over time."""
    days = min(int(request.args.get("days", 30)), 180)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    docs = list(
        db_sync["player_snapshots"]
        .find({"tag": f"#{clean_tag(tag)}", "date": {"$gte": cutoff}}, {"_id": 0, "date": 1, "trophies": 1})
        .sort("date", 1)
    )
    return jsonify(docs)


@web_bp.route("/api/player/<tag>/season-comparison")
def api_player_season_comparison(tag):
    """Idea #22: this season vs. last season's end-of-season trophies. CR seasons
    reset roughly monthly, so this approximates "last season end" as the trophy
    snapshot closest to 30 days ago — an approximation noted here since exact
    season boundaries aren't stored locally."""
    clean = clean_tag(tag)
    current = fetch_cr_api(f"players/%23{clean}")
    if not current:
        return jsonify({"error": "player not found"}), 404
    month_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    last_season_snap = db_sync["player_snapshots"].find_one(
        {"tag": f"#{clean}", "date": {"$lte": month_ago}}, sort=[("date", -1)]
    )
    return jsonify({
        "current_trophies": current.get("trophies", 0),
        "last_season_trophies": last_season_snap.get("trophies") if last_season_snap else None,
        "note": "Approximated using the closest snapshot ~30 days ago, not an exact CR season boundary.",
    })


@web_bp.route("/api/clan/boat-battle")
def api_clan_boat_battle():
    """Surfaces the war-tracking fields that get harvested into Mongo every
    30-minute cycle (and after every completed river race) but were never read
    back by anything: per-participant `repairPoints`/`boatAttacks` from the
    live `war_tracking` snapshot, and the clan-level `warTrophiesChange` from
    the most recently completed race in `war_history`."""
    latest = db_sync["war_tracking"].find_one({}, sort=[("harvest_time", -1)])
    participants = []
    if latest:
        clan = latest.get("clan") or {}
        for p in clan.get("participants", []):
            boat_attacks = p.get("boatAttacks", 0) or 0
            repair_points = p.get("repairPoints", 0) or 0
            if boat_attacks or repair_points:
                participants.append({
                    "name": p.get("name", "Unknown"),
                    "tag": p.get("tag", ""),
                    "boat_attacks": boat_attacks,
                    "repair_points": repair_points,
                })
    participants.sort(key=lambda p: (p["boat_attacks"], p["repair_points"]), reverse=True)

    last_war_trophies_change = None
    last_race = db_sync["war_history"].find_one(
        {}, sort=[("data.seasonId", -1), ("data.sectionIndex", -1)]
    )
    if last_race:
        race_clan = (last_race.get("data") or {}).get("clan") or {}
        last_war_trophies_change = race_clan.get("warTrophiesChange")

    return jsonify({
        "leaderboard": participants[:10],
        "last_war_trophies_change": last_war_trophies_change,
    })


@web_bp.route("/api/compare/<tag1>/<tag2>")
def api_compare_players(tag1, tag2):
    """Idea #25: head-to-head comparison — pick two clan members, see key stats
    side by side. Uses the same fetch/aggregation path as a normal player page."""
    p1 = get_player_analytical_data(tag1)
    p2 = get_player_analytical_data(tag2)
    if not p1 or not p2:
        return jsonify({"error": "one or both players not found"}), 404
    fields = ["trophies", "bestTrophies", "wins", "losses", "threeCrownWins", "donations",
              "warDayWins", "expLevel", "personal_best_win_streak", "deck_grade"]
    return jsonify({
        "player_1": {f: p1.get(f) for f in fields} | {"name": p1.get("name"), "tag": p1.get("tag")},
        "player_2": {f: p2.get(f) for f in fields} | {"name": p2.get("name"), "tag": p2.get("tag")},
    })


@web_bp.route("/api/player/<tag>/card-mastery")
def api_player_card_mastery(tag):
    """Idea #26: card mastery grouped by category (Troop/Spell/Building) instead
    of one flat "cards maxed" number. Category data isn't in the CR API, so this
    uses the CARD_CATEGORIES lookup defined near the top of this file.

    Also folds in rarity/count/elixirCost -- these ride along on every card in
    the CR API's own `cards` array (same as name/level/maxLevel), but nothing
    ever read them before now."""
    clean = clean_tag(tag)
    player = fetch_cr_api(f"players/%23{clean}")
    if not player:
        return jsonify({"error": "player not found"}), 404
    by_category = {}
    by_rarity = {}
    most_duplicated = None
    total_copies = 0
    elixir_total = 0
    elixir_count = 0
    for c in player.get("cards", []):
        maxed = c.get("level", 0) >= c.get("maxLevel", 999)

        cat = categorize_card(c.get("name", ""))
        entry = by_category.setdefault(cat, {"total": 0, "maxed": 0})
        entry["total"] += 1
        if maxed:
            entry["maxed"] += 1

        rarity = str(c.get("rarity") or "Unknown").title()
        r_entry = by_rarity.setdefault(rarity, {"total": 0, "maxed": 0})
        r_entry["total"] += 1
        if maxed:
            r_entry["maxed"] += 1

        count = c.get("count", 0) or 0
        total_copies += count
        if count and (most_duplicated is None or count > most_duplicated["count"]):
            most_duplicated = {"name": c.get("name", "Unknown"), "count": count}

        elixir = c.get("elixirCost")
        if isinstance(elixir, (int, float)):
            elixir_total += elixir
            elixir_count += 1

    return jsonify({
        "by_category": {cat: v for cat, v in by_category.items()},
        "by_rarity": by_rarity,
        "most_duplicated_card": most_duplicated,
        "total_card_copies": total_copies,
        "avg_elixir_cost": round(elixir_total / elixir_count, 2) if elixir_count else None,
    })


@web_bp.route("/api/player/<tag>/flair", methods=["GET", "POST"])
def api_player_flair(tag):
    """Idea #27: a lightweight self-set "personality" field (motto + favourite
    card note) for a more social profile page. Only the linked Discord account
    for this tag (or an admin) may edit it — anyone can read it, since it's meant
    to be public flair, not private data."""
    clean = clean_tag(tag)
    if request.method == "GET":
        doc = db_sync["player_profiles"].find_one({"tag": f"#{clean}"}, {"flair_motto": 1, "flair_note": 1, "custom_badge": 1}) or {}
        return jsonify({"motto": doc.get("flair_motto", ""), "note": doc.get("flair_note", ""), "custom_badge": doc.get("custom_badge", "")})

    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    user = db_sync["users"].find_one({"discord_id": session["discord_id"]})
    owns_tag = user and clean_tag(user.get("cr_tag", "")) == clean
    if not (owns_tag or is_admin()):
        return jsonify({"error": "You can only edit your own profile flair."}), 403

    data = request.get_json(silent=True) or {}
    db_sync["player_profiles"].update_one(
        {"tag": f"#{clean}"},
        {"$set": {
            "flair_motto": str(data.get("motto", ""))[:80],
            "flair_note": str(data.get("note", ""))[:120],
        }},
        upsert=True,
    )
    return jsonify({"success": True})


@web_bp.route("/api/player/<tag>/activity-heatmap")
def api_player_activity_heatmap(tag):
    """Idea #32: which hours/days this player battles most, from battle_time
    already logged in battle_history. Returns a day-of-week x hour-of-day grid."""
    clean = clean_tag(tag)
    battles = db_sync["battle_history"].find({"player_tag": clean}, {"battle_time": 1})
    grid = {}
    for b in battles:
        raw = b.get("battle_time", "")
        # CR API format: 20240312T183045.000Z
        try:
            dt = datetime.strptime(raw[:15], "%Y%m%dT%H%M%S")
        except ValueError:
            continue
        key = f"{dt.weekday()}_{dt.hour}"
        grid[key] = grid.get(key, 0) + 1
    return jsonify({"grid": grid, "note": "key is 'weekday(0=Mon)_hour(0-23 UTC)'"})


@web_bp.route("/player/<tag>/card.png")
def player_card_png(tag):
    """Idea #29: a downloadable/shareable player-card image, Spotify-Wrapped
    style, for social bragging rights. Renders server-side with Pillow — no
    browser screenshot tooling needed. 404s with a plain-text message if Pillow
    isn't installed rather than crashing the whole app."""
    if not _PIL_AVAILABLE:
        return "Player card image generation requires Pillow (see requirements.txt).", 501
    clean = clean_tag(tag)
    player = fetch_cr_api(f"players/%23{clean}")
    if not player:
        return "Player not found", 404

    W, H = 600, 340
    bg = (11, 12, 16)       # matches --bg / --surface dark theme
    accent = (0, 229, 255)  # matches --accent
    text_dim = (136, 136, 136)
    img = Image.new("RGB", (W, H), bg)
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, W - 1, H - 1], outline=accent, width=2)
    try:
        font_big = ImageFont.load_default(size=28)
        font_small = ImageFont.load_default(size=16)
    except TypeError:
        # Older Pillow without the `size` kwarg on load_default.
        font_big = font_small = ImageFont.load_default()

    draw.text((30, 30), player.get("name", "Unknown"), font=font_big, fill=(255, 255, 255))
    draw.text((30, 70), player.get("tag", ""), font=font_small, fill=text_dim)
    lines = [
        f"Trophies: {player.get('trophies', 0)}  (Best: {player.get('bestTrophies', 0)})",
        f"Wins: {player.get('wins', 0)}  Losses: {player.get('losses', 0)}",
        f"3-Crown Wins: {player.get('threeCrownWins', 0)}",
        f"Level {player.get('expLevel', 0)}",
        "graveyardbot.onrender.com",
    ]
    y = 120
    for line in lines:
        draw.text((30, y), line, font=font_small, fill=accent if line.startswith("Trophies") else (200, 200, 200))
        y += 32

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")


@web_bp.route("/player/<tag>/qr.png")
def player_qr_png(tag):
    """Idea #40: QR code encoding a shareable link straight to this player's
    profile, for quick mobile sharing to Discord."""
    if not _QRCODE_AVAILABLE:
        return "QR code generation requires the qrcode package (see requirements.txt).", 501
    clean = clean_tag(tag)
    url = f"https://graveyardbot.onrender.com/player/{clean}"
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")


@web_bp.route("/admin/api/player/<tag>/global-rank")
def admin_player_global_rank(tag):
    """Idea #30: global rank badge. This IS a new live CR API call (the global
    leaderboard endpoint), so admin-gated per this project's existing rule for
    any new per-view external call."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    clean = clean_tag(tag)
    # CR API's global player leaderboard; location id 'global' is the documented
    # special value for the worldwide ranking list.
    leaderboard = fetch_cr_api("locations/global/rankings/players?limit=200")
    if not leaderboard:
        return jsonify({"global_rank": None, "note": "Leaderboard fetch failed or player outside top 200."})
    items = leaderboard.get("items", [])
    match = next((i for i in items if i.get("tag", "").replace("#", "") == clean), None)
    return jsonify({"global_rank": match.get("rank") if match else None})


@web_bp.route("/admin/api/player/<tag>/chests")
def admin_player_chests(tag):
    """Idea #31: season pass / chest cycle tracker. Uses the CR API's own
    upcomingchests endpoint — a new live call per view, so admin-gated per the
    existing project convention (same treatment as global rank above)."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    clean = clean_tag(tag)
    data = fetch_cr_api(f"players/%23{clean}/upcomingchests")
    if not data:
        return jsonify({"error": "Could not fetch chest cycle data."})
    return jsonify(data)


# ---------------------------------------------------------------------------
# 6. FUTURE STUBS
# ---------------------------------------------------------------------------
@web_bp.route("/api/lfg", methods=["POST"])
def look_for_group():
    if "discord_id" not in session: return jsonify({"error": "Unauthorized"}), 403
    db_sync["pending_actions"].insert_one({
        "kind": "lfg_ping",
        "user": session.get("discord_name", "A clan member"),
        "created_at": datetime.now(timezone.utc),
        "processed": False,
    })
    return jsonify({"message": "Ping queued."})

@web_bp.route("/api/predict/matchmaking/<tag>")
def predict_matchmaking(tag): return jsonify({"message": "Analysis pending."})


# ---------------------------------------------------------------------------
# 8. ONBOARDING & MEMBER EXPERIENCE (250-ideas section 7, items 116-130)
# ---------------------------------------------------------------------------

@web_bp.route("/api/clan/active-now")
def api_clan_active_now():
    """Idea #119: "who's currently active" social-proof indicator for the
    roster page. The CR API doesn't expose real online-status for a member
    list, so this approximates activity via logged battles in the last 24h —
    documented approximation, same pattern as this project's others."""
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).strftime("%Y%m%d")
    active_tags = db_sync["battle_history"].distinct("player_tag", {"battle_time": {"$gte": cutoff}})
    return jsonify({"active_last_24h": len(active_tags)})


@web_bp.route("/how-it-works")
def how_it_works():
    """Idea #124: a public 'how the clan works' page (requirements, war
    schedule, culture, channel structure) as its own route, instead of that
    information only living in the /link page's sidebar."""
    return render_sandboxed(get_template("how_it_works"), is_admin=is_admin())


@web_bp.route("/clan-card.png")
def clan_card_png():
    """Idea #189: an auto-generated, print/share-optimized "clan card" banner
    (name, tag, member count, average trophies) for recruiting posts outside
    Discord. Same server-side Pillow approach as the existing player card.png
    route (idea #29) — 501s with a plain message if Pillow isn't installed
    rather than crashing the app."""
    if not _PIL_AVAILABLE:
        return "Clan card image generation requires Pillow (see requirements.txt).", 501
    clan_data, _ = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    members = clan_data.get("memberList", []) or []
    member_count = clan_data.get("memberCount", len(members))
    avg_trophies = round(sum(m.get("trophies", 0) for m in members) / len(members)) if members else 0
    clan_name = clan_data.get("name", "Graveyard Squad")

    W, H = 1000, 420
    bg = (11, 12, 16)        # matches --gy-bg
    accent = (0, 229, 255)   # matches --gy-accent
    ok = (0, 224, 150)       # matches --gy-ok
    text_dim = (136, 136, 136)
    text_bright = (240, 240, 240)  # matches --gy-text-bright (idea #171)
    img = Image.new("RGB", (W, H), bg)
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, W - 1, H - 1], outline=accent, width=3)
    try:
        font_huge = ImageFont.load_default(size=54)
        font_big = ImageFont.load_default(size=26)
        font_small = ImageFont.load_default(size=18)
    except TypeError:
        font_huge = font_big = font_small = ImageFont.load_default()

    draw.text((40, 40), "☠", font=font_huge, fill=accent)
    draw.text((110, 50), clan_name, font=font_huge, fill=text_bright)
    draw.text((40, 130), f"#{CLAN_TAG}", font=font_big, fill=text_dim)

    stats = [
        (f"{member_count}/50", "Members"),
        (f"{avg_trophies:,}", "Avg. Trophies"),
        (clan_data.get("clanWarTrophies", "—"), "War Trophies"),
    ]
    x = 40
    for value, label in stats:
        draw.text((x, 210), str(value), font=font_huge, fill=ok)
        draw.text((x, 280), label, font=font_small, fill=text_dim)
        x += 320

    draw.text((40, H - 60), "graveyardbot.onrender.com  •  Recruiting now", font=font_small, fill=accent)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")


@web_bp.route("/admin/api/mentor", methods=["GET", "POST"])
def admin_mentor_pairs():
    """Idea #125: buddy/mentor assignment for brand-new members. Manual,
    admin-assigned pairing (not an automatic matcher) — same reasoning as the
    user's explicit preference for a manual beta-role sync over an automatic
    one: a person picking the mentor produces a better match than a heuristic."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        mentee_tag = clean_tag(data.get("mentee_tag", ""))
        mentor_tag = clean_tag(data.get("mentor_tag", ""))
        if not mentee_tag or not mentor_tag:
            return jsonify({"error": "mentee_tag and mentor_tag are required"}), 400
        mentee = fetch_cr_api(f"players/%23{mentee_tag}") or {}
        mentor = fetch_cr_api(f"players/%23{mentor_tag}") or {}
        db_sync["mentor_pairs"].update_many({"mentee_tag": f"#{mentee_tag}", "active": True}, {"$set": {"active": False}})
        db_sync["mentor_pairs"].insert_one({
            "mentee_tag": f"#{mentee_tag}", "mentee_name": mentee.get("name", ""),
            "mentor_tag": f"#{mentor_tag}", "mentor_name": mentor.get("name", ""),
            "assigned_at": datetime.now(timezone.utc), "active": True,
        })
        log_admin_activity("Assigned mentor", target=mentee_tag, details=f"mentor={mentor_tag}")
        return jsonify({"success": True})
    pairs = list(db_sync["mentor_pairs"].find({"active": True}, {"_id": 0}).sort("assigned_at", -1))
    return jsonify({"pairs": pairs})


@web_bp.route("/admin/api/mentor/<mentee_tag>", methods=["DELETE"])
def admin_mentor_unassign(mentee_tag):
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    clean = clean_tag(mentee_tag)
    db_sync["mentor_pairs"].update_many({"mentee_tag": f"#{clean}", "active": True}, {"$set": {"active": False}})
    log_admin_activity("Removed mentor pairing", target=clean)
    return jsonify({"success": True})


@web_bp.route("/api/onboarding", methods=["GET", "POST"])
def api_onboarding():
    """Ideas #116 (skippable 3-step onboarding) + #117 (preference-based
    tailoring). GET returns current progress/state; POST records step
    completion (interests + notification prefs) and can mark the whole flow
    done (including an explicit "skip"), so it only ever nags a member once."""
    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    discord_id = session["discord_id"]
    if request.method == "GET":
        user = db_sync["users"].find_one({"discord_id": discord_id}) or {}
        return jsonify({
            "onboarding_completed": bool(user.get("onboarding_completed", False)),
            "is_linked": bool(user.get("cr_tag")),
            "interest_tags": user.get("interest_tags", []),
            "notif_prefs": user.get("notif_prefs", {"war_reminders": True}),
            "email": user.get("email", ""),
        })
    data = request.get_json(silent=True) or {}
    update = {}
    if "interest_tags" in data:
        allowed = {"war", "social", "competitive"}
        update["interest_tags"] = [t for t in data.get("interest_tags", []) if t in allowed]
    if "notif_prefs" in data:
        update["notif_prefs"] = {"war_reminders": bool((data.get("notif_prefs") or {}).get("war_reminders", True))}
    # Idea #138: optional email for the digest-email option (delivery itself is
    # stubbed — see clash_cog.py's weekly_digest_email handler — until real SMTP
    # credentials are configured for this project).
    if "email" in data:
        email = str(data.get("email") or "").strip()[:200]
        update["email"] = email if "@" in email else ""
    if data.get("complete") or data.get("skip"):
        update["onboarding_completed"] = True
    if update:
        db_sync["users"].update_one({"discord_id": discord_id}, {"$set": update}, upsert=True)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# 9. NOTIFICATIONS & REMINDERS (250-ideas section 8, items 131-145)
# ---------------------------------------------------------------------------

@web_bp.route("/api/notifications")
def api_notifications():
    """Idea #140: in-dashboard notification bell/inbox — strikes, promotions,
    and DMs relevant to the logged-in member, independent of Discord delivery."""
    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    notifs = list(db_sync["notifications"].find(
        {"discord_id": session["discord_id"]}, {"_id": 0}
    ).sort("created_at", -1).limit(20))
    unread_count = db_sync["notifications"].count_documents({"discord_id": session["discord_id"], "read": False})
    return jsonify({"notifications": notifs, "unread_count": unread_count})


@web_bp.route("/api/notifications/mark-read", methods=["POST"])
def api_notifications_mark_read():
    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    db_sync["notifications"].update_many({"discord_id": session["discord_id"], "read": False}, {"$set": {"read": True}})
    return jsonify({"success": True})


@web_bp.route("/api/notifications/snooze", methods=["POST"])
def api_notifications_snooze():
    """Idea #141: snooze automated war-nudge DMs for a few hours instead of
    only silence-until-next-scheduled-ping. Read by data_harvester.py's
    _eligible_discord_ids() before queuing the next nudge."""
    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    hours = max(1, min(24, int(data.get("hours", 4))))
    snoozed_until = datetime.now(timezone.utc) + timedelta(hours=hours)
    db_sync["users"].update_one({"discord_id": session["discord_id"]}, {"$set": {"snoozed_until": snoozed_until}}, upsert=True)
    return jsonify({"success": True, "snoozed_until": snoozed_until.isoformat()})


@web_bp.route("/calendar.ics")
def calendar_ics():
    """Idea #142: an ICS export of recurring war days for members who like
    external calendar reminders. War days are Thursday-Sunday per the clan's
    stated schedule (documented in how_it_works.html) — a recurring weekly
    all-day event, not tied to any single season's exact CR API timestamps."""
    from flask import Response
    now = datetime.now(timezone.utc)
    # Next Thursday (weekday 3), used as the RRULE start anchor.
    days_until_thursday = (3 - now.weekday()) % 7
    start = (now + timedelta(days=days_until_thursday)).strftime("%Y%m%d")
    ics = (
        "BEGIN:VCALENDAR\r\n"
        "VERSION:2.0\r\n"
        "PRODID:-//Graveyard Squad//War Schedule//EN\r\n"
        "BEGIN:VEVENT\r\n"
        f"UID:graveyard-squad-war-days@dashboard\r\n"
        f"DTSTART;VALUE=DATE:{start}\r\n"
        "SUMMARY:Graveyard Squad War Day\r\n"
        "DESCRIPTION:Clash Royale River Race war day — use all 4 decks!\r\n"
        "RRULE:FREQ=WEEKLY;COUNT=4\r\n"
        "END:VEVENT\r\n"
        "END:VCALENDAR\r\n"
    )
    return Response(ics, mimetype="text/calendar", headers={"Content-Disposition": "attachment; filename=graveyard-squad-war-days.ics"})


@web_bp.route("/calendar/google")
def calendar_google_add():
    """Idea #241: extends the existing ICS export (idea #142, calendar_ics()
    above) with a one-click "Add to Google Calendar" link — some members find
    downloading+importing an .ics file more friction than a direct add-link.
    Computes the same next-Thursday/weekly-recurring war day and redirects to
    Google Calendar's render endpoint with the event pre-filled; Google adds it
    to the signed-in user's own calendar client-side, no OAuth/API key needed
    on our end. Apple Calendar / other clients still use the .ics download —
    they don't have an equivalent unauthenticated add-link format."""
    now = datetime.now(timezone.utc)
    days_until_thursday = (3 - now.weekday()) % 7
    start_date = (now + timedelta(days=days_until_thursday)).strftime("%Y%m%d")
    end_date = (now + timedelta(days=days_until_thursday + 1)).strftime("%Y%m%d")
    params = {
        "action": "TEMPLATE",
        "text": "Graveyard Squad War Day",
        "dates": f"{start_date}/{end_date}",
        "details": "Clash Royale River Race war day — use all 4 decks! (Graveyard Squad)",
        "recur": "RRULE:FREQ=WEEKLY;COUNT=4",
    }
    from urllib.parse import urlencode
    return redirect("https://calendar.google.com/calendar/render?" + urlencode(params))


WEBHOOK_EVENTS = ["war_end", "harvest_complete"]

@web_bp.route("/admin/api/webhooks", methods=["GET", "POST"])
def admin_webhooks():
    """Idea #242/#246: outbound webhook subscriptions — see
    DataHarvester._fire_webhooks() in data_harvester.py for the dispatch side.
    A subscription is just {url, events: [...]}; any tool that can receive a
    POST works, including Zapier's "Webhooks by Zapier" trigger or a
    Make.com "Custom webhook" module, so idea #246 (Zapier/Make hooks) is
    covered by the same mechanism rather than a separate integration."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    col = db_sync["webhooks"]
    if request.method == "GET":
        hooks = list(col.find({}))
        for h in hooks:
            h["id"] = str(h.pop("_id"))
        return jsonify({"webhooks": hooks, "available_events": WEBHOOK_EVENTS})

    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    events = [e for e in (data.get("events") or []) if e in WEBHOOK_EVENTS]
    if not url or not url.startswith(("http://", "https://")):
        return jsonify({"error": "A valid http(s) URL is required."}), 400
    if not events:
        return jsonify({"error": "Select at least one event to subscribe to."}), 400
    doc = {"url": url, "events": events, "created_at": datetime.now(timezone.utc)}
    result = col.insert_one(doc)
    log_admin_activity("Added webhook subscription", target=url, details=", ".join(events))
    return jsonify({"success": True, "id": str(result.inserted_id)})


@web_bp.route("/admin/api/webhooks/<hook_id>", methods=["DELETE"])
def admin_delete_webhook(hook_id):
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    from bson import ObjectId
    try:
        oid = ObjectId(hook_id)
    except Exception:
        return jsonify({"error": "Invalid webhook id."}), 400
    db_sync["webhooks"].delete_one({"_id": oid})
    log_admin_activity("Removed webhook subscription", target=hook_id)
    return jsonify({"success": True})


@web_bp.route("/api/v1/clan")
def public_api_clan():
    """Idea #244: a public, read-only, rate-limited API exposing the same
    aggregated stats the public roster page (/) already shows in HTML, for
    members who'd rather build their own tool/bot against JSON than scrape
    the page. No auth required — everything returned here is already visible
    to any logged-out visitor on the roster page. Rate-limited by IP (60
    requests / 10 minutes, same `rate_limited()` helper idea #147 added for
    /login and /link) so this can't be used to indirectly hammer the CR API
    past fetch_cr_api()'s own cache. CORS is wide open since this is meant to
    be called from arbitrary third-party frontends/scripts."""
    if rate_limited("public_api", max_attempts=60, window_seconds=600):
        return jsonify({"error": "Rate limit exceeded. Max 60 requests per 10 minutes."}), 429
    clan_data, clan_data_is_stale = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    members = clan_data.get("memberList", [])
    resp = jsonify({
        "name": clan_data.get("name"),
        "tag": clan_data.get("tag"),
        "description": clan_data.get("description"),
        "clan_score": clan_data.get("clanScore"),
        "member_count": len(members),
        "required_trophies": clan_data.get("requiredTrophies"),
        "clan_war_trophies": clan_data.get("clanWarTrophies"),
        "is_stale": clan_data_is_stale,
        "members": [
            {
                "name": m.get("name"),
                "role": m.get("role"),
                "trophies": m.get("trophies"),
                "donations": m.get("donations"),
                "donations_received": m.get("donationsReceived"),
                "last_seen": m.get("lastSeen"),
            }
            for m in members
        ],
    })
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


@web_bp.route("/api/v1/war-history")
def public_api_war_history():
    """Idea #244 continued: recent war outcomes (win/loss/rank only — no
    per-member fame or Discord data, which stays admin-only), same rate limit
    bucket as public_api_clan above."""
    if rate_limited("public_api", max_attempts=60, window_seconds=600):
        return jsonify({"error": "Rate limit exceeded. Max 60 requests per 10 minutes."}), 429
    own_tag = f"#{CLAN_TAG}"
    races = list(db_sync["war_history"].find({}, {"_id": 0}).sort("unique_war_id", -1).limit(10))
    out = []
    for race in races:
        data = race.get("data", {})
        standings = sorted(data.get("standings", []) or [], key=lambda s: s.get("rank", 99))
        own_rank = next((s.get("rank") for s in standings if (s.get("clan") or {}).get("tag") == own_tag), None)
        out.append({
            "unique_war_id": race.get("unique_war_id"),
            "season_id": data.get("seasonId"),
            "section_index": data.get("sectionIndex"),
            "rank": own_rank,
            "outcome": "win" if own_rank == 1 else ("loss" if own_rank else "unknown"),
        })
    resp = jsonify({"races": out})
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


@web_bp.route("/admin/api/config-backups")
def admin_config_backups():
    """Idea #152: visibility into the daily config-collection backups
    (data_harvester.py's backup_config_collection())."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    # Bugfix: the projection excluded "docs" but not "_id" — Mongo's ObjectId
    # isn't JSON-serializable, so jsonify() 500'd the moment a real backup
    # document existed (empty results serialize fine as `[]`, which is why
    # this went unnoticed until the daily backup job actually produced one).
    backups = list(db_sync["config_backups"].find({}, {"docs": 0}).sort("created_at", -1).limit(20))
    for b in backups:
        b["id"] = str(b.pop("_id"))
    return jsonify({"backups": backups})


@web_bp.route("/admin/api/settings/secret-rotation", methods=["GET", "POST"])
def admin_secret_rotation():
    """Idea #160: track "last rotated" dates for the Discord client secret and
    CR API token, surfaced as a Diagnostics checklist item — this only records
    dates an admin confirms; it can't detect an actual rotation automatically
    since those secrets live in environment variables, not Mongo."""
    if not is_admin(): return jsonify({"error": "unauthorized"}), 403
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    log_data = bot_settings.get("secret_rotation_log", {})
    if request.method == "GET":
        return jsonify({"secret_rotation_log": log_data})
    data = request.get_json(silent=True) or {}
    secret_name = data.get("secret_name")
    if secret_name not in ("discord_client_secret", "cr_token", "flask_secret"):
        return jsonify({"error": "Unknown secret_name."}), 400
    log_data[secret_name] = datetime.now(timezone.utc).isoformat()
    db_sync["config"].update_one({"_id": "bot_settings"}, {"$set": {"secret_rotation_log": log_data}}, upsert=True)
    log_admin_activity("Marked secret as rotated", target=secret_name)
    return jsonify({"success": True})


@web_bp.route("/api/status")
def api_status():
    """Idea #156: graceful clanwide degradation messaging — the Diagnostics tab
    already detects CR API rate-limiting for admins; this surfaces the same
    signal on the public roster page so members see a clear "data may be
    stale" note instead of silently wondering why the roster looks empty."""
    health = db_sync["config"].find_one({"_id": "cr_api_health"}, {"_id": 0}) or {}
    last_rate_limited = _as_aware_utc(health.get("last_rate_limited_at"))
    recently_degraded = bool(last_rate_limited and (datetime.now(timezone.utc) - last_rate_limited) < timedelta(minutes=15))
    return jsonify({"cr_api_degraded": recently_degraded})


@web_bp.route("/api/war/countdown")
def api_war_countdown():
    """Idea #145: "next war starts in..." countdown for the roster header.
    Same Thu-Sun approximation as the ICS export and check_war_start_reminders()
    — the CR API doesn't expose an exact next-war-start timestamp."""
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    reset_hour = int(bot_settings.get("war_reset_hour_utc", 10))
    now = datetime.now(timezone.utc)
    days_until_thursday = (3 - now.weekday()) % 7
    next_war = (now + timedelta(days=days_until_thursday)).replace(hour=reset_hour, minute=0, second=0, microsecond=0)
    if next_war <= now:
        next_war += timedelta(days=7)
    is_war_day_now = now.weekday() in (3, 4, 5, 6)  # Thu-Sun
    return jsonify({"next_war_starts_at": next_war.isoformat(), "is_war_day_now": is_war_day_now})


# ---------------------------------------------------------------------------
# 7. GAMIFICATION & ENGAGEMENT (250-ideas section 6, items 101-115)
# ---------------------------------------------------------------------------

# Idea #101: custom, clan-culture badges — not just Supercell's own achievement
# badges. Each entry is (id, label, emoji, check(profile, battles) -> bool).
# Kept as simple functions over already-logged battle_history/player_profiles
# data — no new CR API calls, so this is safe to compute on every page load.
def _badge_comeback_king(profile, battles):
    # NOTE: the flattened battle_history doesn't retain round-by-round crown
    # timing, so "came back from behind" is approximated as a win where the
    # opponent still landed at least 2 crowns — the closest signal available
    # (documented approximation, same pattern as this project's other ones).
    return sum(1 for b in battles if b.get("result") == "win" and b.get("opponent_crowns", 0) >= 2) >= 3

def _badge_iron_wall(profile, battles):
    return sum(1 for b in battles if b.get("result") == "win" and b.get("opponent_crowns", 0) == 0) >= 5

def _badge_marathon_runner(profile, battles):
    return len(battles) >= 500

def _badge_streak_master(profile, battles):
    return (profile.get("war_participation_streak") or 0) >= 5

def _badge_century_club(profile, battles):
    return (profile.get("wins") or 0) >= 100

# Each entry's description explains exactly how that badge is earned, so
# api_player_badges can hand it straight to the frontend as tooltip text
# instead of leaving players to guess what "Iron Wall" is supposed to mean.
BADGE_DEFINITIONS = [
    ("comeback_king",    "Comeback King",     "🔥", _badge_comeback_king,
     "Won 3+ battles where the opponent still landed 2 or more crowns — a close win, not a clean sweep."),
    ("iron_wall",        "Iron Wall",         "🛡️", _badge_iron_wall,
     "Won 5+ battles as a shutout — opponent finished with 0 crowns."),
    ("marathon_runner",  "Marathon Runner",   "🏃", _badge_marathon_runner,
     "Logged 500+ total battles since we started tracking this player."),
    ("streak_master",    "Streak Master",     "⚡", _badge_streak_master,
     "Currently on a 5+ week war-participation streak (used all 4 war-day battles, week after week)."),
    ("century_club",     "Century Club",      "💯", _badge_century_club,
     "100+ lifetime wins, per Clash Royale's own win counter."),
]


@web_bp.route("/api/player/<tag>/badges")
def api_player_badges(tag):
    """Idea #101 (clan-culture achievement badges) + #113 (comeback recognition,
    folded into the comeback_king badge above rather than a separate system —
    they're the same underlying signal)."""
    clean = clean_tag(tag)
    profile = db_sync["player_profiles"].find_one({"tag": f"#{clean}"}, {"wins": 1, "war_participation_streak": 1}) or {}
    battles = list(db_sync["battle_history"].find(
        {"player_tag": clean}, {"result": 1, "opponent_crowns": 1}
    ).limit(1000))
    earned = [
        {"id": bid, "label": label, "emoji": emoji, "description": description}
        for bid, label, emoji, check, description in BADGE_DEFINITIONS
        if check(profile, battles)
    ]
    return jsonify({"badges": earned})


# Idea #109: seasonal challenges. Defined here as simple, code-level rules
# rather than a full admin-editable challenge builder (out of scope for this
# pass) — each has an id, description, weekly target, and a counter function
# over this week's battle_history.
def _challenge_three_crown_wins(battles):
    return sum(1 for b in battles if b.get("result") == "win" and b.get("team_crowns", 0) == 3)

def _challenge_shutout_wins(battles):
    return sum(1 for b in battles if b.get("result") == "win" and b.get("opponent_crowns", 0) == 0)

SEASONAL_CHALLENGES = [
    {"id": "three_crown_5x", "desc": "Get a 3-crown win 5 times this week", "target": 5, "counter": _challenge_three_crown_wins},
    {"id": "shutout_3x",     "desc": "Win without conceding a crown, 3 times this week", "target": 3, "counter": _challenge_shutout_wins},
]


@web_bp.route("/api/player/<tag>/challenges")
def api_player_challenges(tag):
    """Idea #109: weekly seasonal challenges with in-dashboard progress tracking."""
    clean = clean_tag(tag)
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y%m%d")
    battles = list(db_sync["battle_history"].find(
        {"player_tag": clean, "battle_time": {"$gte": week_ago}}, {"result": 1, "team_crowns": 1, "opponent_crowns": 1}
    ))
    result = [
        {"id": c["id"], "desc": c["desc"], "progress": min(c["counter"](battles), c["target"]), "target": c["target"],
         "complete": c["counter"](battles) >= c["target"]}
        for c in SEASONAL_CHALLENGES
    ]
    return jsonify({"challenges": result})


@web_bp.route("/api/clan/spotlights")
def api_clan_spotlights():
    """10 community/improvement-flavored superlatives (War MVP, Rising Star,
    Most Helpful, Most Improved, etc.) — computed by data_harvester.py's
    compute_weekly_spotlights() every harvest cycle. Each category
    independently falls back to its last known non-zero leader when nobody
    qualifies this cycle (see _merge_leaderboard_entry) — a "stale": true flag
    on a category means it's showing that last-known result, not this cycle's."""
    doc = db_sync["config"].find_one({"_id": "weekly_spotlights"}, {"_id": 0}) or {}
    return jsonify(doc)


@web_bp.route("/api/clan/hall-of-fame")
def api_clan_hall_of_fame():
    """10 quantitative "who's on top right now" superlatives (donations,
    trophies, win rate, streaks, etc.) — computed by data_harvester.py's
    compute_weekly_hall_of_fame() every harvest cycle. Same
    {computed_at, categories: {key: {tag, name, value, value_label, stale, as_of}}}
    shape and zero-fallback behavior as /api/clan/spotlights above."""
    doc = db_sync["config"].find_one({"_id": "weekly_hall_of_fame"}, {"_id": 0}) or {}
    return jsonify(doc)


@web_bp.route("/api/clan/legends")
def api_clan_legends():
    """Idea #110: all-time clan 'legends' record book, separate from the
    current-week Hall of Fame — computed by data_harvester.py's compute_clan_legends()."""
    doc = db_sync["config"].find_one({"_id": "clan_legends"}, {"_id": 0}) or {}
    return jsonify(doc)


@web_bp.route("/api/clan/members-lite")
def api_clan_members_lite():
    """Minimal public {tag, name} list of current clan members — backs the
    tag-autofill/autocomplete on the "Compare With Another Member" box on
    player.html (idea: "autofill based on our players in the text boxes").
    Deliberately just tag+name, nothing else: this is meant to populate an
    HTML <datalist>, not double as a stats API (that's what /api/v1/clan is
    for). Public/no-auth since a member's tag and name are already visible
    to any logged-out visitor on the roster page itself; relies on
    fetch_cr_api()'s own cache rather than adding a separate rate limit."""
    clan_data, _ = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    return jsonify([
        {"tag": m.get("tag", "").replace("#", ""), "name": m.get("name", "")}
        for m in clan_data.get("memberList", [])
        if m.get("tag")
    ])


DISCORD_WIDGET_CACHE_TTL_SECONDS = 60  # same order of magnitude as CR_API_CACHE_TTL_SECONDS

@web_bp.route("/api/discord/widget")
def api_discord_widget():
    """Backs the custom-built variant of the "Our Discord" roster card
    (bot_settings.discord_widget_style == "custom") -- the alternative to
    Discord's own official embeddable iframe, for admins who'd rather match
    the site's own dark theme/typography exactly than use Discord's stock
    widget chrome. Calls Discord's public, unauthenticated
    `GET /guilds/{id}/widget.json`, which only returns data if "Server
    Widget" is switched on in Discord's own Server Settings -- same
    prerequisite the official iframe option depends on. Degrades gracefully
    to {"enabled": false, "invite_url": ...} rather than erroring if that
    setting is off or the request fails for any reason. Cached briefly in
    Redis (same non-fatal-on-Redis-down pattern as fetch_cr_api) so this
    doesn't hit Discord on every single roster page visit."""
    invite_url = (db_sync["config"].find_one({"_id": "bot_settings"}) or {}).get("discord_invite_url") or None
    if not GUILD_ID:
        return jsonify({"enabled": False, "invite_url": invite_url})

    cache_key = f"discord_widget:{GUILD_ID}"
    try:
        cached = redis_sync_client.get(cache_key)
        if cached:
            return jsonify(json.loads(cached))
    except Exception as e:
        log.warning(f"Discord widget cache read failed (non-fatal): {e}")

    try:
        r = requests.get(f"{DISCORD_API}/guilds/{GUILD_ID}/widget.json", timeout=8)
        if r.status_code != 200:
            result = {"enabled": False, "invite_url": invite_url}
        else:
            d = r.json()
            result = {
                "enabled": True,
                "name": d.get("name"),
                "presence_count": d.get("presence_count", 0),
                "instant_invite": d.get("instant_invite") or invite_url,
                # Cap the member list -- widget.json can return a large roster of
                # everyone currently online, and this card only ever shows a
                # handful as a "who's around" flavor, not a full member browser.
                "members": [
                    {"name": m.get("username"), "avatar_url": m.get("avatar_url"), "status": m.get("status")}
                    for m in (d.get("members") or [])[:12]
                ],
            }
    except Exception as e:
        log.warning(f"Discord widget fetch failed (non-fatal, degrading to disabled state): {e}")
        result = {"enabled": False, "invite_url": invite_url}

    try:
        redis_sync_client.setex(cache_key, DISCORD_WIDGET_CACHE_TTL_SECONDS, json.dumps(result))
    except Exception as e:
        log.warning(f"Discord widget cache write failed (non-fatal): {e}")
    return jsonify(result)


@web_bp.route("/api/clan/progress")
def api_clan_progress():
    """Idea #106: clan-wide progress bar toward a shared monthly donation goal,
    for the roster page. Goal is configurable via bot_settings.monthly_donation_goal
    (default 50,000, a reasonable round number for a ~50-member clan)."""
    clan_data, clan_data_is_stale = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    total_donations = sum(m.get("donations", 0) for m in clan_data.get("memberList", []))
    goal = int((db_sync["config"].find_one({"_id": "bot_settings"}) or {}).get("monthly_donation_goal", 50000))
    return jsonify({
        "total_donations": total_donations, "goal": goal,
        "pct": round(min(total_donations / goal, 1.0) * 100, 1) if goal else 0,
        "is_stale": clan_data_is_stale,
    })


@web_bp.route("/api/player/<tag>/rivalry")
def api_player_rivalry(tag):
    """Idea #115: this member's currently-assigned friendly rivalry (re-paired
    roughly weekly by data_harvester.py's assign_weekly_rivalries()), with a
    head-to-head record computed on read from logged battle_history."""
    clean = clean_tag(tag)
    pair = db_sync["rivalries"].find_one(
        {"active": True, "$or": [{"tag_a": f"#{clean}"}, {"tag_b": f"#{clean}"}]}, {"_id": 0}
    )
    if not pair:
        return jsonify({"rivalry": None})
    opponent_tag = pair["tag_b"] if pair["tag_a"] == f"#{clean}" else pair["tag_a"]
    opponent_name = pair["name_b"] if pair["tag_a"] == f"#{clean}" else pair["name_a"]
    h2h = list(db_sync["battle_history"].find(
        {"player_tag": clean, "opponent_tag": opponent_tag.replace("#", "")}, {"result": 1}
    ))
    record = {"wins": sum(1 for b in h2h if b.get("result") == "win"),
              "losses": sum(1 for b in h2h if b.get("result") == "loss")}
    return jsonify({"rivalry": {"opponent_tag": opponent_tag.replace("#", ""), "opponent_name": opponent_name, "head_to_head": record}})


@web_bp.route("/api/leaderboards/public")
def api_public_leaderboards():
    """Idea #108: opt-in public leaderboards (trophies/win-rate/war fame), unlike
    the admin-only analytics leaderboard — members who've set leaderboard_optout
    are excluded entirely rather than just anonymized."""
    optout_tags = {p["tag"] for p in db_sync["player_profiles"].find({"leaderboard_optout": True}, {"tag": 1})}
    clan_data, clan_data_is_stale = fetch_cr_api_with_fallback(f"clans/%23{CLAN_TAG}")
    clan_data = clan_data or {}
    members = [m for m in clan_data.get("memberList", []) if m.get("tag") not in optout_tags]
    top_trophies = sorted(members, key=lambda m: m.get("trophies", 0), reverse=True)[:10]

    war_data = fetch_cr_api(f"clans/%23{CLAN_TAG}/currentriverrace") or {}
    fame_rows = [p for p in war_data.get("clan", {}).get("participants", []) if p.get("tag") not in optout_tags]
    fame_is_last_race = False
    if not any(p.get("fame", 0) for p in fame_rows):
        # Fame resets to 0 for everyone between war days -- most of the week,
        # this river race response is genuinely all zeroes, not broken. Rather
        # than show a wall of 0s, fall back to the most recently *completed*
        # race, same data already used for the Boat Battle Leaderboard.
        last_race = db_sync["war_history"].find_one(
            {}, sort=[("data.seasonId", -1), ("data.sectionIndex", -1)]
        )
        if last_race:
            race_clan = (last_race.get("data") or {}).get("clan") or {}
            race_participants = [p for p in race_clan.get("participants", []) if p.get("tag") not in optout_tags]
            if any(p.get("fame", 0) for p in race_participants):
                fame_rows = race_participants
                fame_is_last_race = True

    top_fame = sorted(fame_rows, key=lambda p: p.get("fame", 0), reverse=True)[:10]
    return jsonify({
        "top_trophies": [{"name": m.get("name"), "value": m.get("trophies", 0)} for m in top_trophies],
        "top_trophies_is_stale": clan_data_is_stale,
        "top_fame": [{"name": p.get("name"), "value": p.get("fame", 0)} for p in top_fame],
        "top_fame_is_last_race": fame_is_last_race,
    })


@web_bp.route("/api/player/<tag>/leaderboard-optout", methods=["GET", "POST"])
def api_player_leaderboard_optout(tag):
    """Idea #108: self-service (or admin) toggle for opting out of the public leaderboards."""
    clean = clean_tag(tag)
    if request.method == "GET":
        doc = db_sync["player_profiles"].find_one({"tag": f"#{clean}"}, {"leaderboard_optout": 1}) or {}
        return jsonify({"optout": bool(doc.get("leaderboard_optout", False))})
    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    user = db_sync["users"].find_one({"discord_id": session["discord_id"]})
    if not (user and clean_tag(user.get("cr_tag", "")) == clean) and not is_admin():
        return jsonify({"error": "You can only change your own leaderboard visibility."}), 403
    data = request.get_json(silent=True) or {}
    db_sync["player_profiles"].update_one({"tag": f"#{clean}"}, {"$set": {"leaderboard_optout": bool(data.get("optout"))}}, upsert=True)
    return jsonify({"success": True})


@web_bp.route("/api/player/<tag>/points")
def api_player_points(tag):
    """Idea #104: clan points/currency balance, earned via war participation
    (data_harvester.py's _update_war_streaks_and_points)."""
    clean = clean_tag(tag)
    doc = db_sync["player_profiles"].find_one({"tag": f"#{clean}"}, {"clan_points": 1, "war_participation_streak": 1, "streak_shields": 1}) or {}
    return jsonify({
        "clan_points": doc.get("clan_points", 0),
        "war_participation_streak": doc.get("war_participation_streak", 0),
        "streak_shields": doc.get("streak_shields", 0),
    })


# Idea #104/#112: a tiny "shop" — spend banked clan points on cosmetic profile
# perks. Kept intentionally small (one perk) rather than a full storefront,
# since the underlying currency/points plumbing is the part with real reuse value.
SHOP_BADGE_COST = 100

@web_bp.route("/api/player/<tag>/shop/redeem", methods=["POST"])
def api_player_shop_redeem(tag):
    """Idea #112: spend clan points to set a custom profile badge (shown next to
    the flair motto on the player page). Self-service only — an admin can't
    spend a member's points on their behalf."""
    clean = clean_tag(tag)
    if "discord_id" not in session:
        return jsonify({"error": "unauthorized"}), 403
    user = db_sync["users"].find_one({"discord_id": session["discord_id"]})
    if not (user and clean_tag(user.get("cr_tag", "")) == clean):
        return jsonify({"error": "You can only redeem points on your own profile."}), 403
    data = request.get_json(silent=True) or {}
    badge_text = str(data.get("badge_text", "")).strip()[:24]
    if not badge_text:
        return jsonify({"error": "badge_text is required"}), 400
    profile = db_sync["player_profiles"].find_one({"tag": f"#{clean}"}, {"clan_points": 1}) or {}
    if (profile.get("clan_points", 0) or 0) < SHOP_BADGE_COST:
        return jsonify({"error": f"Not enough clan points — need {SHOP_BADGE_COST}, have {profile.get('clan_points', 0)}."}), 400
    db_sync["player_profiles"].update_one(
        {"tag": f"#{clean}"},
        {"$inc": {"clan_points": -SHOP_BADGE_COST}, "$set": {"custom_badge": badge_text}},
    )
    return jsonify({"success": True, "custom_badge": badge_text})


@web_bp.route("/api/player/<tag>/season-recap")
def api_player_season_recap(tag):
    """Idea #114: a Spotify-Wrapped-style end-of-season recap, built entirely
    from data this project already computes elsewhere (get_player_analytical_data)
    rather than a second parallel stats pipeline. Available on demand at any
    time — not gated to only display right at season end — since a member might
    want to check progress mid-season too."""
    player = get_player_analytical_data(tag)
    if not player:
        return jsonify({"error": "player not found"}), 404
    # war_participation_streak/clan_points live only in the Mongo player_profiles
    # doc (idea #103/#104's gamification fields), not in get_player_analytical_data's
    # live-CR-API-rooted dict, so they're fetched separately here.
    db_profile = db_sync["player_profiles"].find_one(
        {"tag": f"#{clean_tag(tag)}"}, {"war_participation_streak": 1, "clan_points": 1}
    ) or {}
    return jsonify({
        "name": player.get("name"),
        "season_trophies": player.get("trophies"),
        "personal_best_trophies_tracked": player.get("personal_best_trophies_tracked"),
        "wins": player.get("wins"), "losses": player.get("losses"),
        "three_crown_wins": player.get("threeCrownWins"),
        "longest_win_streak": player.get("personal_best_win_streak"),
        "war_participation_streak": db_profile.get("war_participation_streak", 0),
        "clan_points": db_profile.get("clan_points", 0),
        "deck_grade": player.get("deck_grade"),
    })


# ============================================================================
# 250-ideas implementation pass — Section 12: Social & community (191-205)
# ============================================================================

def _current_linked_user():
    """Shared helper for section 12+: the `users` doc for whoever is logged in
    via Discord OAuth, or None if not logged in / no session at all. Several
    social features below are open to any linked member, not just admins —
    callers still need to separately check `.get("cr_tag")` for "actually
    linked" vs. just "has a Discord session"."""
    if "discord_id" not in session:
        return None
    return db_sync["users"].find_one({"discord_id": session["discord_id"]})


@web_bp.route("/api/player/<tag>/comments", methods=["GET", "POST"])
def api_player_comments(tag):
    """Idea #191: lightweight comments/reactions on a player's profile —
    teammates congratulating a good war day, separate from admin notes (which
    stay private to leadership). GET is public (same visibility as the profile
    itself); POST requires a linked account so comments aren't anonymous."""
    clean = clean_tag(tag)
    if request.method == "GET":
        comments = list(db_sync["profile_comments"].find(
            {"tag": f"#{clean}"}, {"_id": 0}
        ).sort("created_at", -1).limit(50))
        return jsonify({"comments": comments})
    user = _current_linked_user()
    if not user or not user.get("cr_tag"):
        return jsonify({"error": "Link your account to comment."}), 403
    data = request.get_json(silent=True) or {}
    message = str(data.get("message", "")).strip()[:280]
    reaction = str(data.get("reaction", "")).strip()[:8]  # a single emoji, optional
    if not message and not reaction:
        return jsonify({"error": "Say something or react with an emoji first."}), 400
    db_sync["profile_comments"].insert_one({
        "tag": f"#{clean}", "from_discord_id": user["discord_id"],
        "from_name": session.get("discord_name", "A teammate"),
        "message": message, "reaction": reaction,
        "created_at": datetime.now(timezone.utc),
    })
    return jsonify({"success": True})


@web_bp.route("/api/gallery", methods=["GET", "POST"])
def api_gallery():
    """Idea #194: a lightweight shared clan photo/meme board. Kept simple —
    members paste an image URL rather than this app hosting file uploads, so
    there's no new storage/CDN dependency. Any linked member can post; an
    admin can remove anything via the DELETE route below.

    Idea #250: also the concrete example wired up for the opt-in "beta
    features" toggle — the gallery board stays off until an admin flips
    `bot_settings.beta_features_enabled` on, so leadership can trial it
    before rolling it out clan-wide."""
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    if not bot_settings.get("beta_features_enabled") and not is_admin():
        return jsonify({"enabled": False, "items": []})
    if request.method == "GET":
        items = list(db_sync["clan_gallery"].find({}, {"_id": 0, "gallery_id": 1, "url": 1, "caption": 1, "posted_by_name": 1, "created_at": 1}).sort("created_at", -1).limit(60))
        return jsonify({"items": items})
    user = _current_linked_user()
    if not user or not user.get("cr_tag"):
        return jsonify({"error": "Link your account to post."}), 403
    data = request.get_json(silent=True) or {}
    url = str(data.get("url", "")).strip()
    if not url.startswith(("http://", "https://")):
        return jsonify({"error": "Please provide a valid image URL."}), 400
    gallery_id = secrets.token_hex(8)
    db_sync["clan_gallery"].insert_one({
        "gallery_id": gallery_id, "url": url[:500],
        "caption": str(data.get("caption", "")).strip()[:200],
        "posted_by_discord_id": user["discord_id"], "posted_by_name": session.get("discord_name", "Member"),
        "created_at": datetime.now(timezone.utc),
    })
    return jsonify({"success": True, "gallery_id": gallery_id})


@web_bp.route("/api/gallery/<gallery_id>", methods=["DELETE"])
def api_gallery_delete(gallery_id):
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    db_sync["clan_gallery"].delete_one({"gallery_id": gallery_id})
    return jsonify({"success": True})


@web_bp.route("/api/polls", methods=["GET", "POST"])
def api_polls():
    """Idea #195: a poll/voting widget for clan decisions embedded on the
    dashboard. POST (create) is admin-only — this is for leadership to put a
    decision to a vote, not an open-floor voting board; voting itself (see
    /api/polls/<id>/vote below) is open to any linked member."""
    if request.method == "GET":
        active_only = request.args.get("active_only", "1") != "0"
        query = {"active": True} if active_only else {}
        polls = list(db_sync["polls"].find(query, {"voters": 0}).sort("created_at", -1).limit(10))
        for p in polls:
            p["_id"] = str(p["_id"])
        return jsonify({"polls": polls})
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    question = str(data.get("question", "")).strip()[:200]
    options = [str(o).strip()[:80] for o in (data.get("options") or []) if str(o).strip()]
    if not question or len(options) < 2:
        return jsonify({"error": "A poll needs a question and at least 2 options."}), 400
    result = db_sync["polls"].insert_one({
        "question": question,
        "options": [{"text": o, "votes": 0} for o in options],
        "created_by": session.get("discord_name", "Admin"),
        "created_at": datetime.now(timezone.utc),
        "active": True, "voters": [],
    })
    return jsonify({"success": True, "poll_id": str(result.inserted_id)})


@web_bp.route("/api/polls/<poll_id>/vote", methods=["POST"])
def api_poll_vote(poll_id):
    from bson import ObjectId
    from bson.errors import InvalidId
    user = _current_linked_user()
    if not user or not user.get("cr_tag"):
        return jsonify({"error": "Link your account to vote."}), 403
    data = request.get_json(silent=True) or {}
    try:
        option_index = int(data.get("option_index"))
        oid = ObjectId(poll_id)
    except (TypeError, ValueError, InvalidId):
        return jsonify({"error": "Invalid poll or option."}), 400
    poll = db_sync["polls"].find_one({"_id": oid})
    if not poll or not poll.get("active"):
        return jsonify({"error": "That poll is no longer active."}), 404
    if user["discord_id"] in (poll.get("voters") or []):
        return jsonify({"error": "You've already voted in this poll."}), 400
    if not (0 <= option_index < len(poll.get("options", []))):
        return jsonify({"error": "Invalid option."}), 400
    db_sync["polls"].update_one(
        {"_id": oid},
        {"$inc": {f"options.{option_index}.votes": 1}, "$addToSet": {"voters": user["discord_id"]}},
    )
    return jsonify({"success": True})


@web_bp.route("/admin/api/polls/<poll_id>/close", methods=["POST"])
def admin_close_poll(poll_id):
    from bson import ObjectId
    from bson.errors import InvalidId
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    try:
        oid = ObjectId(poll_id)
    except InvalidId:
        return jsonify({"error": "invalid poll id"}), 400
    db_sync["polls"].update_one({"_id": oid}, {"$set": {"active": False}})
    return jsonify({"success": True})


@web_bp.route("/api/kudos", methods=["GET", "POST"])
def api_kudos():
    """Idea #197: a lightweight "thank a teammate" kudos system, separate from
    the formal strikes/admin-notes system — positive peer recognition rather
    than anything leadership-gated. GET with ?tag=... returns kudos received
    by that player; POST sends one."""
    if request.method == "GET":
        tag = clean_tag(request.args.get("tag", ""))
        if not tag:
            return jsonify({"error": "tag query param required"}), 400
        items = list(db_sync["kudos"].find({"to_tag": f"#{tag}"}, {"_id": 0}).sort("created_at", -1).limit(20))
        return jsonify({"kudos": items, "count": len(items)})
    user = _current_linked_user()
    if not user or not user.get("cr_tag"):
        return jsonify({"error": "Link your account to send kudos."}), 403
    data = request.get_json(silent=True) or {}
    to_tag = clean_tag(data.get("to_tag", ""))
    message = str(data.get("message", "")).strip()[:200]
    if not to_tag or not message:
        return jsonify({"error": "to_tag and message are required"}), 400
    if to_tag == clean_tag(user.get("cr_tag", "")):
        return jsonify({"error": "You can't send kudos to yourself."}), 400
    db_sync["kudos"].insert_one({
        "from_discord_id": user["discord_id"], "from_name": session.get("discord_name", "A teammate"),
        "to_tag": f"#{to_tag}", "message": message, "created_at": datetime.now(timezone.utc),
    })
    return jsonify({"success": True})


@web_bp.route("/api/feature-suggestion", methods=["POST"])
def api_feature_suggestion():
    """Idea #201: a "suggest a feature" box embedded in the dashboard footer.
    Stored in its own small collection (reviewed via the admin panel's
    Recruiting tab alongside applications) rather than the pending_actions
    queue, since this isn't an action that needs dispatching anywhere — it's
    just a durable inbox for leadership to read later."""
    user = _current_linked_user()
    message = str((request.get_json(silent=True) or {}).get("message", "")).strip()[:500]
    if not message:
        return jsonify({"error": "Please write a suggestion first."}), 400
    db_sync["feature_suggestions"].insert_one({
        "message": message,
        "from_discord_id": (user or {}).get("discord_id"),
        "from_name": session.get("discord_name", "Anonymous") if user else "Anonymous",
        "created_at": datetime.now(timezone.utc), "status": "new",
    })
    return jsonify({"success": True})


@web_bp.route("/admin/api/feature-suggestions")
def admin_feature_suggestions():
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    items = list(db_sync["feature_suggestions"].find({}, {"_id": 0}).sort("created_at", -1).limit(100))
    return jsonify({"suggestions": items})


@web_bp.route("/admin/api/feature-suggestions/status", methods=["POST"])
def admin_feature_suggestion_status():
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    db_sync["feature_suggestions"].update_one(
        {"created_at": data.get("created_at")},  # created_at doubles as a stable-enough key since there's no exposed _id
        {"$set": {"status": str(data.get("status", "reviewed"))[:20]}},
    )
    return jsonify({"success": True})


@web_bp.route("/admin/api/scrims", methods=["GET", "POST"])
def admin_scrims():
    """Idea #202: cross-clan friendly-battle results tracker, for clans that
    regularly scrim other clans. Admin-managed — there's no CR API endpoint
    for "friendly battle vs. another clan", so this is manually logged."""
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    if request.method == "GET":
        items = list(db_sync["scrim_results"].find({}, {"_id": 0}).sort("date", -1).limit(50))
        return jsonify({"scrims": items})
    data = request.get_json(silent=True) or {}
    opponent = str(data.get("opponent_clan_name", "")).strip()[:60]
    if not opponent:
        return jsonify({"error": "opponent_clan_name is required"}), 400
    db_sync["scrim_results"].insert_one({
        "opponent_clan_name": opponent,
        "opponent_clan_tag": clean_tag(data.get("opponent_clan_tag", "")),
        "date": str(data.get("date", "")).strip()[:10] or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "result": str(data.get("result", "")).strip()[:10],  # "win"/"loss"/"draw"
        "our_score": data.get("our_score"), "their_score": data.get("their_score"),
        "notes": str(data.get("notes", "")).strip()[:300],
        "created_at": datetime.now(timezone.utc),
    })
    return jsonify({"success": True})


@web_bp.route("/admin/api/spotlight", methods=["GET", "POST"])
def admin_member_spotlight():
    """Idea #193: a monthly member spotlight/interview leadership can publish —
    a short bio/Q&A block for a chosen member, shown on the clan history page."""
    if request.method == "GET":
        spotlight = db_sync["config"].find_one({"_id": "member_spotlight"}, {"_id": 0}) or {}
        return jsonify(spotlight)
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    doc = {
        "tag": clean_tag(data.get("tag", "")),
        "name": str(data.get("name", "")).strip()[:60],
        "bio": str(data.get("bio", "")).strip()[:1000],
        "month": str(data.get("month", "")).strip()[:20],
        "updated_at": datetime.now(timezone.utc),
    }
    db_sync["config"].update_one({"_id": "member_spotlight"}, {"$set": doc}, upsert=True)
    return jsonify({"success": True})


@web_bp.route("/admin/api/milestones", methods=["GET", "POST"])
def admin_clan_milestones():
    """Ideas #192/#196: manually-curated historical clan milestones (biggest
    war win, longest streak, founding-era events, ...) feeding the public
    /clan-history timeline, on top of the auto-computed clan_legends record
    book (idea #110) which only tracks a few fixed record categories."""
    if request.method == "GET":
        items = list(db_sync["clan_milestones"].find({}, {"_id": 0}).sort("date", 1).limit(200))
        return jsonify({"milestones": items})
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    title = str(data.get("title", "")).strip()[:120]
    if not title:
        return jsonify({"error": "title is required"}), 400
    db_sync["clan_milestones"].insert_one({
        "title": title,
        "description": str(data.get("description", "")).strip()[:400],
        "date": str(data.get("date", "")).strip()[:10] or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "created_by": session.get("discord_name", "Admin"),
        "created_at": datetime.now(timezone.utc),
    })
    return jsonify({"success": True})


@web_bp.route("/clan-history")
def clan_history_page():
    """Ideas #192 (wall of fame), #193 (member spotlight), #196 (public
    timeline), #203 (founding-day info), #204 (clan culture doc) — consolidated
    into one page rather than five separate near-empty ones."""
    legends = db_sync["config"].find_one({"_id": "clan_legends"}, {"_id": 0}) or {}
    spotlight = db_sync["config"].find_one({"_id": "member_spotlight"}, {"_id": 0}) or {}
    milestones = list(db_sync["clan_milestones"].find({}, {"_id": 0}).sort("date", 1).limit(200))
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    return render_sandboxed(
        get_template("clan_history"),
        legends=legends, spotlight=spotlight, milestones=milestones,
        culture_content=bot_settings.get("culture_page_content", ""),
        founding_date=bot_settings.get("clan_founding_date", ""),
        is_admin=is_admin(),
    )


@web_bp.route("/family")
def family_page():
    """Idea #198 — per your note, built but hidden until you actually need it:
    the nav link/route stay dark unless `bot_settings.family_page_enabled` is
    explicitly turned on in Settings. When enabled, aggregates whichever
    linked clan tags are configured in `bot_settings.family_clan_tags` into
    one combined roster/leaderboard view."""
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    if not bot_settings.get("family_page_enabled"):
        return "The clan family page isn't enabled yet — an admin can turn it on in Settings once there's a second linked clan.", 404
    family_tags = bot_settings.get("family_clan_tags", []) or []
    clans = []
    any_stale = False
    for t in family_tags:
        # fetch_cr_api_with_fallback caches per-endpoint, so each family clan
        # tag gets its own independent last-known snapshot -- one clan's live
        # fetch failing no longer silently drops it from the page entirely.
        data, is_stale = fetch_cr_api_with_fallback(f"clans/%23{clean_tag(t)}")
        if data:
            clans.append(data)
            any_stale = any_stale or is_stale
    return render_sandboxed(get_template("family"), clans=clans, is_admin=is_admin(), clan_data_is_stale=any_stale)


@web_bp.route("/alumni")
def alumni_page():
    """Idea #199: alumni page for members who left on good terms, preserving
    their historical stats. Reuses the `left_clan_at` tracking already added
    for the 23-week departed-member retention purge — this just surfaces
    anyone still in that window (and not explicitly flagged
    `departed_on_good_terms: false` by an admin) as a friendly "where are they
    now" list instead of only ever being invisible until purge."""
    alumni = list(db_sync["player_profiles"].find(
        {"left_clan_at": {"$ne": None}, "departed_on_good_terms": {"$ne": False}},
        {"_id": 0, "tag": 1, "name": 1, "left_clan_at": 1, "joined_clan_at": 1, "bestTrophies": 1, "clan_points": 1},
    ).sort("left_clan_at", -1).limit(100))
    return render_sandboxed(get_template("alumni"), alumni=alumni, is_admin=is_admin())


@web_bp.route("/admin/api/player/<tag>/departed-status", methods=["POST"])
def admin_player_departed_status(tag):
    """Companion to the alumni page: lets an admin exclude a specific departed
    member from the "good terms" alumni list without touching their retained
    stats otherwise."""
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    db_sync["player_profiles"].update_one(
        {"tag": f"#{clean_tag(tag)}"},
        {"$set": {"departed_on_good_terms": bool(data.get("good_terms", True))}},
    )
    return jsonify({"success": True})


@web_bp.route("/api/testimonials", methods=["GET", "POST"])
def api_testimonials():
    """Idea #205 — per your note, hidden until enabled: GET only returns
    approved testimonials, and only if `bot_settings.testimonials_enabled` is
    on. POST lets any linked member submit one for admin approval (see
    /admin/api/testimonials/<id>/approve)."""
    bot_settings = db_sync["config"].find_one({"_id": "bot_settings"}) or {}
    if request.method == "GET":
        if not bot_settings.get("testimonials_enabled"):
            return jsonify({"enabled": False, "testimonials": []})
        items = list(db_sync["testimonials"].find({"approved": True}, {"_id": 0}).limit(20))
        return jsonify({"enabled": True, "testimonials": items})
    user = _current_linked_user()
    if not user or not user.get("cr_tag"):
        return jsonify({"error": "Link your account to submit a testimonial."}), 403
    data = request.get_json(silent=True) or {}
    quote = str(data.get("quote", "")).strip()[:300]
    if not quote:
        return jsonify({"error": "quote is required"}), 400
    db_sync["testimonials"].insert_one({
        "tag": user.get("cr_tag"), "name": session.get("discord_name", "Member"),
        "quote": quote, "approved": False, "created_at": datetime.now(timezone.utc),
    })
    return jsonify({"success": True, "note": "Submitted for admin approval."})


@web_bp.route("/admin/api/testimonials")
def admin_list_testimonials():
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    items = list(db_sync["testimonials"].find({}, {"_id": 0}).sort("created_at", -1).limit(100))
    return jsonify({"testimonials": items})


@web_bp.route("/admin/api/testimonials/approve", methods=["POST"])
def admin_approve_testimonial():
    if not is_admin():
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    db_sync["testimonials"].update_one(
        {"tag": data.get("tag"), "created_at": data.get("created_at")},
        {"$set": {"approved": bool(data.get("approved", True))}},
    )
    return jsonify({"success": True})